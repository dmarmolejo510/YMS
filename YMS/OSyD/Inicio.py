from flask import request,session,render_template,current_app
from cryptography.fernet import Fernet
from datetime import datetime,date,timedelta
import sys
import json
import os
from Componentes import LibDM_2023
import csv
import openpyxl
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill,colors
BD_Nombre = "public"
Bandera_Dock = "YMS"
fernet = Fernet(LibDM_2023.Compartido().Dame_K2())


def Inicio():
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    Cur = ""
    try:
        Activo = "YMS"
        Compartido = LibDM_2023.Compartido()
        Menu = LibDM_2023.Menu().Menu(Activo,request.url_root,session["IDu"])
        Titulo = LibDM_2023.Menu().Get_Titulo(Activo)
        Contenido = ""

        Contenido += """<div class='text-end pe-1 pt-1'><small class='link-primary' style='cursor:pointer' onclick='Llamar_Funcion(\""""+str(request.url)+"""\");'>Actualizar <i class='mdi mdi-refresh'></i></small></div>"""
        Contenido += "<div class='h2 fw-lighter mt-1 mb-1 text-center border-bottom'><i class='mdi mdi-alert'></i> OS&D</div>"
        Contenido += """
        <div class='container-fluid'>
             <ul class="nav nav-tabs nav-fill">
                <li class="nav-item">
                    <a id='Tab_Abierto' class="nav-link active bg-primary fw-bold text-light" href="#" onclick='Cargar_Abiertos()'>OS&D Abiertos <span class="badge text-bg-secondary"></span></a>
                </li>
                <li class="nav-item">
                    <a id='Tab_Liberados' class="nav-link" href="#" onclick='Cargar_Liberados()'>OS&D Liberados <span class="badge text-bg-secondary"></span></a>
                </li>
                <li class="nav-item">
                    <a id='Tab_Reporte' class="nav-link" href="#" onclick='Cargar_Menu_Reporte()'>OS&D Reporte</a>
                </li>
            </ul>
            <div id='Res'></div>
        </div>
        <style>
            .tabulator .tabulator-row.tabulator-selectable:hover .tabulator-cell{
                background-color: rgba(255, 221, 221, 1)
            }
            .tabulator-row .tabulator-cell[tabulator-field="Ultimo Comentario"] {
                white-space: pre-wrap;
            }
        </style>
        <script>
            $( document ).ready(function() {
               Cargar_Abiertos();
            });
            function Cargar_Abiertos(){
                Mostrar_Ventana_Cargando(false);
                $(".nav-link").removeClass('active bg-primary fw-bold text-light');
                $("#Tab_Abierto").addClass('active bg-primary fw-bold text-light');
                var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Abiertos".encode()).decode("utf-8"))+"""'};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Res").html(Resultado["Contenido"]);
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,'Process error ['+ jqXHR.status + " | " + textStatus  + " | " + errorThrown +']');}
                });
            }
            function Cargar_Liberados(){
                Mostrar_Ventana_Cargando(false);
                $(".nav-link").removeClass('active bg-primary fw-bold text-light')
                $("#Tab_Liberados").addClass('active bg-primary fw-bold text-light')
                $("#Res").html("Loading...");
                var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Liberados".encode()).decode("utf-8"))+"""'};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Res").html(Resultado["Contenido"]);
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ jqXHR.status + " | " + textStatus  + " | " + errorThrown +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                    }
                });
            }
            function Ver_Detalle(IDMaster){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm').addClass('modal-lg')
                $("#Vent_1").modal("show").find(".modal-body").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>").parent().find(".modal-title").html(pad(IDMaster,5));
                var parametros = {"Fun":'"""+str(fernet.encrypt("Ver_Detalle".encode()).decode("utf-8"))+"""',"IDMaster":IDMaster};
                $.ajax({data:  parametros,url:   \""""+str(request.url)+"""\",
                    type:  "post",
                    success:  function (response){ var Resultado = JSON.parse(response); $("#Vent_1").find(".modal-body").html(Resultado["Contenido"]); swal.close();},
                    error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,'Process error ['+ jqXHR.status + " | " + textStatus  + " | " + errorThrown +']');}
                });
            }
            function Generar_Archivo_Ruteo(IDMaster){
                Mostrar_Ventana_Cargando(false);
                var parametros = {"Fun":'"""+str(fernet.encrypt("Generar_Archivo".encode()).decode("utf-8"))+"""',"IDMaster":IDMaster};
                $.ajax({data:  parametros,url:   \""""+str(request.url)+"""\",
                    type:  "post",
                    success:  function (response){ 
                        var Resultado = JSON.parse(response);
                        if(Resultado["Estado"] == 1){
                            swal.close();
                            var win = window.open('"""+str(request.url_root)+"""/Portal_File/Gen/'+Resultado["Archivo"], '_blank');
                            win.focus();
                        }else{
                            Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ Resultado["Contenido"] +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                        }
                    },
                   error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,'Process error ['+ jqXHR.status + " | " + textStatus  + " | " + errorThrown +']');}
                });
            }
            function pad(num, size) {
                num = num.toString();
                while (num.length < size) num = "0" + num;
                return num;
            }
            function Ver_Historico(ID,Folio){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm').addClass('modal-lg')
                $("#Vent_1").modal("show").find(".modal-body").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>").parent().find(".modal-title").html("<i class='mdi mdi-history'></i> History ["+Folio+"]");
                var parametros = {"Fun":'"""+str(fernet.encrypt("Ver_Historico".encode()).decode("utf-8"))+"""',"ID":ID};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    type:  "post",
                    success:  function (response){var Resultado = JSON.parse(response); $("#Vent_1").find(".modal-body").html(Resultado["Contenido"]);swal.close();},
                    error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,'Process error ['+ jqXHR.status + " | " + textStatus  + " | " + errorThrown +']');}
                });
            }
            function Cargar_Menu_Reporte(){
                Mostrar_Ventana_Cargando(false);
                $(".nav-link").removeClass('active bg-primary fw-bold text-light')
                $("#Tab_Reporte").addClass('active bg-primary fw-bold text-light')
                $("#Res").html("Loading...");
                var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Menu_Reporte".encode()).decode("utf-8"))+"""'};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Res").html(Resultado["Contenido"]);
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ jqXHR.status + " | " + textStatus  + " | " + errorThrown +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                    }
                });
            }
        </script>
        """

        Cur += render_template("general.html",Contenido=Contenido,Componentes=Compartido.Complementos(None),Menu=Menu,Titulo=Titulo)
    except:
        Cur += str(sys.exc_info())
    return Cur
def Cargar_Abiertos(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Contador_Abierto = {"valor":0,"clase":"text-bg-secondary"}
        Contador_Liberados = {"valor":0,"clase":"text-bg-secondary"}
        Contador_Pre_Liberados = {"valor":0,"clase":"text-bg-secondary"}
        Resultado["Contenido"] += "<div class='h3 text-center mt-2'>OS&D Abiertos</div>"
        Tabla_Datos_Ruteo = []

        Proveedores = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cproveedores")
        
        Tabla_Datos_Porduccion = []
        Tabla_Datos_Cerrados = []
        for PakingSplip in DB.Get_Dato("""
        select MASTER.*,
        (select HIS.cosyd_fecha from """+str(BD_Nombre)+""".cosyd_historico as HIS where HIS.cosyd_master = MASTER.cosyd_id  order by cosyd_fecha desc limit 1) as "Fecha_Ultimo",
        (select HIS.cosyd_usuario  from """+str(BD_Nombre)+""".cosyd_historico as HIS where HIS.cosyd_master = MASTER.cosyd_id  order by cosyd_fecha desc limit 1) as "Usuario",
        (select HIS.cosyd_movimiento from """+str(BD_Nombre)+""".cosyd_historico as HIS where HIS.cosyd_master = MASTER.cosyd_id  order by cosyd_fecha desc limit 1) as "Tipo",
        (select HIS.cosyd_evidencia  from """+str(BD_Nombre)+""".cosyd_historico as HIS where HIS.cosyd_master = MASTER.cosyd_id  order by cosyd_fecha desc limit 1) as "Evidencia"
        FROM """+str(BD_Nombre)+""".cosyd  MASTER 
        left join """+str(BD_Nombre)+""".cosyd_partes PARTES ON PARTES.cosyd_master = MASTER.cosyd_id
        WHERE MASTER.cosyd_estado IN (1,2,4)
        group by MASTER.cosyd_id
        """):
            if int(PakingSplip["cosyd_estado"]) == 1:
                Contador_Abierto["valor"] += 1
            if int(PakingSplip["cosyd_estado"]) == 2:
                Contador_Liberados["valor"] += 1
            if int(PakingSplip["cosyd_estado"]) == 4:
                Contador_Pre_Liberados["valor"] += 1
            OK = 0
            Numeros = 0
            Problemas = []
            Historico = int(DB.Get_Dato("SELECT COUNT(*) AS \"Numero\" FROM "+str(BD_Nombre)+".cosyd_historico WHERE cosyd_master = '"+str(PakingSplip["cosyd_id"])+"'")[0]["Numero"])
            for Partes in DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cosyd_partes WHERE cosyd_master = '"+str(PakingSplip["cosyd_id"])+"'"):
                if int(Partes["cosyd_p_1_damage"]) == 1 and "1. Damage" not in Problemas:
                    Problemas.append("1. Damage")
                if int(Partes["cosyd_p_2_shortage"]) == 1 and "2. Shortage" not in Problemas:
                    Problemas.append("2. Shortage")
                if int(Partes["cosyd_p_3_surpluse"]) == 1 and "3. Surplus" not in Problemas:
                    Problemas.append("3. Surplus")
                if int(Partes["cosyd_p_4_asn_issue"]) == 1 and "4. ASN Issue" not in Problemas:
                    Problemas.append("4. ASN Issue")
                if int(Partes["cosyd_p_5_missing_doc_in_prisma"]) == 1 and "5. Missing doc in Prisma" not in Problemas:
                    Problemas.append("5. Missing doc in Prisma")
                Numeros += 1
            for K in PakingSplip.keys():
                if PakingSplip[K] is None:
                    PakingSplip[K]  = ""
            Aux_Datos = {}
            Aux_Datos["Estado"] = OK
            Opciones = """
            <div class="btn-group btn-group-sm" role="group">

            <button class='btn btn-warning btn-sm p-0 ps-1 pe-1' onclick='Ver_Historico("""+str(PakingSplip["cosyd_id"])+""",\""""+str(PakingSplip["cosyd_id"]).zfill(5)+"""\");'><i class='mdi mdi-history'></i></button>
            """
            if Numeros == 0:
                Aux_Datos["Estado"] = 2
                Opciones += """
                <button class='btn btn-primary btn-sm p-0 ps-1 pe-1' onclick='Completar_Ruteo("""+str(PakingSplip["cosyd_id"])+""");'><i class='mdi mdi-file-document-edit'></i></button>
                """
                if Historico == 1:
                    Opciones += """<button class='btn btn-danger btn-sm p-0 ps-1 pe-1' onclick='Eliminar_Ruteo("""+str(PakingSplip["cosyd_id"])+""",\""""+str(PakingSplip["cosyd_id"]).zfill(5)+"""\");'><i class='mdi mdi-trash-can'></i></button>"""
            else:
                Opciones += """
                <button class='btn btn-primary btn-sm p-0 ps-1 pe-1' onclick='Modificar_Ruteo("Editar","""+str(PakingSplip["cosyd_id"])+""",\""""+str(PakingSplip["cosyd_id"]).zfill(5)+"""\",\"ABIERTOS\",\""""+str(",".join(Problemas))+"""\");'><i class='mdi mdi-file-document-edit'></i></button>
                <button class='btn btn-dark btn-sm p-0 ps-1 pe-1' onclick='Generar_Archivo_Ruteo("""+str(PakingSplip["cosyd_id"])+""");'><i class='mdi mdi-printer'></i></button>
                """
                if Historico == 1:
                    Opciones += """<button class='btn btn-danger btn-sm p-0 ps-1 pe-1' onclick='Eliminar_Ruteo("""+str(PakingSplip["cosyd_id"])+""",\""""+str(PakingSplip["cosyd_id"]).zfill(5)+"""\");'><i class='mdi mdi-trash-can'></i></button>"""
            Opciones += """
            </div>"""
            Aux_Datos[" "] = str(Opciones)
            
            Lleva = ""
            Diff = datetime.now() - PakingSplip["Fecha_Ultimo"]
            if Diff.days > 0:
                Lleva += str(Diff.days)+" Day(s) "
            segundos = Diff.seconds
            horas, segundos = divmod(segundos, 3600)
            minutos, segundos = divmod(segundos, 60)

            if len(str(horas)) == 1:
                horas = "0"+str(horas)
            if len(str(minutos)) == 1:
                minutos = "0"+str(minutos)
            if len(str(segundos)) == 1:
                segundos = "0"+str(segundos)
            Lleva += str(horas)+":"+str(minutos)+":"+str(segundos)

            Aux_Datos["Ultima modificacion"] = str(Lleva)
            Aux_Datos["Folio"] = str(PakingSplip["cosyd_id"]).zfill(5)

            Aux_Datos["Ruta"] = str(PakingSplip["cosyd_ruta"]) + " ["+str(PakingSplip["cosyd_ruta_fecha_hora"])+"]"
            Aux_Datos["Fecha de Ruta"] = str(PakingSplip["cosyd_ruta_fecha_hora"])

            Aux_Datos["Fecha Alta"] = str(PakingSplip["cosyd_alta"])
            # if PakingSplip["cosyd_fecha_envio"] is not None and PakingSplip["cosyd_fecha_envio"] != "":
            #     Aux_Datos["Fecha de envío del proveedor"] = PakingSplip["cosyd_fecha_envio"].strftime("%y-%m-%d")
            Aux_Datos["Fecha de envío del proveedor"] = ""
            Aux_Datos["Proveedor"] = PakingSplip["cosyd_proveedor"]
            #Aux_Datos["PackingSlip"] = "<a class='link-primary' style='cursor:pointer' onclick='Ver_Detalle("+str(PakingSplip["cosyd_id"])+",\""+str(PakingSplip["cosyd_packingslip"])+"\")'>"+str(PakingSplip["cosyd_packingslip"])+"</a>"
            #Aux_Datos["Packing Slip"] = str(PakingSplip["cosyd_packingslip"])

            Aux_Datos["Problema"] = str(",".join(Problemas))

            if PakingSplip["Fecha_Ultimo"] is None or "Fecha_Ultimo" not in PakingSplip.keys():
                PakingSplip["Fecha_Ultimo"] = datetime.now()
            

            Aux_Datos["Destino"] = PakingSplip["cosyd_destino"]
            Aux_Datos["Contenedor"] = PakingSplip["cosyd_caja"]
            Aux_Datos["SCAC"] = PakingSplip["cosyd_scac"]
            Clase_Hora = ""
            if DB.Dame_Hora() <= PakingSplip["Fecha_Ultimo"] + timedelta(minutes=30):
                Clase_Hora = "text-bg-success"
            elif DB.Dame_Hora() <= PakingSplip["Fecha_Ultimo"] + timedelta(minutes=60):
                Clase_Hora = "text-bg-warning"
            elif DB.Dame_Hora() >= PakingSplip["Fecha_Ultimo"] + timedelta(hours=3):
                Clase_Hora = "Pulsa_Texto_Rojo"
            else:
                Clase_Hora = 'text-muted'

            Aux_Datos["Ultimo Comentario"] = "<small><span class='"+str(Clase_Hora)+" ps-1 pe-1 fw-lighter'>"+str(PakingSplip["Fecha_Ultimo"])+"</span> <span class='fw-bold'>"+str(DB.Dame_Nombre_IDUsuario(PakingSplip["Usuario"])) + "</span></small><br>"+str(PakingSplip["cosyd_comentario"])
            Aux_Datos["Numeros"] = "<a class='link-primary' style='cursor:pointer' onclick='Ver_Detalle("+str(PakingSplip["cosyd_id"])+")'>"+str(Numeros)+"</a>" 
            Aux_Datos["Numeros_de_Parte"] = str(Numeros)
            Aux_Datos["Archivos"] = ""
            Aux_Datos["Packing_Slip"] = PakingSplip["cosyd_packingslip"]
            for Archivo in str(PakingSplip["Evidencia"]).split(","):
                if Archivo.strip() != "":
                    if "pdf" in Archivo:
                        Aux_Datos["Archivos"] += "<a href='http://10.4.7.219:8080/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-file-pdf-box ms-1'></a>"
                    else:
                        Aux_Datos["Archivos"] += "<a href='http://10.4.7.219:8080/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-image ms-1'></a>"

            
            if str(PakingSplip["cosyd_tipo"]) == "RUTEO" and int(PakingSplip["cosyd_estado"]) == 1:
                Tabla_Datos_Ruteo.append(Aux_Datos)
            if str(PakingSplip["cosyd_tipo"]) == "PRODUCCION" and int(PakingSplip["cosyd_estado"]) == 1:
                Tabla_Datos_Porduccion.append(Aux_Datos)
            if int(PakingSplip["cosyd_estado"]) != 1:
                Aux_Datos[" "] = """<button class='btn btn-warning btn-sm p-0 ps-1 pe-1' onclick='Ver_Historico("""+str(PakingSplip["cosyd_id"])+""",\""""+str(PakingSplip["cosyd_id"]).zfill(5)+"""\");'><i class='mdi mdi-history'></i></button>"""
                Aux_Datos["Cerrado"] = str(PakingSplip["cosyd_baja"])
                Aux_Datos["Tipo"] = str(PakingSplip["cosyd_tipo"])
                if int(PakingSplip["cosyd_estado"]) == 0: 
                    Aux_Datos["Estado"] = 3
                Tabla_Datos_Cerrados.append(Aux_Datos)
        
        if Contador_Abierto["valor"] > 0:
            Contador_Abierto["clase"] = "text-bg-danger"
        if Contador_Liberados["valor"] > 0:
            Contador_Liberados["clase"] = "text-bg-danger"
        if Contador_Pre_Liberados["valor"] > 0:
            Contador_Pre_Liberados["clase"] = "text-bg-danger"
        Resultado["Contenido"] += """
        <div class='row'>
            <div class='col-12'>
                <div class='h3'><i class='mdi mdi-go-kart-track'></i> Ruteo</div>
                <div class='w-100 d-flex justify-content-end'>
                    <button class='btn btn-success mb-1 me-1' onclick='Nueva_Ruteo()'><i class='mdi mdi-plus'></i> Nuevo</button>
                    <button class='btn btn-primary mb-1' id="download-xlsx-ruteo"><i class='mdi mdi-microsoft-excel'></i> Descargar Excel</button>
                </div>
                <div id='Tabla-Ruteo' class='border border-dark bg-dark-subtle'></div>
                <script>
                    delete table_abiertos;
                    var Col = [
                        {title: ' ', field: ' ', formatter: 'html',download: false,width:100,headerSort:false}, 
                        {title: 'Folio', field: 'Folio', headerFilter: 'input',width:65},
                        {title: 'Packing Slip', field: 'Packing_Slip', headerFilter: 'input',width:100},
                        {title: 'Ruta', field: 'Ruta', headerFilter: 'input',width:140}, 
                        {title: 'Fecha Alta', field: 'Fecha Alta', headerFilter: 'input',width:100}, 
                        {title: 'Proveedor', field: 'Proveedor', headerFilter: 'input',width:90}, 
                        {title: 'Destino', field: 'Destino', headerFilter: 'input',width:80}, 
                        {title: 'Numeros de Parte', field: 'Numeros', formatter: 'html',download: false,width:30,hozAlign:"center"},
                        {title: 'Numeros_de_Parte', field: 'Numeros_de_Parte',download: true,visible:false}, 
                        {title: 'Problema', field: 'Problema',width:150},
                        {title: 'Ultimo Comentario', field: 'Ultimo Comentario', headerFilter: 'input',formatter:'html'}, 
                        {title: 'Archivo(s)', field: 'Archivos', formatter: 'html',download: false,headerSort:false,width:90}];
                    var table_abiertos = new Tabulator("#Tabla-Ruteo", {
                        minHeight:500,
                        data:"""+str(Tabla_Datos_Ruteo)+""",
                        initialSort:[
                            {column:"Fecha Alta", dir:"asc"}
                        ],
                        layout:"fitColumns",
                        pagination:"local",
                        rowFormatter:function(row){
                        if(row.getData().Estado == 1){
                                row.getElement().style.backgroundColor = "#5dff67";
                            }
                        },
                        paginationSize:50,  
                        paginationCounter:"rows",
                        columns:Col
                    });
                    document.getElementById("download-xlsx-ruteo").addEventListener("click", function(){
                        table_abiertos.download("xlsx", "Ruteo.xlsx", {sheetName:"My Data"});
                    });
                </script>

            </div>
            <div class='col-12'>
                <hr>
                <div class='h3'><i class='mdi mdi-forklift'></i> Operaciones</div>
                <div class='w-100 d-flex justify-content-end'>
                    <button class='btn btn-success mb-1' id="download-xlsx-produccion"><i class='mdi mdi-microsoft-excel'></i> Descargar Excel</button>
                </div>
                <div id='Tabla-Produccion' class='border border-dark bg-dark-subtle'></div>
                <script>
                    delete table_p;
                    var Col = [
                        {title: ' ', field: ' ', formatter: 'html',download: false,width:100,headerSort:false}, 
                        {title: 'Folio', field: 'Folio', headerFilter: 'input',width:65},
                        {title: 'Packing Slip', field: 'Packing_Slip', headerFilter: 'input',width:100},
                        {title: 'Ruta', field: 'Ruta', headerFilter: 'input',width:140}, 
                        {title: 'Fecha Alta', field: 'Fecha Alta', headerFilter: 'input',width:100}, 
                        {title: 'Proveedor', field: 'Proveedor', headerFilter: 'input',width:90}, 
                        {title: 'Destino', field: 'Destino', headerFilter: 'input',width:80}, 
                        {title: 'Numeros de Parte', field: 'Numeros', formatter: 'html',download: false,width:30,hozAlign:"center"},
                        {title: 'Numeros_de_Parte', field: 'Numeros_de_Parte',download: true,visible:false},
                        {title: 'Problema', field: 'Problema',width:150},
                        {title: 'Ultimo Comentario', field: 'Ultimo Comentario', headerFilter: 'input',formatter:'html'}, 
                        {title: 'Archivo(s)', field: 'Archivos', formatter: 'html',download: false,headerSort:false,width:90}];
                    var table_p = new Tabulator("#Tabla-Produccion", {
                        minHeight:500,
                        data:"""+str(Tabla_Datos_Porduccion)+""",
                        initialSort:[
                            {column:"Fecha Alta", dir:"desc"}
                        ],
                        layout:"fitColumns",
                        pagination:"local",
                        rowFormatter:function(row){
                            if(row.getData().Estado == 1){
                                    row.getElement().style.backgroundColor = "#5dff67";
                                }
                            if(row.getData().Estado == 2){
                                    row.getElement().style.backgroundColor = "#cbe00f";
                                }
                        },
                        paginationSize:50,  
                        paginationCounter:"rows",
                        columns:Col,
                    });
                    document.getElementById("download-xlsx-produccion").addEventListener("click", function(){
                        table_p.download("xlsx", "Produccion.xlsx", {sheetName:"My Data"});
                    });
                </script>
            </div
        </div>
        <script>
            $( document ).ready(function() {
                $('.tabulator-header-contents').addClass('bg-body-secondary').find('.tabulator-col').addClass('bg-body-secondary');
                $("#Tab_Abierto").find('.badge').html("""+str(Contador_Abierto["valor"])+""").removeClass('text-bg-danger text-bg-secondary').addClass('"""+str(Contador_Abierto["clase"])+"""');
                $("#Tab_Liberados").find('.badge').html("""+str(Contador_Liberados["valor"])+""").removeClass('text-bg-danger text-bg-secondary').addClass('"""+str(Contador_Liberados["clase"])+"""');
                $("#Tab_Pre_Liberados").find('.badge').html("""+str(Contador_Pre_Liberados["valor"])+""").removeClass('text-bg-danger text-bg-secondary').addClass('"""+str(Contador_Pre_Liberados["clase"])+"""');
            })
            function Nueva_Ruteo()
            {   
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-plus'></i> Nuevo OS&D de Ruteo");
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm').addClass('modal-xl');
                var parametros = {"Fun":'"""+str(fernet.encrypt("Nueva_Ruteo".encode()).decode("utf-8"))+"""'};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,'Process error ['+ jqXHR.status + " | " + textStatus  + " | " + errorThrown +']');}
                });
            }
            function Completar_Ruteo(ID){
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm').addClass('modal-lg')
                $("#Vent_1").modal("show").find(".modal-body").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>").parent().find(".modal-title").html("<i class='mdi mdi-plus'></i> Nuevo OS&D de Ruteo");
                var parametros = {"Fun":'"""+str(fernet.encrypt("Completar_Ruteo".encode()).decode("utf-8"))+"""',"ID_User":'"""+str(Datos["ID_User"])+"""',"ID":ID};
                $.ajax({
                    data:  parametros,
                    url:   \""""+str(request.url)+"""\",
                    type:  "post",
                    success:  function (response){var Resultado = JSON.parse(response); $("#Vent_1").find(".modal-body").html(Resultado["Contenido"]);},
                    error: function (jqXHR, textStatus, errorThrown){$("#Vent_1").find(".modal-body").html("<span style='color:red;'><b>ERROR</b></span> <hr>"+textStatus);}
                });
            }
            function Modificar_Ruteo(Accion,ID,Folio,De_Donde,Problemas){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm').addClass('modal-lg')
                $("#Vent_1").modal("show").find(".modal-body").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>").parent().find(".modal-title").html("<i class='mdi mdi-file-document-edit'></i> Edit ["+Folio+"] "+Problemas);
                var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_Ruteo".encode()).decode("utf-8"))+"""',"Accion":Accion,"ID":ID,"De_Donde":De_Donde};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    type:  "post",
                    success:  function (response){var Resultado = JSON.parse(response); $("#Vent_1").find(".modal-body").html(Resultado["Contenido"]);swal.close();},
                    error: function (jqXHR, textStatus, errorThrown){$("#Vent_1").find(".modal-body").html("<span style='color:red;'><b>ERROR</b></span> <hr>"+textStatus);}
                });
            }
            function Eliminar_Ruteo(ID,Folio){
                Swal.fire({
                title: '¿Estás seguro de eliminar el OS&D ['+Folio+']?',
                input: 'text',buttonsStyling: false,showCancelButton: true,confirmButtonText: "<i class='mdi mdi-check'></i> Si",cancelButtonText: "<i class='mdi mdi-close'></i> No",showLoaderOnConfirm: true,
                customClass: {confirmButton: 'btn btn-success ms-1 me-1',cancelButton: 'btn btn-danger ms-1 me-1'},
                preConfirm: (Comentario) => {
                    
                    if(Comentario.trim() == ""){
                        Mensaje(1,"Agrega comentario");
                    }else{
                        Mostrar_Ventana_Cargando(false);
                        var parametros = {"Fun":'"""+str(fernet.encrypt("Eliminar_Ruteo".encode()).decode("utf-8"))+"""',"Comentario":Comentario,"ID":ID};
                        $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                            success:  function (response)
                            {
                                var Resultado = JSON.parse(response);
                                if(Resultado["Estado"] == 1)
                                {
                                    $("#Vent_1").modal("hide");
                                    Mensaje(2);
                                    Cargar_Abiertos();
                                }
                                else
                                    Mensaje(0,Resultado["Contenido"]);
                                    
                            },
                            error: function (jqXHR, textStatus, errorThrown )
                            {
                                Mensaje(0,textStatus);
                            }
                        });
                    }
                    
                }
                })
            }
        </script>
        """

        
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Nueva_Ruteo(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Estado":0,"Contenido":""}
    try:
        Resultado["Contenido"] += "<div class='row'><div class='col'>"
        Formulario = {"Col":"", "Campos": [],"Clase": "Formato" }
        Formulario["Campos"].append({"tipo":"texto","campo":"Packing Slip","id":"PS","titulo":"Packing Slip","editable":True,"Requerido":1,"min":1,"max":20,"valor":"","Col":12})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += "</div><div class='col-auto pt-4' onclick='Cargar_Tally()'> <button class='btn btn-primary mt-3'><i class='mdi mdi-upload'></i> Cargar Tally</button>"
        
        Resultado["Contenido"] += "</div>"
        Resultado["Contenido"] += "</div>"
        Resultado["Contenido"] += """
        <div class='mb-1'>
            <button id="reactivity-add" class='btn btn-success btn-sm'><i class='mdi mdi-plus'></i> Add New Number Part</button>
            <button id="reactivity-delete" class='btn btn-danger btn-sm'><i class='mdi mdi-close'></i> Remove Last Number Part</button>
        </div>
        <div id="example-table" class='border border-dark bg-dark-subtle'></div>
        <hr>
        """

        Formulario = {"Col":"", "Campos": [],"Clase": "Formato" }
        Formulario["Campos"].append({"tipo":"texto","campo":"Ruta","titulo":"Ruta","editable":True,"Requerido":1,"min":1,"max":20,"valor":"","Col":6})
        Formulario["Campos"].append({"tipo":"fecha","campo":"Fecha de Rura","titulo":"Fecha de Ruta","Requerido":1,"Col":6,"valor":"","editable":True})
        Formulario["Campos"].append({"tipo":"seleccion","campo":"SCAC","titulo":"SCAC","Tipo_Opciones":"Query","Opciones":"SELECT cca_nombre as Valor, cca_nombre as Texto FROM "+str(BD_Nombre)+".ccarrier WHERE cca_activo = 1","Requerido":1,"Col":6,"valor":"","editable":True})
        Formulario["Campos"].append({"tipo":"texto","campo":"Contenedor-Caja","titulo":"Contenedor-Caja","editable":True,"Requerido":1,"min":1,"max":100,"valor":"","Col":6})
        Formulario["Campos"].append({"tipo":"texto","campo":"Destino","titulo":"Destino","editable":True,"Requerido":1,"min":1,"max":50,"valor":"","Col":6})
        Formulario["Campos"].append({"tipo":"seleccion","campo":"Proveedor","id":"Proveedor_Nuevo","titulo":"Proveedor","Tipo_Opciones":"Query","Opciones":"SELECT crilcpr_codigo as Valor, CONCAT(crilcpr_codigo,' - ',crilcpr_nombre) AS Texto FROM "+str(BD_Nombre)+".cproveedores WHERE crilcpr_activo = '1' ORDER BY crilcpr_codigo","Requerido":1,"Col":6,"valor":"","editable":True})
        Formulario["Campos"].append({"tipo":"multitexto","campo":"Comentario","titulo":"Comentario","editable":True,"Requerido":1,"min":1,"max":1500,"valor":"","Col":12})
        Formulario["Campos"].append({"tipo":"multitexto","campo":"Lista_distribucion","id":"Lista_distribucion_Nuevo","titulo":"Lista de distribución (Correos) separado por ´;´","editable":True,"Requerido":0,"min":0,"max":1500,"valor":"","Col":12})
        Formulario["Campos"].append({"tipo":"archivo","campo":"Archivos","titulo":"Archivo(s)","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["application/pdf","image/png","image/jpeg","image/gif"],"valor":"","editable":True})
        #Resultado["Contenido"] += "<div class='text-end mb-1'><button onclick='Nuevo_Proveedor()' class='btn btn-sm btn-success'><i class='mdi mdi-plus'></i> Agregar nuevo proveedor</button></div>"
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += """
        <script>
            delete table;
            var tabledata = [];
            var table = new Tabulator("#example-table", {
                height:"311px",
                layout:"fitColumns",
                reactiveData:true, //turn on data reactivity
                data:tabledata,
                columns:[
                    {title:"Numero de Parte", field:"Numero de Parte", editor:"input"},
                    {title:"Cantidad de ASN", field:"Cantidad de ASN", sorter:"number", editor:"input"},
                    {title:"Cantidad Real", field:"Cantidad Real", sorter:"number", editor:"input"},
        """
        for D in ["1 Damage","2 Shortage","3 Surplus","4 ASN Issue","5 Missing doc in Prisma"]:
            Resultado["Contenido"] += """{title:'"""+str(D)+"""', field:'"""+str(D)+"""', hozAlign:"center", editor:true, formatter:"tickCross",width:100,headerSort:false,headerHozAlign:"center"},"""
        Resultado["Contenido"] += """
                   {title:"6 Other (Explain in Comments)", field:"6 Other",editor:"input"}
                ],
            });
            table.on("cellEdited", function(cell){

            });
            table.on("tableBuilt", function(){
            """
        for D in ["1 Damage","2 Shortage","3 Surplus","4 ASN Issue","5 Missing doc in Prisma"]:
            Resultado["Contenido"] += """
                $(".tabulator-col[tabulator-field='"""+str(D)+"""']").html('<div class="w-100 h-100"><div class="text-center"><input value=" """+str(D)+""" " onclick="Seleccinar_Todos(this);" class="form-check-input" type="checkbox"></div><small>"""+str(D)+"""</small></div>');
            """
        Resultado["Contenido"] += """
            });

            document.getElementById("reactivity-add").addEventListener("click", function(){
                tabledata.push({"Numero de Parte":"","Cantidad de ASN":"","Cantidad Real":"","1 Damage":false,"2 Shortage":false,"3 Surplus":false,"4 ASN Issue":false,"5 Missing doc in Prisma":false,"6 Other":""});
            });
            document.getElementById("reactivity-delete").addEventListener("click", function(){
                tabledata.pop();
            });

            $("#Proveedor_Nuevo").on( "change", function() {
                Cargar_Lista_distribucion($("#Proveedor_Nuevo").find('option:selected').val());
            } );
            
            function Seleccinar_Todos(Control){
               if($(Control).is(':checked')){
                    tabledata.forEach((element) => {
                        try {
                            element[$(Control).val().trim()] = 1;
                        } catch (error) {
                        }
                    });
               }
                else{
                    tabledata.forEach((element) => {
                        try {
                            element[$(Control).val().trim()] = 0;
                        } catch (error) {
                        }
                    });
                }
            }

            function Cargar_Lista_distribucion(Codigo){
                $("#Lista_distribucion_Nuevo").val("Loading...");
                var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Lista_distribucion".encode()).decode("utf-8"))+"""',"Codigo":Codigo};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response){
                        var obj = JSON.parse(response);
                        if(obj["Estado"] == 1)
                        {
                            $("#Lista_distribucion_Nuevo").val(obj["Contenido"]);
                        }
                        else
                        {
                            $("#Lista_distribucion_Nuevo").val('Process error: '+obj["Contenido"]);
                        }
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                         $("#Lista_distribucion_Nuevo").val('Process error :'+ textStatus);
                    }
                });
            }
        </script>
        <hr>
        """
        Resultado["Contenido"] += """
        <br>
        <div class='w-100 text-center'><button class='btn btn-success w-50' onclick='Guardar_Nueva_Ruteo()'><i class='mdi mdi-floppy'></i> Save</button></div>
        <script>

            Actualizar_Cambia_Texto();
            function Guardar_Nueva_Ruteo()
            {
                var Ok = true;
                tabledata.forEach((element) => {
                if(element["Numero de Parte"].trim() == "")
                    Ok = false;
                if(element["Cantidad de ASN"].trim() == "")
                    Ok = false;
                if(element["1 Damage"] == false && element["2 Shortage"] == false && element["3 Surplus"] == false && element["4 ASN Issue"] == false && element["5 Missing doc in Prisma"] == false && element["6 Other"].trim() == "")
                    Ok = false;
                });
                if(tabledata.length == 0)
                    Ok = false;
                var Info = Dame_Formulario(".Formato",true);
                if(Info != null)
                {
                    if(Ok == false)
                    {
                        Swal.fire({icon: 'warning',position: 'top-end',title: 'Verifique lista de numero de parte, no puede haber campos en blanco y debe de incluir por lo menos un numero de parte',showConfirmButton: false,toast: true,background : "#ffeb96",timer: 1500,timerProgressBar: true});
                    }
                    else{
                        Info["Partes"] = tabledata;
                        Mostrar_Ventana_Cargando(false);
                        var parametros = {"Fun":'"""+str(fernet.encrypt("Guardar_Nueva_Ruteo".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info)};
                            $.ajax({
                                data:  parametros,url:   \""""+str(request.url)+"""\",type:  "post",
                                success:  function (response){
                                    var obj = JSON.parse(response);
                                    if(obj["Estado"] == 1)
                                    {
                                        $("#Vent_1").modal("hide");
                                        Swal.fire({icon: 'success',position: 'top-end',title: '¡Guardado con éxito!',showConfirmButton: false,toast: true,background : "#c9fad7",timer: 1500,timerProgressBar: true});
                                        Cargar_Abiertos();
                                    }
                                    else
                                    {
                                        Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ obj["Contenido"] +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                                    }
                                },
                                error: function (jqXHR, textStatus, errorThrown )
                                {
                                    Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ textStatus +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                                }
                        });
                    }

                    
                }

            }
            function Nuevo_Proveedor()
            {
                Mostrar_Ventana_Cargando(false);
                $("#Vent_2").modal("show").find(".modal-body").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>").parent().find(".modal-title").html("<i class='mdi mdi-plus'></i> New supplier");
                var parametros = {"Fun":'"""+str(fernet.encrypt("Nuevo_Proveedor".encode()).decode("utf-8"))+"""'};
                $.ajax({
                    data:  parametros,
                    url:   \""""+str(request.url)+"""\",
                    type:  "post",
                    success:  function (response){$("#Vent_2").find(".modal-body").html(response);swal.close();},
                    error: function (jqXHR, textStatus, errorThrown){$("#Vent_2").find(".modal-body").html("<span style='color:red;'><b>ERROR</b></span> <hr>"+textStatus);swal.close();}
                });
            }
            function Cargar_Tally(){
                if($("#PS").val().trim()==""){
                    Mensaje(1,"Agrega el número de Packing Slip");
                }else{
                    Mostrar_Ventana_Cargando(false);
                    $("#Vent_2").modal("show").find(".modal-body").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>").parent().find(".modal-title").html("<i class='mdi mdi-upload'></i> Cargar Tally");
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Tally".encode()).decode("utf-8"))+"""',"PS":$("#PS").val()};
                    $.ajax({
                        data:  parametros,
                        url:   \""""+str(request.url)+"""\",
                        type:  "post",
                        success:  function (response){var obj = JSON.parse(response); $("#Vent_2").find(".modal-body").html(obj["Contenido"]);swal.close();},
                        error: function (jqXHR, textStatus, errorThrown){$("#Vent_2").find(".modal-body").html("<span style='color:red;'><b>ERROR</b></span> <hr>"+textStatus);swal.close();}
                    });
                }
            }
        </script>
        """
    except:
        Resultado["Contenido"] += str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Cargar_Tally(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = { "Estado" :0, "Contenido":""}
    try:
        Formulario = {"Col":"", "Campos": [],"Clase": "Manifiesto" }
        Formulario["Campos"].append({"tipo":"archivo","campo":"Tally","titulo":"Tally","editable":True,"Requerido":1,"tipo_archivo":["application/vnd.ms-excel", "text/csv","application/csv"],"min":1,"max":1,"valor":"","Col":12})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += """
        <script>
            $('.Archivo-0').on('FilePond:processfile', function (e) {
                Mostrar_Ventana_Cargando(false);
                setTimeout(Cargar_Tally_Guardar, 1000);
            });
            function Cargar_Tally_Guardar(){
                var Archivos = [];
                $('.Archivo-0').find('.filepond--data').children().each(function() {
                    Archivos.push($(this).attr("archivo"));
                });
                if(Archivos.length > 0)
                    if(Archivos == "")
                    {
                        Mensaje(0,"Error al cargar archivo. Favor de valida.");
                        $("#Vent_2").modal("hide");
                    }
                    else{
                        Mostrar_Ventana_Cargando(false);
                        var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Tally_Guardar".encode()).decode("utf-8"))+"""',"Archivo":Archivos[0],"PS":'"""+str(Datos["PS"])+"""'};
                        $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                            success:  function (response)
                            {
                                var Resultado = JSON.parse(response);
                                if(Resultado["Estado"] == 1){
                                    $(".Formato[campo='Ruta']").val(Resultado["Ruta"]);
                                    $(".Formato[campo='Fecha de Rura']").val(Resultado["Fecha"]);
                                    $(".Formato[campo='Contenedor-Caja']").val(Resultado["Contenedor-Caja"]);
                                    $(".Formato[campo='Proveedor']").val(Resultado["Proveedor"]);
                                    $(".Formato[campo='Destino']").val(Resultado["Destino"]);
                                    Cargar_Lista_distribucion(Resultado["Proveedor"]);
                                    $("#Vent_2").modal("hide");
                                    Actualizar_Cambia_Texto();

                                    Resultado["Partes"].forEach((element) => {
                                        tabledata.push({"Numero de Parte":element[0],"Cantidad de ASN":element[1],"Cantidad Real":element[1],"1 Damage":false,"2 Shortage":false,"3 Surplus":false,"4 ASN Issue":false,"5 Missing doc in Prisma":false,"6 Other":""});
                                    });

                                    Mensaje(2);
                                }else{
                                    Mensaje(0,Resultado["Contenido"]);
                                    $("#Vent_2").modal("hide");
                                }
                                
                                
                            },
                            error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,textStatus);}
                        });
                       
                    }
                else
                    setTimeout(Cargar_Tally_Guardar, 500);
            }
            </script>
        """
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Cargar_Tally_Guardar(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Resulato = []
        #Resultado["Contenido"] += str("I:/Portal_File/"+str(Datos["Archivo"]))
        with open(str(str(current_app.root_path)+"/Files/"+str(Datos["Archivo"])), newline='') as csvfile:
            spamreader = csv.reader(csvfile, delimiter=',')
            for row in spamreader:
                Resulato.append(row)
        #Resultado["Contenido"] += str(Resulato)
        Resultado["Ruta"] = str(Resulato[1][2]).split(":")[1].strip()
        Resultado["Fecha"] = datetime.strptime(str(Resulato[1][4]).split(":")[1].strip(),"%m/%d/%y").strftime("%Y-%m-%d")
        Resultado["Contenedor-Caja"] = str(Resulato[1][6]).split(":")[1].strip()
        Resultado["Partes"] = []
        index = 0
        for R in Resulato:
            if index >= 4 and len(R) > 8 :
                if str(R[3]).strip() == str(Datos["PS"]).strip():
                    if str(R[1]).strip() != "":
                        Resultado["Proveedor"] = str(R[1]).strip()
                    if str(R[2]).strip() != "":
                        Resultado["Destino"] = str(R[2]).strip().split("-")[0]
                    Aux = []
                    Aux.append(str(R[5]).strip())
                    Aux.append(str(R[7]).replace(",","").strip())
                    Resultado["Partes"].append(Aux)
            index += 1
        os.remove(str(str(current_app.root_path)+"/Files/"+str(Datos["Archivo"])))
        Resultado["Estado"] = 1
        # A = []
        # with open("I:/Portal_File/"+str(Datos["Archivo"]), mode ='r') as file:
        #     A = csv.reader(file, delimiter=',')
        # Resultado["Contenido"] += str(A)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Cargar_Lista_distribucion(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = { "Estado" :0, "Contenido":""}
    try:
        Error = ""
        Corres = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cproveedores WHERE crilcpr_codigo = '"+str(Datos["Codigo"])+"' AND crilcpr_activo = '1' ")[0]["crilcpr_email_s"]
        if Corres is None:
            Resultado["Contenido"] += ""
        else:
            Resultado["Contenido"] += str(Corres)
        if Error == "":
            Resultado["Estado"] = 1
        Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Guardar_Nueva_Ruteo(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = { "Estado" :0, "Contenido":""}
    try:
        Info = json.loads(Datos["Info"])
        Error = ""
        Error += DB.Instruccion("""
        INSERT INTO """+str(BD_Nombre)+""".cosyd
        (cosyd_tipo,cosyd_proveedor,cosyd_usuario,cosyd_alta,cosyd_caja,cosyd_scac,cosyd_comentario,cosyd_correo,cosyd_archivos,cosyd_ruta,cosyd_ruta_fecha_hora)
        VALUES
        ('RUTEO','"""+str(Info["Proveedor"])+"""','"""+str(Datos["ID_User"])+"""',NOW(),'"""+str(Info["Contenedor-Caja"])+"""','"""+str(Info["SCAC"])+"""','"""+str(Info["Comentario"])+"""','"""+str(Info["Lista_distribucion"])+"""','"""+str(','.join(Info["Archivos"]))+"""','"""+str(Info["Ruta"])+"""','"""+str(Info["Fecha de Rura"])+"""')
        """)
        if Error == "":
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".cproveedores SET crilcpr_email_s = '"+str(Info["Lista_distribucion"])+"' WHERE crilcpr_codigo = '"+str(Info["Proveedor"])+"'")
            ID = DB.Get_Dato("SELECT MAX(cosyd_id) as \"ID\" FROM "+str(BD_Nombre)+".cosyd WHERE cosyd_tipo = 'RUTEO' AND cosyd_usuario = '"+str(Datos["ID_User"])+"' ")[0]["ID"]
            Error += DB.Instruccion("""
            INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
            (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
            VALUES
            ('"""+str(ID)+"""','"""+str(Datos["ID_User"])+"""','"""+str(Info["Comentario"])+"""','"""+str(','.join(Info["Archivos"]))+"""','ALTA',NOW())
            """)
            if Error == "":
                for Parte in Info["Partes"]:
                    cosyd_p_1_Damage = 0
                    cosyd_p_2_Shortage = 0
                    cosyd_p_3_Surplus = 0
                    cosyd_p_4_ASN_Issue = 0
                    cosyd_p_5_Missing_doc_in_Prisma = 0
                    cosyd_p_6_Other = "null"
                    if str(Parte["1 Damage"]) == "True" or str(Parte["1 Damage"]) == "1":
                        cosyd_p_1_Damage = 1
                    if str(Parte["2 Shortage"]) == "True" or str(Parte["2 Shortage"]) == "1":
                        cosyd_p_2_Shortage = 1
                    if str(Parte["3 Surplus"]) == "True" or str(Parte["3 Surplus"]) == "1":
                        cosyd_p_3_Surplus = 1
                    if str(Parte["4 ASN Issue"]) == "True" or str(Parte["4 ASN Issue"]) == "1":
                        cosyd_p_4_ASN_Issue = 1
                    if str(Parte["5 Missing doc in Prisma"]) == "True" or str(Parte["5 Missing doc in Prisma"]) == "1":
                        cosyd_p_5_Missing_doc_in_Prisma = 1
                    if str(Parte["6 Other"]).strip() != "":
                        cosyd_p_6_Other = "'"+str(Parte["6 Other"]).strip()+"'"
                    Error += DB.Instruccion("""
                    INSERT INTO """+str(BD_Nombre)+""".cosyd_partes
                    (cosyd_master,cosyd_parte,cosyd_cantidad_asn,cosyd_cantidad_real,cosyd_p_pakingslip,cosyd_p_destino,cosyd_p_1_damage,cosyd_p_2_shortage,cosyd_p_3_surpluse,cosyd_p_4_asn_issue,cosyd_p_5_missing_doc_in_prisma,cosyd_p_6_other)
                    VALUES
                    ('"""+str(ID)+"""','"""+str(Parte["Numero de Parte"])+"""','"""+str(Parte["Cantidad de ASN"])+"""','"""+str(Parte["Cantidad Real"])+"""','"""+str(Info["Packing Slip"])+"""','"""+str(Info["Destino"])+"""',"""+str(cosyd_p_1_Damage)+""","""+str(cosyd_p_2_Shortage)+""","""+str(cosyd_p_3_Surplus)+""","""+str(cosyd_p_4_ASN_Issue)+""","""+str(cosyd_p_5_Missing_doc_in_Prisma)+""","""+str(cosyd_p_6_Other)+""")
                    """)
        if Error == "":
            Resultado["Estado"] = 1
        Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Completar_Ruteo(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Estado":0,"Contenido":""}
    DB = LibDM_2023.DataBase()
    try:


        Info_Gen = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cosyd inner join "+str(BD_Nombre)+".cosyd_historico  WHERE  cosyd_id = '"+str(Datos["ID"])+"' ")[0]

        Formulario_2 = {"Col":"", "Campos": [],"Clase": "Formulario_2" }
        Formulario_2["Campos"].append({"tipo":"texto","campo":"Usuario","titulo":"Usuario","editable":False,"Requerido":1,"min":1,"max":50,"valor":DB.Dame_Nombre_IDUsuario(Info_Gen["cosyd_usuario"]),"Col":12})
        Formulario_2["Campos"].append({"tipo":"archivo","campo":"Archivos Producción","titulo":"Archivo(s) Producción","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["application/pdf","image/png","image/jpeg","image/gif"],"valor":Info_Gen["cosyd_archivos"],"editable":False})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario_2))

        
        Formulario = {"Col":"", "Campos": [],"Clase": "Formato" }
        Formulario["Campos"].append({"tipo":"texto","campo":"Ruta","titulo":"Ruta","editable":True,"Requerido":1,"min":1,"max":20,"valor":"","Col":6})
        Formulario["Campos"].append({"tipo":"fecha","campo":"Fecha de Rura","titulo":"Fecha de Ruta","Requerido":1,"Col":6,"valor":"","editable":True})
        Formulario["Campos"].append({"tipo":"seleccion","campo":"SCAC","titulo":"SCAC","Tipo_Opciones":"Query","Opciones":"SELECT cca_nombre as Valor, cca_nombre as Texto FROM "+str(BD_Nombre)+".ccarrier WHERE cca_activo = 1","Requerido":1,"Col":6,"valor":"","editable":True})
        Formulario["Campos"].append({"tipo":"texto","campo":"Contenedor-Caja","titulo":"Contenedor-Caja","editable":True,"Requerido":1,"min":1,"max":100,"valor":str(Info_Gen["cosyd_caja"]),"Col":6})
        Formulario["Campos"].append({"tipo":"texto","campo":"Destino","titulo":"Destino","editable":True,"Requerido":1,"min":1,"max":50,"valor":"","Col":6})
        Formulario["Campos"].append({"tipo":"seleccion","campo":"Proveedor","id":"Proveedor_Nuevo","titulo":"Proveedor","Tipo_Opciones":"Query","Opciones":"SELECT crilcpr_codigo as Valor, CONCAT(crilcpr_codigo,' - ',crilcpr_nombre) AS Texto FROM "+str(BD_Nombre)+".cproveedores WHERE crilcpr_activo = '1' ORDER BY crilcpr_codigo","Requerido":1,"Col":6,"valor":"","editable":True})
        Formulario["Campos"].append({"tipo":"multitexto","campo":"Comentario","titulo":"Comentario","editable":True,"Requerido":1,"min":1,"max":1500,"valor":"","Col":12})
        Formulario["Campos"].append({"tipo":"multitexto","campo":"Lista_distribucion","titulo":"Lista de distribución (Correos) separado por ´;´;","editable":True,"Requerido":0,"min":0,"max":1500,"valor":"","Col":12})
        Formulario["Campos"].append({"tipo":"archivo","campo":"Archivos","titulo":"Archivo(s)","Requerido":0,"Col":12,"min":0,"max":5,"tipo_archivo":["application/pdf","image/png","image/jpeg","image/gif"],"valor":"","editable":True})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))

        Resultado["Contenido"] += """
        <hr>
        <div class='mb-1'>
            <button id="reactivity-add" class='btn btn-success btn-sm'><i class='mdi mdi-plus'></i> Add New Number Part</button>
            <button id="reactivity-delete" class='btn btn-danger btn-sm'><i class='mdi mdi-close'></i> Remove Last Number Part</button>
        </div>
        <div id="example-table" class='border border-dark bg-dark-subtle'></div>
        <script>
            var tabledata = [];
            var table = new Tabulator("#example-table", {
                height:"311px",
                layout:"fitColumns",
                reactiveData:true, //turn on data reactivity
                data:tabledata,
                columns:[
                    {title:"Packing Slip", field:"Packing Slip", editor:"input"},
                    {title:"Numero de Parte", field:"Numero de Parte", editor:"input"},
                    {title:"Cantidad de ASN", field:"Cantidad de ASN", sorter:"number", editor:"input"},
                    {title:"Cantidad Real", field:"Cantidad Real", sorter:"number", editor:"input"},
                    /*{title:"Pallets", field:"Pallets", sorter:"number", editor:"input"},*/
                    {title:"OS&D", field:"OSD", editor:"list", editorParams:{values:{"1.- Damage":"1.- Damage","2.- Shortage":"2.- Shortage","3.- Surplus":"3.- Surplus","4.- ASN Issue":"4.- ASN Issue","5.- Missing doc. in Prisma":"5.- Missing doc. in Prisma","6.- Other (Explain in Comments)":"6.- Other (Explain in Comments)"}}},
                    {title:"Comentario", field:"Comentario",editor:"input"},
                ],
            });
            document.getElementById("reactivity-add").addEventListener("click", function(){
                var Ultimo_PS = "";
                tabledata.forEach((element) => {
                    Ultimo_PS = element["Packing Slip"].trim();
                });
                tabledata.push({"Packing Slip":Ultimo_PS,"Numero de Parte":"","Cantidad de ASN":"","Cantidad Real":"","OSD":"","Comentario":"","Pallets":"1"});
            });
            document.getElementById("reactivity-delete").addEventListener("click", function(){
                tabledata.pop();
            });
            $("#Proveedor_Nuevo").on( "change", function() {
                Cargar_Lista_distribucion($("#Proveedor_Nuevo").find('option:selected').val());
            } );
            function Cargar_Lista_distribucion(Codigo){
                $("#Lista_distribucion").val("Loading...");
                var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Lista_distribucion".encode()).decode("utf-8"))+"""',"Codigo":Codigo};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response){
                        var obj = JSON.parse(response);
                        if(obj["Estado"] == 1)
                        {
                            $("#Lista_distribucion").val(obj["Contenido"]);
                        }
                        else
                        {
                            $("#Lista_distribucion").val('Process error: '+obj["Contenido"]);
                        }
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                         $("#Lista_distribucion").val('Process error :'+ textStatus);
                    }
                });
            }
        </script>
        <hr>
        """
        
        html = "<table><thead><tr><th>Parte</th></tr></thead><tbody><tr><td>A</td></tr></tbody></table>"

        Resultado["Contenido"] += """
        <br>
        <div class='w-100 text-center'><button class='btn btn-outline-success w-50' onclick='Guardar_Completar_Ruteo("""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
        <script>

            Actualizar_Cambia_Texto();
            function Guardar_Completar_Ruteo(ID)
            {
                var Ok = true;
                tabledata.forEach((element) => {
                if(element["Numero de Parte"].trim() == "")
                    Ok = false;
                if(element["Cantidad de ASN"].trim() == "")
                    Ok = false;
                if(element["Pallets"].trim() == "")
                    Ok = false;
                /*if(element["Cantidad Real"].trim() == "")
                    Ok = false;*/
                if(element["OSD"].trim() == "")
                    Ok = false;
                if(element["Packing Slip"].trim() == "")
                    Ok = false;
                });
                if(tabledata.length == 0)
                    Ok = false;
                var Info = Dame_Formulario(".Formato",true);
                if(Info != null)
                {
                    if(Ok == false)
                    {
                        Swal.fire({icon: 'warning',position: 'top-end',title: 'Verifique lista de numero de parte, no puede haber campos en blanco y debe de incluir por lo menos un numero de parte',showConfirmButton: false,toast: true,background : "#ffeb96",timer: 1500,timerProgressBar: true});
                    }
                    else{
                        Info["Partes"] = tabledata;
                        Mostrar_Ventana_Cargando(false);
                        var parametros = {"Fun":'"""+str(fernet.encrypt("Guardar_Completar_Ruteo".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID_User":'"""+str(Datos["ID_User"])+"""',"ID":ID};
                            $.ajax({
                                data:  parametros,url:   \""""+str(request.url)+"""\",type:  "post",
                                success:  function (response){
                                    var obj = JSON.parse(response);
                                    if(obj["Estado"] == 1)
                                    {
                                        $("#Vent_1").modal("hide");
                                        Swal.fire({icon: 'success',position: 'top-end',title: '¡Guardado con éxito!',showConfirmButton: false,toast: true,background : "#c9fad7",timer: 1500,timerProgressBar: true});
                                        Cargar_Abiertos();
                                    }
                                    else
                                    {
                                        Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ obj["Contenido"] +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                                    }
                                },
                                error: function (jqXHR, textStatus, errorThrown )
                                {
                                    Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ textStatus +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                                }
                        });
                    }

                    
                }

            }

        </script>
        """
    except:
        Resultado["Contenido"] += str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Guardar_Completar_Ruteo(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = { "Estado" :0, "Contenido":""}
    try:
        Info = json.loads(Datos["Info"])
        Error = ""

        Tipo = None
        for Parte in Info["Partes"]:
            Tipo = str(Parte["OSD"])

        if str(Tipo) == "1.- Damage":
            Error += DB.Instruccion("""
            UPDATE """+str(BD_Nombre)+""".cosyd SET 
            cosyd_proveedor = '"""+str(Info["Proveedor"])+"""',
            cosyd_caja = '"""+str(Info["Contenedor-Caja"])+"""',
            cosyd_destino = '"""+str(Info["Destino"])+"""',
            cosyd_scac = '"""+str(Info["SCAC"])+"""',
            cosyd_comentario = '"""+str(Info["Comentario"])+"""',
            cosyd_correo = '"""+str(Info["Lista_distribucion"])+"""',
            cosyd_archivos = '"""+str(','.join(Info["Archivos"]))+"""',
            cosyd_ruta = '"""+str(Info["Ruta"])+"""',
            cosyd_ruta_fecha_hora = '"""+str(Info["Fecha de Rura"])+"""',
            cosyd_estado = 2
            WHERE cosyd_id = '"""+str(Datos["ID"])+"""'""")
        else:
            Error += DB.Instruccion("""
            UPDATE """+str(BD_Nombre)+""".cosyd SET 
            cosyd_proveedor = '"""+str(Info["Proveedor"])+"""',
            cosyd_caja = '"""+str(Info["Contenedor-Caja"])+"""',
            cosyd_destino = '"""+str(Info["Destino"])+"""',
            cosyd_scac = '"""+str(Info["SCAC"])+"""',
            cosyd_comentario = '"""+str(Info["Comentario"])+"""',
            cosyd_correo = '"""+str(Info["Lista_distribucion"])+"""',
            cosyd_archivos = '"""+str(','.join(Info["Archivos"]))+"""',
            cosyd_ruta = '"""+str(Info["Ruta"])+"""',
            cosyd_ruta_fecha_hora = '"""+str(Info["Fecha de Rura"])+"""',
            cosyd_estado = 1
            WHERE cosyd_id = '"""+str(Datos["ID"])+"""'""")

        Error += DB.Instruccion("""
        INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
        (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
        VALUES
        ('"""+str(Datos["ID"])+"""','"""+str(Datos["ID_User"])+"""','"""+str(Info["Comentario"])+"""','"""+str(','.join(Info["Archivos"]))+"""','COMPLETAR',NOW())
        """)
        if Error == "":
            for Parte in Info["Partes"]:
                Error += DB.Instruccion("""
                INSERT INTO """+str(BD_Nombre)+""".cosyd_partes
                (cosyd_master,cosyd_parte,cosyd_cantidad_asn,cosyd_cantidad_real,cosyd,cosyd_comentario,cosyd_p_pallets,cosyd_p_pakingslip,cosyd_p_destino)
                VALUES
                ('"""+str(Datos["ID"])+"""','"""+str(Parte["Numero de Parte"])+"""','"""+str(Parte["Cantidad de ASN"])+"""','"""+str(Parte["Cantidad Real"])+"""','"""+str(Parte["OSD"])+"""','"""+str(Parte["Comentario"])+"""','"""+str(Parte["Pallets"])+"""','"""+str(Parte["Packing Slip"])+"""','"""+str(Info["Destino"])+"""')
                """)

        if Error == "":
            Resultado["Estado"] = 1
        Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Modificar_Ruteo(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Estado":0,"Contenido":""}
    DB = LibDM_2023.DataBase()
    try:
        Formulario = {"Col":"", "Campos": [],"Clase": "Formato" }

        Info_Gen = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cosyd inner join "+str(BD_Nombre)+".cosyd_historico on public.cosyd.cosyd_id = public.cosyd_historico.cosyd_master  WHERE  cosyd_id = '"+str(Datos["ID"])+"' ")[0]

        Formulario["Campos"].append({"tipo":"texto","campo":"Ruta","titulo":"Ruta","editable":False,"Requerido":1,"min":1,"max":20,"valor":Info_Gen["cosyd_ruta"],"Col":6})
        Formulario["Campos"].append({"tipo":"fecha","campo":"Fecha de Rura","titulo":"Fecha de Ruta","Requerido":1,"Col":6,"valor":Info_Gen["cosyd_ruta_fecha_hora"],"editable":False})
        Formulario["Campos"].append({"tipo":"texto","campo":"SCAC","titulo":"SCAC","editable":False,"Requerido":1,"min":1,"max":20,"valor":Info_Gen["cosyd_scac"],"Col":6})
        Formulario["Campos"].append({"tipo":"texto","campo":"Contenedor-Caja","titulo":"Contenedor","editable":False,"Requerido":1,"min":0,"max":100,"valor":Info_Gen["cosyd_caja"],"Col":6})
        Formulario["Campos"].append({"tipo":"texto","campo":"Destino","titulo":"Destino","editable":False,"Requerido":0,"min":0,"max":50,"valor":Info_Gen["cosyd_destino"],"Col":6})
        Formulario["Campos"].append({"tipo":"seleccion","campo":"Proveedor","titulo":"Proveedor","editable":False,"Tipo_Opciones":"Query","Opciones":"SELECT crilcpr_codigo as Valor, CONCAT(crilcpr_codigo,' - ',crilcpr_nombre) AS Texto FROM "+str(BD_Nombre)+".cproveedores ORDER BY crilcpr_codigo","Requerido":1,"Col":6,"valor":Info_Gen["cosyd_proveedor"]})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Historico = DB.Get_Dato("select MASTER.cosyd_id,MASTER.cosyd_tipo,MASTER.cosyd_proveedor,MASTER.cosyd_packingslip,MASTER.cosyd_estado,MASTER.cosyd_alta,MASTER.cosyd_baja,MASTER.cosyd_caja,MASTER.cosyd_destino,MASTER.cosyd_scac,MASTER.cosyd_fecha_envio,MASTER.cosyd_correo,MASTER.cosyd_archivos,MASTER.cosyd_ruta,MASTER.cosyd_ruta_fecha_hora,HISTORICO.cosyd_master,HISTORICO.cosyd_usuario,HISTORICO.cosyd_comentario,HISTORICO.cosyd_evidencia,HISTORICO.cosyd_movimiento,HISTORICO.cosyd_fecha from "+str(BD_Nombre)+".cosyd MASTER inner join "+str(BD_Nombre)+".cosyd_historico HISTORICO on MASTER.cosyd_id = HISTORICO.cosyd_master  where MASTER.cosyd_id = '"+str(Datos["ID"])+"'")
        Tabla_Datos = []
        Columnas = []
        Columnas.append({"title":"Fecha y Hora", "field":"Fecha y Hora"})
        Columnas.append({"title":"Usuario", "field":"Usuario"})
        Columnas.append({"title":"Tipo", "field":"Tipo"})
        Columnas.append({"title":"Comentario", "field":"Comentario","formatter":"textarea"})
        Columnas.append({"title":"Archivo(s)", "field":"Archivos","formatter":"html"})
        for H in Historico:
            Aux_Datos ={}
            Aux_Datos["Fecha y Hora"] = H["cosyd_fecha"].strftime("%Y-%m-%d %H:%M:%S")
            Aux_Datos["Usuario"] = str(DB.Dame_Nombre_IDUsuario(H["cosyd_usuario"]))
            Aux_Datos["Tipo"]  = H["cosyd_movimiento"]
            Aux_Datos["Comentario"]  = H["cosyd_comentario"]
            Aux_Datos["Archivos"] = ""
            for Archivo in str(H["cosyd_evidencia"]).split(","):
                if Archivo.strip() != "":
                    if "pdf" in Archivo:
                        Aux_Datos["Archivos"] += "<a href='"+str(request.url_root)+"/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-file-pdf-box ms-1'></a>"
                    else:
                        Aux_Datos["Archivos"] += "<a href='"+str(request.url_root)+"/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-image ms-1'></a>"
            Tabla_Datos.append(Aux_Datos)

        Resultado["Contenido"] += """
        <br>
        <div><i class='mdi mdi-history'></i> Historial de comentarios</div>
        <div id='Tabla-Detalles' class='border border-dark bg-dark-subtle'></div>
        <script>
            delete table;
            var Col = """+str(Columnas)+""";
            var table = new Tabulator("#Tabla-Detalles", {
                height:300,
                data:"""+str(Tabla_Datos)+""",
                layout:"fitColumns",
                rowFormatter:function(row){
                if(row.getData().Estado == 1){
                        row.getElement().style.backgroundColor = "#5dff67";
                    }
                },
                columns:Col,
                initialSort:[
                    {column:"Fecha y Hora", dir:"desc"}
                ],
                selectable:false
            });
            document.getElementById("download-xlsx-ruteo").addEventListener("click", function(){
                table.download("xlsx", "Ruteo.xlsx", {sheetName:"My Data"});
            });
            $( document ).ready(function() {
                $('.tabulator-header-contents').addClass('bg-body-secondary').find('.tabulator-col').addClass('bg-body-secondary');
            })
        </script>
        """

        Formulario = {"Col":"", "Campos": [],"Clase": "Formato" }
        Formulario["Campos"].append({"tipo":"multitexto","campo":"Nuevo Comentario","titulo":"Nuevo Comentario","Requerido":1,"min":1,"max":1500,"valor":"","Col":12})
        Formulario["Campos"].append({"tipo":"archivo","campo":"Nuevo Archivos","titulo":"Nuevo Archivo(s)","Requerido":0,"Col":12,"min":0,"max":5,"tipo_archivo":["application/pdf","image/png","image/jpeg","image/gif"],"valor":""})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))

            
        Resultado["Contenido"] += """
        <br>
        <div class='w-100 text-center'>
            <button class='btn btn-warning w-75 mb-1' onclick='Guardar_Modificar_Ruteo("""+str(Datos["ID"])+""")'><i class='mdi mdi-message-alert'></i> Agregar Comentario/Evidencia OS&D</button>
        """
        
        if Datos["De_Donde"] == "ABIERTOS":
            Resultado["Contenido"] += """<button class='btn btn-success w-75' onclick='Cerrar_OSyD("""+str(Datos["ID"])+""",\"ABIERTOS\")'><i class='mdi mdi-check-bold'></i> Liberar OS&D</button>"""
        if Datos["De_Donde"] == "LIBERADOS":
            Resultado["Contenido"] += """<button class='btn btn-danger w-75 mb-1' onclick='Cerrar_OSyD("""+str(Datos["ID"])+""",\"REGRESAR\")'><i class='mdi mdi-arrow-left-bold'></i> Regresar a Abiertos</button>"""
            Resultado["Contenido"] += """<button class='btn btn-success w-75' onclick='Cerrar_OSyD("""+str(Datos["ID"])+""",\"CERRAR\")'><i class='mdi mdi-check-bold'></i> Cerrar OS&D</button>"""
        
        if Datos["De_Donde"] == "PRE_LIBERADO":
            Resultado["Contenido"] += """<button class='btn btn-danger w-75 mb-1' onclick='Cerrar_OSyD("""+str(Datos["ID"])+""",\"REGRESAR\")'><i class='mdi mdi-arrow-left-bold'></i> Regresar a Abiertos</button>"""
            Resultado["Contenido"] += """<button class='btn btn-success w-75' onclick='Cerrar_OSyD("""+str(Datos["ID"])+""",\"LIBERAR\")'><i class='mdi mdi-check-bold'></i> Liberar OS&D</button>"""

        Resultado["Contenido"] += """
        </div>
        <script>

            Actualizar_Cambia_Texto();
            function Guardar_Modificar_Ruteo(ID)
            {
                var Info = Dame_Formulario(".Formato",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Guardar_Modificar_Ruteo".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID":ID};
                        $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                            success:  function (response){
                                var obj = JSON.parse(response);
                                if(obj["Estado"] == 1)
                                {
                                    $("#Vent_1").modal("hide");
                                    Swal.fire({icon: 'success',position: 'top-end',title: '¡Guardado con éxito!',showConfirmButton: false,toast: true,background : "#c9fad7",timer: 1500,timerProgressBar: true});
        """
        if Datos["De_Donde"] == "ABIERTOS":
            Resultado["Contenido"] += "Cargar_Abiertos();"
        else:
            Resultado["Contenido"] += "Cargar_Liberados();"
        Resultado["Contenido"] += """
                                    
                                }
                                else
                                {
                                    Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ obj["Contenido"] +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                                }
                             },
                            error: function (jqXHR, textStatus, errorThrown )
                            {
                                Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ textStatus +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                            }
                    });
                }

            }
            function Cerrar_OSyD(ID,De_Donde){
                var Info = Dame_Formulario(".Formato",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Cerrar_OSyD".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID":ID,"De_Donde":De_Donde};
                        $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                            success:  function (response){
                                var obj = JSON.parse(response);
                                if(obj["Estado"] == 1)
                                {
                                    $("#Vent_1").modal("hide");
                                    Swal.fire({icon: 'success',position: 'top-end',title: '¡Guardado con éxito!',showConfirmButton: false,toast: true,background : "#c9fad7",timer: 1500,timerProgressBar: true});
        """
        if Datos["De_Donde"] == "ABIERTOS":
            Resultado["Contenido"] += "Cargar_Abiertos();"
        elif Datos["De_Donde"] == "PRE_LIBERADO":
            Resultado["Contenido"] += "Cargar_Pre_Liberados();"
        else:
            Resultado["Contenido"] += "Cargar_Liberados();"
        Resultado["Contenido"] += """
                                }
                                else
                                {
                                    Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ obj["Contenido"] +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                                }
                             },
                            error: function (jqXHR, textStatus, errorThrown )
                            {
                                Swal.fire({icon: 'error',position: 'top-end',title: 'Process error ['+ textStatus +']',showConfirmButton: false,toast: true,background : "#fac9c9",timer: 3500,timerProgressBar: true});
                            }
                    });
                }
            }
        </script>
        """
    except:
        Resultado["Contenido"] += str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Guardar_Modificar_Ruteo(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = { "Estado" :0, "Contenido":""}
    try:
        Info = json.loads(Datos["Info"])
        Error = ""
        Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".cosyd SET cosyd_comentario = '"+str(Info["Nuevo Comentario"])+"', cosyd_archivos = '"+str(','.join(Info["Nuevo Archivos"]))+"' WHERE cosyd_id = '"+str(Datos["ID"])+"'")
        Error += DB.Instruccion("""
        INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
        (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
        VALUES
        ('"""+str(Datos["ID"])+"""','"""+str(Datos["ID_User"])+"""','"""+str(Info["Nuevo Comentario"])+"""','"""+str(','.join(Info["Nuevo Archivos"]))+"""','ACTUALIZAR',NOW())
        """)
        if Error == "":
            Resultado["Estado"] = 1
        Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Cerrar_OSyD(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = { "Estado" :0, "Contenido":""}
    try:
        Info = json.loads(Datos["Info"])
        Error = ""
        if Datos["De_Donde"] == "ABIERTOS":
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".cosyd SET cosyd_comentario = '"+str(Info["Nuevo Comentario"])+"', cosyd_archivos = '"+str(','.join(Info["Nuevo Archivos"]))+"',cosyd_estado = '2' WHERE cosyd_id = '"+str(Datos["ID"])+"'")
            Error += DB.Instruccion("""
            INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
            (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
            VALUES
            ('"""+str(Datos["ID"])+"""','"""+str(Datos["ID_User"])+"""','"""+str(Info["Nuevo Comentario"])+"""','"""+str(','.join(Info["Nuevo Archivos"]))+"""','LIBERADO',NOW())
            """)
        elif Datos["De_Donde"] == "REGRESAR":
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".cosyd SET cosyd_comentario = '"+str(Info["Nuevo Comentario"])+"', cosyd_archivos = '"+str(','.join(Info["Nuevo Archivos"]))+"',cosyd_estado = '1' WHERE cosyd_id = '"+str(Datos["ID"])+"'")
            Error += DB.Instruccion("""
            INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
            (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
            VALUES
            ('"""+str(Datos["ID"])+"""','"""+str(Datos["ID_User"])+"""','"""+str(Info["Nuevo Comentario"])+"""','"""+str(','.join(Info["Nuevo Archivos"]))+"""','REGRESAR',NOW())
            """)
        elif Datos["De_Donde"] == "LIBERAR":
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".cosyd SET cosyd_comentario = '"+str(Info["Nuevo Comentario"])+"', cosyd_archivos = '"+str(','.join(Info["Nuevo Archivos"]))+"',cosyd_estado = '1' WHERE cosyd_id = '"+str(Datos["ID"])+"'")
            Error += DB.Instruccion("""
            INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
            (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
            VALUES
            ('"""+str(Datos["ID"])+"""','"""+str(Datos["ID_User"])+"""','"""+str(Info["Nuevo Comentario"])+"""','"""+str(','.join(Info["Nuevo Archivos"]))+"""','LIBERADO',NOW())
            """)
        else:
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".cosyd SET cosyd_comentario = '"+str(Info["Nuevo Comentario"])+"', cosyd_archivos = '"+str(','.join(Info["Nuevo Archivos"]))+"',cosyd_estado = '3', cosyd_baja = NOW() WHERE cosyd_id = '"+str(Datos["ID"])+"'")
            Error += DB.Instruccion("""
            INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
            (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
            VALUES
            ('"""+str(Datos["ID"])+"""','"""+str(Datos["ID_User"])+"""','"""+str(Info["Nuevo Comentario"])+"""','"""+str(','.join(Info["Nuevo Archivos"]))+"""','CERRADO',NOW())
            """)
        if Error == "":
            Resultado["Estado"] = 1
        Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Eliminar_Ruteo(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = { "Estado" :0, "Contenido":""}
    try:
        Error = ""
        Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".cosyd SET cosyd_comentario = '"+str(Datos["Comentario"])+"',cosyd_estado = 0 WHERE cosyd_id = '"+str(Datos["ID"])+"'")
        Error += DB.Instruccion("""
        INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
        (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
        VALUES
        ('"""+str(Datos["ID"])+"""','"""+str(Datos["ID_User"])+"""','"""+str(Datos["Comentario"])+"""','','ELIMINAR',NOW())
        """)
        if Error == "":
            Resultado["Estado"] = 1
        Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Ver_Historico(Datos):
    Compartido_2023 = LibDM_2023.Compartido()
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = {"Estado":0,"Contenido":""}
    try:
        Historico = DB.Get_Dato("select MASTER.cosyd_id,MASTER.cosyd_tipo,MASTER.cosyd_proveedor,MASTER.cosyd_packingslip,MASTER.cosyd_estado,MASTER.cosyd_alta,MASTER.cosyd_baja,MASTER.cosyd_caja,MASTER.cosyd_destino,MASTER.cosyd_scac,MASTER.cosyd_fecha_envio,MASTER.cosyd_correo,MASTER.cosyd_archivos,MASTER.cosyd_ruta,MASTER.cosyd_ruta_fecha_hora,HISTORICO.cosyd_master,HISTORICO.cosyd_usuario,HISTORICO.cosyd_comentario,HISTORICO.cosyd_evidencia,HISTORICO.cosyd_movimiento,HISTORICO.cosyd_fecha from "+str(BD_Nombre)+".cosyd MASTER inner join "+str(BD_Nombre)+".cosyd_historico HISTORICO on MASTER.cosyd_id = HISTORICO.cosyd_master  where MASTER.cosyd_id = '"+str(Datos["ID"])+"'")
        Tabla_Datos = []
        Columnas = []
        Columnas.append({"title":"Fecha y Hora", "field":"Fecha y Hora"})
        Columnas.append({"title":"Usuario", "field":"Usuario"})
        Columnas.append({"title":"Tipo", "field":"Tipo"})
        Columnas.append({"title":"Comentario", "field":"Comentario"})
        Columnas.append({"title":"Archivo(s)", "field":"Archivos","formatter":"html"})
        for H in Historico:
            Aux_Datos ={}
            Aux_Datos["Fecha y Hora"] = H["cosyd_fecha"].strftime("%Y-%m-%d %H:%M:%S")
            Aux_Datos["Usuario"] = str(DB.Dame_Nombre_IDUsuario(H["cosyd_usuario"]))
            Aux_Datos["Tipo"]  = H["cosyd_movimiento"]
            Aux_Datos["Comentario"]  = H["cosyd_comentario"]
            Aux_Datos["Archivos"] = ""
            for Archivo in str(H["cosyd_evidencia"]).split(","):
                if Archivo.strip() != "":
                    if "pdf" in Archivo:
                        Aux_Datos["Archivos"] += "<a href='"+str(request.url_root)+"/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-file-pdf-box ms-1'></a>"
                    else:
                        Aux_Datos["Archivos"] += "<a href='"+str(request.url_root)+"/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-image ms-1'></a>"
            Tabla_Datos.append(Aux_Datos)

        Resultado["Contenido"] += """
        <div id='Tabla-Detalles' class='border border-dark bg-dark-subtle'></div>
        <script>
            delete table;
            var Col = """+str(Columnas)+""";
            var table = new Tabulator("#Tabla-Detalles", {
                minHeight:300,
                data:"""+str(Tabla_Datos)+""",
                layout:"fitColumns",
                pagination:"local",
                rowFormatter:function(row){
                if(row.getData().Estado == 1){
                        row.getElement().style.backgroundColor = "#5dff67";
                    }
                },
                paginationSize:50,  
                paginationCounter:"rows",
                columns:Col,
                selectable:false
            });
            document.getElementById("download-xlsx-ruteo").addEventListener("click", function(){
                table.download("xlsx", "Ruteo.xlsx", {sheetName:"My Data"});
            });
            $( document ).ready(function() {
                $('.tabulator-header-contents').addClass('bg-body-secondary').find('.tabulator-col').addClass('bg-body-secondary');
            })
        </script>
        """
    except:
        Resultado["Contenido"] += str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Ver_Detalle(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Estado":0,"Contenido":""}
    DB = LibDM_2023.DataBase()
    try:
        Formulario = {"Col":"", "Campos": [],"Clase": "Formato" }
        Tabla_Datos = []
        Columnas = []
        Columnas.append({"title":"Packing Slip", "field":"Packing Slip","headerFilter":"input"})
        Columnas.append({"title":"Numero de Parte", "field":"Numero de Parte","headerFilter":"input"})
        Columnas.append({"title":"Destino", "field":"Destino","headerFilter":"input"})
        Columnas.append({"title":"Cantidad de ASN", "field":"Cantidad de ASN","headerFilter":"input"})
        Columnas.append({"title":"Cantidad Real", "field":"Cantidad Real","headerFilter":"input"})
        for D in ["1 Damage","2 Shortage","3 Surplus","4 ASN Issue","5 Missing doc in Prisma"]:
            Columnas.append({"title":str(D),"field":str(D),"hozAlign":"center","width":50,"headerSort":0,"headerHozAlign":"center"})
        Columnas.append({"title":"6 Other (Explain in Comments)","field":"6 Other","hozAlign":"center","width":100,"headerSort":0,"headerHozAlign":"center"})
        
        for Partes in DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cosyd_partes WHERE cosyd_master = '"+str(Datos["IDMaster"])+"'"):
            for K in Partes.keys():
                if Partes[K] is None:
                    Partes[K]  = ""
            Aux_Datos = {}
            Opciones = ""

            Aux_Datos["ID"] = str(Partes["cosyd_p_id"])
            Aux_Datos["Packing Slip"] = str(Partes["cosyd_p_pakingslip"])
            Aux_Datos["Numero de Parte"] = str(Partes["cosyd_parte"])
            Aux_Datos["Cantidad de ASN"] = str(Partes["cosyd_cantidad_asn"])
            Aux_Datos["Cantidad Real"] = str(Partes["cosyd_cantidad_real"])
            Aux_Datos["Destino"] = str(Partes["cosyd_p_destino"])
            Aux_Datos["Pallets"] = str(Partes["cosyd_p_pallets"])
            Aux_Datos["OSD"] = str(Partes["cosyd"])
            Aux_Datos["Comentario"] = str(Partes["cosyd_comentario"])
            Aux_Datos["Estado"] = int(Partes["cosyd_p_estado"])

            if Partes["cosyd"] != "":
                Partes["cosyd_p_1_damage"] = 'X' if str(Partes["cosyd"]) == "1.- Damage" else ""
                Partes["cosyd_p_2_shortage"] = 'X' if str(Partes["cosyd"]) == "2.- Surplus" else ""
                Partes["cosyd_p_3_surplusr"] = 'X' if str(Partes["cosyd"]) == "3.- Surplus" else ""
                Partes["cosyd_p_4_asn_issue"] = 'X' if str(Partes["cosyd"]) == "4.- ASN Issue" else ""
                Partes["cosyd_p_5_missing_doc_in_prisma"] = 'X' if str(Partes["cosyd"]) == "5.- Missing doc. in Prisma" else ""
                if Partes["cosyd_p_1_damage"] == "" and Partes["cosyd_p_2_shortage"] == "" and Partes["cosyd_p_3_surpluse"] == "" and Partes["cosyd_p_4_asn_issue"] == "" and Partes["cosyd_p_5_missing_doc_in_prisma"] == "":
                    Partes["cosyd_p_6_other"] = "["+str(Partes["cosyd"])+"] - " + str(Partes["cosyd_comentario"])
            else:
                Partes["cosyd_p_1_damage"] = 'X' if int(Partes["cosyd_p_1_damage"]) == 1 else ""
                Partes["cosyd_p_2_shortage"] = 'X' if int(Partes["cosyd_p_2_shortage"]) == 1 else ""
                Partes["cosyd_p_3_surpluse"] = 'X' if int(Partes["cosyd_p_3_surpluse"]) == 1 else ""
                Partes["cosyd_p_4_asn_issue"] = 'X' if int(Partes["cosyd_p_4_asn_issue"]) == 1 else ""
                Partes["cosyd_p_5_missing_doc_in_prisma"] = 'X' if int(Partes["cosyd_p_5_missing_doc_in_prisma"]) == 1 else ""
                Partes["cosyd_p_6_other"] = str(Partes["cosyd_p_6_other"]) if Partes["cosyd_p_6_other"] is not None else ""

            Aux_Datos["1 Damage"] = str(Partes["cosyd_p_1_damage"])
            Aux_Datos["2 Shortage"] = str(Partes["cosyd_p_2_shortage"])
            Aux_Datos["3 Surplus"] = str(Partes["cosyd_p_3_surpluse"])
            Aux_Datos["4 ASN Issue"] = str(Partes["cosyd_p_4_asn_issue"])
            Aux_Datos["5 Missing doc in Prisma"] = str(Partes["cosyd_p_5_missing_doc_in_prisma"])
            Aux_Datos["6 Other"] = str(Partes["cosyd_p_6_other"])

            
            Tabla_Datos.append(Aux_Datos)

        Resultado["Contenido"] += """
        <div class='w-100 d-flex justify-content-around'>
            <button class='btn btn-outline-success mb-1' id="download-xlsx-detalle"><i class='mdi mdi-microsoft-excel'></i> Descargar Excel</button>
        </div>
        <div id='Detalle-Partes' class='border border-dark bg-dark-subtle'></div>
        <script>
            delete table;
            var Col = """+str(Columnas)+""";
            var table = new Tabulator("#Detalle-Partes", {
                data:"""+str(Tabla_Datos)+""",
                layout:"fitColumns",
                rowFormatter:function(row){
                if(row.getData().Estado > 1){
                        row.getElement().style.backgroundColor = "#5dff67";
                    }
                },
                height:500,
                pagination:"local",
                paginationSize:50,  
                paginationCounter:"rows",
                columns:Col,
                selectable:false
            });
            document.getElementById("download-xlsx-detalle").addEventListener("click", function(){
                table.download("xlsx", '"""+str(Datos["IDMaster"])+""".xlsx', {sheetName:"My Data"});
            });
            $( document ).ready(function() {
                $('.tabulator-header-contents').addClass('bg-body-secondary').find('.tabulator-col').addClass('bg-body-secondary');
            })
            table.on("cellEdited", function(cell){

                var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_Parte_Destino".encode()).decode("utf-8"))+"""',"ID":cell.getData()["ID"],"Destino":cell.getValue()};
                 $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    type:  "post",
                    success:  function (response){},
                    error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,'Process error ['+ jqXHR.status + " | " + textStatus  + " | " + errorThrown +']');}
                });
                    
            });
            function Modificar_Parte(Accion,ID,Parte){
                $("#Vent_2").removeClass('modal-xl modal-lg modal-sm').addClass('modal-lg')
                $("#Vent_2").modal("show").find(".modal-body").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>").parent().find(".modal-title").html("<i class='mdi mdi-plus'></i> Nuevo OS&D de Ruteo");
                var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_Parte".encode()).decode("utf-8"))+"""',"ID_User":'"""+str(Datos["ID_User"])+"""',"Accion":Accion,"ID":ID,"Parte":Parte};
                $.ajax({
                    data:  parametros,
                    url:   \""""+str(request.url)+"""\",
                    type:  "post",
                    success:  function (response){var Resultado = JSON.parse(response); $("#Vent_2").find(".modal-body").html(Resultado["Contenido"]);},
                    error: function (jqXHR, textStatus, errorThrown){$("#Vent_2").find(".modal-body").html("<span style='color:red;'><b>ERROR</b></span> <hr>"+textStatus);}
                });
            }
        </script>
        """
    except:
        Resultado["Contenido"] += str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur

def Cargar_Liberados(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Contador_Abierto = {"valor":0,"clase":"text-bg-secondary"}
        Contador_Liberados = {"valor":0,"clase":"text-bg-secondary"}
        Resultado["Contenido"] += "<div class='h3 text-center mt-2'>OS&D Liberados</div>"

        Tabla_Datos_Liberados = []
        for PakingSplip in DB.Get_Dato("""
        SELECT MASTER.*,MAX(HISTORICO.cosyd_fecha) "Fecha_Ultimo",string_agg(DISTINCT(PARTES.cosyd_p_destino),',') "Destinos"  FROM """+str(BD_Nombre)+""".cosyd  MASTER
        inner join """+str(BD_Nombre)+""".cosyd_historico HISTORICO on MASTER.cosyd_id = HISTORICO.cosyd_master 
        left join """+str(BD_Nombre)+""".cosyd_partes PARTES ON PARTES.cosyd_master = MASTER.cosyd_id
        WHERE MASTER.cosyd_estado IN (1,2)
        group by MASTER.cosyd_id  ORDER BY cosyd_alta
        """):
            if int(PakingSplip["cosyd_estado"]) == 1:
                Contador_Abierto["valor"] += 1
            if int(PakingSplip["cosyd_estado"]) == 2:
                Contador_Liberados["valor"] += 1
            OK = 0
            Numeros = 0
            Cumeros_Ok = 0
            Problemas = []
            for Partes in DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cosyd_partes WHERE cosyd_master = '"+str(PakingSplip["cosyd_id"])+"'"):
                if int(Partes["cosyd_p_1_damage"]) == 1 and "1. Damage" not in Problemas:
                    Problemas.append("1. Damage")
                if int(Partes["cosyd_p_2_shortage"]) == 1 and "2. Shortage" not in Problemas:
                    Problemas.append("2. Shortage")
                if int(Partes["cosyd_p_3_surpluse"]) == 1 and "3. Surplus" not in Problemas:
                    Problemas.append("3. Surplus")
                if int(Partes["cosyd_p_4_asn_issue"]) == 1 and "4. ASN Issue" not in Problemas:
                    Problemas.append("4. ASN Issue")
                if int(Partes["cosyd_p_5_missing_doc_in_prisma"]) == 1 and "5. Missing doc in Prisma" not in Problemas:
                    Problemas.append("5. Missing doc in Prisma")
                if int(Partes["cosyd_p_estado"]) != 1:
                    Cumeros_Ok += 1
                Numeros += 1
            for K in PakingSplip.keys():
                if PakingSplip[K] is None:
                    PakingSplip[K]  = ""
            Aux_Datos = {}
            Aux_Datos["Estado"] = OK
            Opciones = """
            <button class='btn btn-warning btn-sm p-0 ps-1 pe-1' onclick='Ver_Historico("""+str(PakingSplip["cosyd_id"])+""",\""""+str(PakingSplip["cosyd_id"]).zfill(5)+"""\");'><i class='mdi mdi-history'></i></button>
            """
            if Numeros == 0:
                Aux_Datos["Estado"] = 2
                Opciones += """
                <button class='btn btn-primary btn-sm p-0 ps-1 pe-1' onclick='Completar_Ruteo("""+str(PakingSplip["cosyd_id"])+""");'><i class='mdi mdi-file-document-edit'></i></button>
                """
            else:
                Opciones += """
                <button class='btn btn-primary btn-sm p-0 ps-1 pe-1' onclick='Modificar_Ruteo("Editar","""+str(PakingSplip["cosyd_id"])+""",\""""+str(PakingSplip["cosyd_id"]).zfill(5)+"""\",\"LIBERADOS\",\""""+str(",".join(Problemas))+"""\");'><i class='mdi mdi-file-document-edit'></i></button>
                <button class='btn btn-dark btn-sm p-0 ps-1 pe-1' onclick='Generar_Archivo_Ruteo("""+str(PakingSplip["cosyd_id"])+""");'><i class='mdi mdi-printer'></i></button>
                """
            
            Diff = datetime.now() - PakingSplip["Fecha_Ultimo"]
            Aux_Datos[" "] = str(Opciones)
            Lleva = ""
            if Diff.days > 0:
                Lleva += str(Diff.days)+" Day(s) "
            segundos = Diff.seconds
            horas, segundos = divmod(segundos, 3600)
            minutos, segundos = divmod(segundos, 60)

            if len(str(horas)) == 1:
                horas = "0"+str(horas)
            if len(str(minutos)) == 1:
                minutos = "0"+str(minutos)
            if len(str(segundos)) == 1:
                segundos = "0"+str(segundos)
            Lleva += str(horas)+":"+str(minutos)+":"+str(segundos)

            Aux_Datos["Ultima modificacion"] = str(Lleva)
            Aux_Datos["Folio"] = str(PakingSplip["cosyd_id"]).zfill(5)
            Aux_Datos["Problema"] = str(",".join(Problemas))

            Aux_Datos["Ruta"] = str(PakingSplip["cosyd_ruta"])
            Aux_Datos["Fecha de Ruta"] = str(PakingSplip["cosyd_ruta_fecha_hora"])

            Aux_Datos["Fecha Alta"] = str(PakingSplip["cosyd_alta"])
            Aux_Datos["Fecha de envío del proveedor"] = ""
            Aux_Datos["Proveedor"] = PakingSplip["cosyd_proveedor"]
            Aux_Datos["Packing Slip"] = str(PakingSplip["cosyd_packingslip"])
            Aux_Datos["Destino"] = PakingSplip["Destinos"]
            Aux_Datos["Contenedor"] = PakingSplip["cosyd_caja"]
            Aux_Datos["SCAC"] = PakingSplip["cosyd_scac"]
            Aux_Datos["Ultimo Comentario"] = PakingSplip["cosyd_comentario"]
            Aux_Datos["Numeros"] = "<a class='link-primary' style='cursor:pointer' onclick='Ver_Detalle("+str(PakingSplip["cosyd_id"])+")'>"+str(Numeros)+"</a>" 
            Aux_Datos["Numeros_de_Parte"] = str(Numeros)
            Aux_Datos["Archivos"] = ""
            for Archivo in str(PakingSplip["cosyd_archivos"]).split(","):
                if Archivo.strip() != "":
                    if "pdf" in Archivo:
                        Aux_Datos["Archivos"] += "<a href='"+str(request.url_root)+"/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-file-pdf-box ms-1'></a>"
                    else:
                        Aux_Datos["Archivos"] += "<a href='"+str(request.url_root)+"/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-image ms-1'></a>"

            
            if int(PakingSplip["cosyd_estado"]) == 2:
                Tabla_Datos_Liberados.append(Aux_Datos)
        

        Tabla_Datos_Liberados_Partes = []
        for PakingSplip in DB.Get_Dato("""
        SELECT 
        PARTES.cosyd_p_pakingslip,MASTER.cosyd_id, MASTER.cosyd_ruta,MASTER.cosyd_ruta_fecha_hora,MASTER.cosyd_alta, MASTER.cosyd_proveedor,PARTES.cosyd_p_destino, PARTES.cosyd_parte,MASTER.cosyd_caja,PARTES.cosyd_cantidad_real,MASTER.cosyd_scac,MASTER.cosyd_comentario,MASTER.cosyd_archivos  
        FROM """+str(BD_Nombre)+""".cosyd  MASTER
        inner join """+str(BD_Nombre)+""".cosyd_partes PARTES ON PARTES.cosyd_master = MASTER.cosyd_id
        WHERE MASTER.cosyd_estado = 2
        ORDER BY cosyd_alta
        """):
            Tabla_Datos_Liberados_Partes.append({"Cantidad":str(PakingSplip["cosyd_cantidad_real"]),"Packing Slip":str(PakingSplip["cosyd_p_pakingslip"]),"Folio":str(PakingSplip["cosyd_id"]),"Ruta":str(PakingSplip["cosyd_ruta"]),"Fecha de Ruta":str(PakingSplip["cosyd_ruta_fecha_hora"]),"Fecha Alta":str(PakingSplip["cosyd_alta"]),"Proveedor":str(PakingSplip["cosyd_proveedor"]),"Destino":str(PakingSplip["cosyd_p_destino"]),"Numero de Parte":str(PakingSplip["cosyd_parte"]),"Contenedor":str(PakingSplip["cosyd_caja"]),"SCAC":str(PakingSplip["cosyd_scac"]),"Ultimo Comentario":str(PakingSplip["cosyd_comentario"])})
            pass
        if Contador_Abierto["valor"] > 0:
            Contador_Abierto["clase"] = "text-bg-danger"
        if Contador_Liberados["valor"] > 0:
            Contador_Liberados["clase"] = "text-bg-danger"
        Resultado["Contenido"] += """
        <div class='row'>
            <div class='col-12'>
                <div class='w-100 d-flex justify-content-end'>
                    <!--<button class='btn btn-outline-primary mb-1' onclick='Nueva_Ruteo()'><i class='mdi mdi-plus'></i> Nuevo</button>-->
                    <button class='btn btn-primary mb-1' id="download-xlsx-ruteo"><i class='mdi mdi-microsoft-excel'></i> Descargar Excel</button>
                    <button class='btn btn-dark mb-1 ms-1' id="download-xlsx-ruteo-partes"><i class='mdi mdi-microsoft-excel'></i> Descargar Excel Partes</button>
                </div>
                <div id='Tabla-Ruteo' class='border border-dark bg-dark-subtle'></div>
                <script>
                    delete table_liberados;
                    var Col = [
                        {title: ' ', field: ' ', formatter: 'html',download: false,width:120}, 
                        {title: 'Packing Slip', field: 'Packing Slip', formatter: 'html'}, 
                        {title: 'Folio', field: 'Folio', headerFilter: 'input'}, 
                        {title: 'Ruta', field: 'Ruta', headerFilter: 'input'}, 
                        {title: 'Fecha de Ruta', field: 'Fecha de Ruta', headerFilter: 'input'}, 
                        {title: 'Fecha Alta', field: 'Fecha Alta', headerFilter: 'input'}, 
                        {title: 'Proveedor', field: 'Proveedor', headerFilter: 'input'}, 
                        {title: 'Destino', field: 'Destino', headerFilter: 'input'}, 
                        {title: 'Numeros de Parte', field: 'Numeros', formatter: 'html',download: false,width:50},
                        {title: 'Numeros_de_Parte', field: 'Numeros_de_Parte',download: true,visible:false,width:50}, 
                        {title: 'Problema', field: 'Problema',width:150},
                        {title: 'Contenedor', field: 'Contenedor', headerFilter: 'input'}, 
                        {title: 'SCAC', field: 'SCAC', headerFilter: 'input'}, 
                        {title: 'Ultimo Comentario', field: 'Ultimo Comentario', headerFilter: 'input'}, 
                        {title: 'Archivo(s)', field: 'Archivos', formatter: 'html',download: false}];
                    var table_liberados = new Tabulator("#Tabla-Ruteo", {
                        minHeight:800,
                        data:"""+str(Tabla_Datos_Liberados)+""",
                        initialSort:[
                            {column:"Fecha Alta", dir:"desc"}
                        ],
                        layout:"fitColumns",
                        pagination:"local",
                        rowFormatter:function(row){
                        if(row.getData().Estado == 1){
                                row.getElement().style.backgroundColor = "#5dff67";
                            }
                        },
                        paginationSize:50,  
                        paginationCounter:"rows",
                        columns:Col,
                        selectable:false
                    });
                    document.getElementById("download-xlsx-ruteo").addEventListener("click", function(){
                        table_liberados.download("xlsx", "Ruteo.xlsx", {sheetName:"My Data"});
                    });
                </script>


                <div style='display:none;' id='Tabla-Ruteo-Partes' class='border border-dark bg-dark-subtle'></div>
                <script>
                    delete table_liberados_Partes;
                    var Col = [
                        {title: 'Packing Slip', field: 'Packing Slip'}, 
                        {title: 'Folio', field: 'Folio', headerFilter: 'input'}, 
                        {title: 'Ruta', field: 'Ruta', headerFilter: 'input'}, 
                        {title: 'Fecha de Ruta', field: 'Fecha de Ruta', headerFilter: 'input'}, 
                        {title: 'Fecha Alta', field: 'Fecha Alta', headerFilter: 'input'}, 
                        {title: 'Proveedor', field: 'Proveedor', headerFilter: 'input'}, 
                        {title: 'Destino', field: 'Destino', headerFilter: 'input'}, 
                        {title: 'Numero de Parte', field: 'Numero de Parte'},
                        {title: 'Cantidad', field: 'Cantidad'},
                        {title: 'Contenedor', field: 'Contenedor', headerFilter: 'input'}, 
                        {title: 'SCAC', field: 'SCAC', headerFilter: 'input'}, 
                        {title: 'Ultimo Comentario', field: 'Ultimo Comentario', headerFilter: 'input'}];
                    var table_liberados_Partes = new Tabulator("#Tabla-Ruteo-Partes", {
                        minHeight:800,
                        data:"""+str(Tabla_Datos_Liberados_Partes)+""",
                        initialSort:[
                            {column:"Fecha Alta", dir:"desc"}
                        ],
                        layout:"fitColumns",
                        pagination:"local",
                        rowFormatter:function(row){
                        if(row.getData().Estado == 1){
                                row.getElement().style.backgroundColor = "#5dff67";
                            }
                        },
                        paginationSize:50,  
                        paginationCounter:"rows",
                        columns:Col,
                        selectable:false
                    });
                    document.getElementById("download-xlsx-ruteo-partes").addEventListener("click", function(){
                        table_liberados_Partes.download("xlsx", "Ruteo.xlsx", {sheetName:"My Data"});
                    });
                    table_liberados_Partes.on("tableBuilt", function(){
                            $("#Tabla-Ruteo-Partes").hide();
                    });
                    
                </script>


            </div>
        </div>
        <script>
            $( document ).ready(function() {
                $('.tabulator-header-contents').addClass('bg-body-secondary').find('.tabulator-col').addClass('bg-body-secondary');
                $("#Tab_Abierto").find('.badge').html("""+str(Contador_Abierto["valor"])+""").removeClass('text-bg-danger text-bg-secondary').addClass('"""+str(Contador_Abierto["clase"])+"""');
                $("#Tab_Liberados").find('.badge').html("""+str(Contador_Liberados["valor"])+""").removeClass('text-bg-danger text-bg-secondary').addClass('"""+str(Contador_Liberados["clase"])+"""');
            })
            function Modificar_Ruteo(Accion,ID,Folio,De_Donde,Problemas){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm').addClass('modal-lg')
                $("#Vent_1").modal("show").find(".modal-body").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>").parent().find(".modal-title").html("<i class='mdi mdi-file-document-edit'></i> Edit ["+Folio+"] "+Problemas);
                var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_Ruteo".encode()).decode("utf-8"))+"""',"Accion":Accion,"ID":ID,"De_Donde":De_Donde};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    type:  "post",
                    success:  function (response){var Resultado = JSON.parse(response); $("#Vent_1").find(".modal-body").html(Resultado["Contenido"]);swal.close();},
                    error: function (jqXHR, textStatus, errorThrown){$("#Vent_1").find(".modal-body").html("<span style='color:red;'><b>ERROR</b></span> <hr>"+textStatus);}
                });
            }
        </script>
        """

        
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Cargar_Menu_Reporte(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Resultado["Contenido"] += "<div class='h3 text-center mt-2'>OS&D Reporte</div>"
        Resultado["Contenido"] += """
        <div class='col-12'>
        """
        Formulario = {"Col":"", "Campos": [],"Clase": "Reporte" }
        Formulario["Campos"].append({"tipo":"fecha-rango","campo":"Rango de fechas","titulo":"Rango de fechas","Requerido":1,"Col":12,"valor":"","editable":True})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += "<div class='w-100 text-center mb-3'><button class='btn btn-primary w-75' onclick='Reporte()'><i class='mdi mdi-download'></i> Cargar</button></div>"
        Resultado["Contenido"] += """
                <div id='Reporte' class='w-100 mb-5 p-1' style='height:600px; background:#e6e6e6'></div>
            </div
        </div>
        <script>
            function Reporte()
            {
                var Info = Dame_Formulario(".Reporte",true);
                if(Info != null)
                {
                    $("#Reporte").html("<div class='w-100 text-center'><div class='spinner-grow' role='status'></div><b> Loading...</b></div>");
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Reporte".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info)};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response){ var Resultado = JSON.parse(response); $("#Reporte").html(Resultado["Contenido"]);},
                        error: function (jqXHR, textStatus, errorThrown){$("#Reporte").html("<span style='color:red;'><b>ERROR</b></span> <hr>"+textStatus);}
                    });   
                }
            }
            $( document ).ready(function() {
                $('.tabulator-header-contents').addClass('bg-body-secondary').find('.tabulator-col').addClass('bg-body-secondary');
                //Reporte();
            })
        </script>
        """

        
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Cargar_Reporte(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Info = json.loads(Datos["Info"])
        Resultado["Contenido"] += "<h3>"+str(Info["Rango de fechas"][0])+" to "+str(Info["Rango de fechas"][1])+"</h3>"
        Tabla_Datos_Cerrados = []
        Proveedores = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cproveedores")
        PakingSplips = DB.Get_Dato("""
        SELECT MASTER.*,PARTES.cosyd_parte,PARTES.cosyd_cantidad_asn,PARTES.cosyd_cantidad_real,PARTES.cosyd,PARTES.cosyd_comentario as Comm_Parte,PARTES.cosyd_p_pakingslip,PARTES.cosyd_p_destino,PARTES.cosyd_p_1_Damage,PARTES.cosyd_p_2_Shortage,PARTES.cosyd_p_3_Surpluse,PARTES.cosyd_p_4_ASN_Issue,PARTES.cosyd_p_5_Missing_doc_in_Prisma,PARTES.cosyd_p_6_Other FROM """+str(BD_Nombre)+""".cosyd  MASTER
        inner join """+str(BD_Nombre)+""".cosyd_partes PARTES on MASTER.cosyd_id = PARTES.cosyd_master
        WHERE MASTER.cosyd_alta between '"""+str(Info["Rango de fechas"][0])+""" 00:00:00' and '"""+str(Info["Rango de fechas"][1])+""" 23:59:59' and cosyd_estado > 0
        ORDER BY cosyd_alta
        """)
        IDMaster = []
        for PakingSplip in PakingSplips:
            IDMaster.append(str(PakingSplip["cosyd_id"]))

        Historicos = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cosyd_historico WHERE cosyd_master IN ("+','.join(IDMaster)+")")

        for PakingSplip in PakingSplips:
            
            H = []
            for Hi in Historicos:
                if int(Hi["cosyd_master"]) == int(PakingSplip["cosyd_id"]):
                    H.append(Hi)

            for K in PakingSplip.keys():
                if PakingSplip[K] is None:
                    PakingSplip[K] = ""
            OK = 0
            Numeros = 0
            Aux_Datos = {}
            Aux_Datos["Estado"] = OK
            Opciones = """
            <button class='btn btn-warning btn-sm p-0 ps-1 pe-1' onclick='Ver_Historico("""+str(PakingSplip["cosyd_id"])+""",\""""+str(PakingSplip["cosyd_id"]).zfill(5)+"""\");'><i class='mdi mdi-history'></i></button>
            """
            Aux_Datos["Folio"] = str(PakingSplip["cosyd_id"]).zfill(5)
            Aux_Datos["Alta"] = str(PakingSplip["cosyd_alta"])
            Aux_Datos["Ruta"] = str(PakingSplip["cosyd_ruta"])
            Aux_Datos["Fecha de Ruta"] = str(PakingSplip["cosyd_ruta_fecha_hora"])


            Aux_Datos["Pre-Liberado"] = "" 
            for Hi in H:
                if Hi["cosyd_movimiento"] == "PRE_LIBERADOS":
                     Aux_Datos["Pre-Liberado"] = str(Hi["cosyd_fecha"])
            
            Aux_Datos["Liberado"] = "" 
            for Hi in H:
                if Hi["cosyd_movimiento"] == "LIBERADO":
                     Aux_Datos["Liberado"] = str(Hi["cosyd_fecha"])
            
            Aux_Datos["Cerrado"] = "" 
            for Hi in H:
                if Hi["cosyd_movimiento"] == "CERRADO":
                     Aux_Datos["Cerrado"] = str(Hi["cosyd_fecha"])


            if int(PakingSplip["cosyd_estado"]) == 0:
                Aux_Datos["Estado"] = "ELIMINADO"
            elif int(PakingSplip["cosyd_estado"]) == 1:
                Aux_Datos["Estado"] = "ABIERTO"
            elif int(PakingSplip["cosyd_estado"]) == 2:
                Aux_Datos["Estado"] = "LIBERADO"
            else:
                Aux_Datos["Estado"] = "CERRADO"
            Aux_Datos[" "] = Opciones
            Aux_Datos["Fecha de envío del proveedor"] = ""
            Aux_Datos["Proveedor"] = PakingSplip["cosyd_proveedor"]
            for P in Proveedores:
                if str(P["crilcpr_codigo"]) ==  str(PakingSplip["cosyd_proveedor"]):
                    Aux_Datos["Proveedor"] += " - " + str(P["crilcpr_nombre"])            
            Aux_Datos["PackingSlip"] = str(PakingSplip["cosyd_p_pakingslip"])
            Aux_Datos["Parte"] = str(PakingSplip["cosyd_parte"])
            Aux_Datos["Cantidad ASN"] = str(PakingSplip["cosyd_cantidad_asn"])
            Aux_Datos["Cantidad Real"] = str(PakingSplip["cosyd_cantidad_real"])
            Aux_Datos["OS&D"] = str(PakingSplip["cosyd"])
            Aux_Datos["Comentario Parte"] = str(PakingSplip["Comm_Parte"])

            if PakingSplip["cosyd"] != "":
                PakingSplip["cosyd_p_1_Damage"] = 'X' if str(PakingSplip["cosyd"]) == "1.- Damage" else ""
                PakingSplip["cosyd_p_2_Shortage"] = 'X' if str(PakingSplip["cosyd"]) == "2.- Surplus" else ""
                PakingSplip["cosyd_p_3_Surplus"] = 'X' if str(PakingSplip["cosyd"]) == "3.- Surplus" else ""
                PakingSplip["cosyd_p_4_ASN_Issue"] = 'X' if str(PakingSplip["cosyd"]) == "4.- ASN Issue" else ""
                PakingSplip["cosyd_p_5_Missing_doc_in_Prisma"] = 'X' if str(PakingSplip["cosyd"]) == "5.- Missing doc. in Prisma" else ""
                if PakingSplip["cosyd_p_1_Damage"] == "" and PakingSplip["cosyd_p_2_Shortage"] == "" and PakingSplip["cosyd_p_3_Surplus"] == "" and PakingSplip["cosyd_p_4_ASN_Issue"] == "" and PakingSplip["cosyd_p_5_Missing_doc_in_Prisma"] == "":
                    PakingSplip["cosyd_p_6_Other"] = "["+str(PakingSplip["cosyd"])+"] - " + str(PakingSplip["cosyd_comentario"])
            else:
                PakingSplip["cosyd_p_1_Damage"] = 'X' if int(PakingSplip["cosyd_p_1_Damage"]) == 1 else ""
                PakingSplip["cosyd_p_2_Shortage"] = 'X' if int(PakingSplip["cosyd_p_2_Shortage"]) == 1 else ""
                PakingSplip["cosyd_p_3_Surplus"] = 'X' if int(PakingSplip["cosyd_p_3_Surplus"]) == 1 else ""
                PakingSplip["cosyd_p_4_ASN_Issue"] = 'X' if int(PakingSplip["cosyd_p_4_ASN_Issue"]) == 1 else ""
                PakingSplip["cosyd_p_5_Missing_doc_in_Prisma"] = 'X' if int(PakingSplip["cosyd_p_5_Missing_doc_in_Prisma"]) == 1 else ""
                PakingSplip["cosyd_p_6_Other"] = str(PakingSplip["cosyd_p_6_Other"]) if PakingSplip["cosyd_p_6_Other"] is not None else ""

            Aux_Datos["1 Damage"] = str(PakingSplip["cosyd_p_1_Damage"])
            Aux_Datos["2 Shortage"] = str(PakingSplip["cosyd_p_2_Shortage"])
            Aux_Datos["3 Surplus"] = str(PakingSplip["cosyd_p_3_Surplus"])
            Aux_Datos["4 ASN Issue"] = str(PakingSplip["cosyd_p_4_ASN_Issue"])
            Aux_Datos["5 Missing doc in Prisma"] = str(PakingSplip["cosyd_p_5_Missing_doc_in_Prisma"])
            Aux_Datos["6 Other"] = str(PakingSplip["cosyd_p_6_Other"])

            
            if PakingSplip["cosyd_p_destino"] != "" and PakingSplip["cosyd_p_destino"] is not None:
                Aux_Datos["Destino"] = PakingSplip["cosyd_p_destino"]
            else:
                Aux_Datos["Destino"] = PakingSplip["cosyd_destino"]
            Aux_Datos["Contenedor"] = PakingSplip["cosyd_caja"]
            Aux_Datos["SCAC"] = PakingSplip["cosyd_scac"]
            Aux_Datos["Ultimo Comentario"] = PakingSplip["cosyd_comentario"]
            Aux_Datos["Numeros"] = str(Numeros)
            Aux_Datos["Archivos"] = ""
            for Archivo in str(PakingSplip["cosyd_archivos"]).split(","):
                if Archivo.strip() != "":
                    if "pdf" in Archivo:
                        Aux_Datos["Archivos"] += "<a href='http://10.4.7.219:8080/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-file-pdf-box ms-1'></a>"
                    else:
                        Aux_Datos["Archivos"] += "<a href='http://10.4.7.219:8080/Portal_File/"+str(Archivo)+"' target='_blank' class='mdi mdi-image ms-1'></a>"
            
            if PakingSplip["cosyd_baja"] is None or str(PakingSplip["cosyd_baja"]) == "":
                Aux_Datos["Cerrado"] = ""
            else:
                Aux_Datos["Cerrado"] = str(PakingSplip["cosyd_baja"])
            Aux_Datos["Tipo"] = str(PakingSplip["cosyd_tipo"])
            if int(PakingSplip["cosyd_estado"]) == 0: 
                Aux_Datos["Estado"] = 3
            Tabla_Datos_Cerrados.append(Aux_Datos)

        Resultado["Contenido"] += """
        <div class='w-100 d-flex justify-content-end'>
            <button class='btn btn-success mb-1' id="download-xlsx-cerrados"><i class='mdi mdi-microsoft-excel'></i> Descargar Excel</button>
        </div>
        <div id='Tabla-Cerrados'></div>
        <script>
            var Col = [
                {title: ' ', field: ' ', formatter: 'html', download: false,width:30}, 
                /*{title: 'Cerrado', field: 'Cerrado'},
                {title: 'Estado', field: 'Estado', headerFilter: 'input'},*/
                {title: 'Folio', field: 'Folio', headerFilter: 'input',width:80}, 
                {title: 'Ruta', field: 'Ruta', headerFilter: 'input',width:80}, 
                {title: 'Fecha de Ruta', field: 'Fecha de Ruta', headerFilter: 'input'}, 
                {title: 'Proveedor', field: 'Proveedor', headerFilter: 'input'}, 
                {title: 'PackingSlip', field: 'PackingSlip', headerFilter: 'input'}, 
                {title: 'Parte', field: 'Parte', headerFilter: 'input'},
                {title: 'Cantidad ASN', field: 'Cantidad ASN', headerFilter: 'input'},
                {title: 'Cantidad Real', field: 'Cantidad Real', headerFilter: 'input'},
                /*{title: 'OS&D', field: 'OS&D', headerFilter: 'input'},*/
        """
        for D in ["1 Damage","2 Shortage","3 Surplus","4 ASN Issue","5 Missing doc in Prisma"]:
            Resultado["Contenido"] += """{title:'"""+str(D)+"""', field:'"""+str(D)+"""', hozAlign:"center",width:100,headerSort:false,headerHozAlign:"center"},"""
        Resultado["Contenido"] += """
                {title:"6 Other (Explain in Comments)", field:"6 Other"},
                /*{title: 'Comentario Parte', field: 'Comentario Parte', headerFilter: 'input'},*/
                {title: 'Destino', field: 'Destino', headerFilter: 'input'}, 
                /*{title: 'Numeros de Parte (Total/Cerrados)', field: 'Numeros', headerFilter: 'input'},*/
                {title: 'Contenedor', field: 'Contenedor', headerFilter: 'input'}, 
                /*{title: 'SCAC', field: 'SCAC', headerFilter: 'input'},*/
                /*{title: 'Ultimo Comentario', field: 'Ultimo Comentario', headerFilter: 'input'},
                {title: 'Archivo(s)', field: 'Archivos', formatter: 'html',download: false}*/
                {title: 'Alta', field: 'Alta', headerFilter: 'input'}, 
                {title: 'Pre-Liberado', field: 'Pre-Liberado', headerFilter: 'input'},
                {title: 'Liberado', field: 'Liberado', headerFilter: 'input'}, 
                {title: 'Cerrado', field: 'Cerrado', headerFilter: 'input'}, 
                ];
            var table_c = new Tabulator("#Tabla-Cerrados", {
                minHeight:600,
                data:"""+str(Tabla_Datos_Cerrados)+""",
                initialSort:[
                    {column:"Fecha Alta", dir:"desc"}
                ],
                layout:"fitColumns",
                pagination:"local",
                rowFormatter:function(row){
                if(row.getData().Estado == 3){
                        row.getElement().style.backgroundColor = "#ff5d5d";
                    }
                },
                paginationSize:50,  
                paginationCounter:"rows",
                columns:Col,
                selectable:false
            });
            document.getElementById("download-xlsx-cerrados").addEventListener("click", function(){
                table_c.download("xlsx", "Cerrados.xlsx", {sheetName:"My Data"});
            });
        </script>

        """
    except:
        Resultado["Contenido"] += str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur

def Generar_Archivo(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = { "Estado" :0, "Contenido":""}
    try:
        wb = openpyxl.Workbook()
        sheet_obj = wb.create_sheet("Sheet1",0)
        a = 1
        i = 1
        for Partes in DB.Get_Dato("select PASTER.cosyd,MASTER.cosyd_id,MASTER.cosyd_destino,MASTER.cosyd_packingslip,MASTER.cosyd_caja,PASTER.cosyd_parte,PASTER.cosyd_cantidad_asn,MASTER.cosyd_alta,PASTER.cosyd_p_pallets,MASTER.cosyd_ruta,MASTER.cosyd_ruta_fecha_hora,PASTER.cosyd_p_pakingslip,PASTER.cosyd_comentario,PASTER.cosyd_p_destino,PASTER.cosyd_p_1_Damage,PASTER.cosyd_p_2_Shortage,PASTER.cosyd_p_3_Surpluse,PASTER.cosyd_p_4_ASN_Issue,PASTER.cosyd_p_5_Missing_doc_in_Prisma,PASTER.cosyd_p_6_Other from "+str(BD_Nombre)+".cosyd MASTER inner join "+str(BD_Nombre)+".cosyd_partes PASTER on PASTER.cosyd_master = MASTER.cosyd_id where MASTER.cosyd_id = '"+str(Datos["IDMaster"])+"' "):
            for Pallets in range(0,1):
                Partes["cosyd"] = ""
                Partes["cosyd"] += '1 Damage,' if int(Partes["cosyd_p_1_damage"]) == 1 else ""
                Partes["cosyd"] += '2 Shortage,' if int(Partes["cosyd_p_2_shortage"]) == 1 else ""
                Partes["cosyd"] += '3 Surplus,' if int(Partes["cosyd_p_3_surpluse"]) == 1 else ""
                Partes["cosyd"] += '4 ASN_Issue,' if int(Partes["cosyd_p_4_asn_issue"]) == 1 else ""
                Partes["cosyd"] += '5 Missing_doc_in_Prisma,' if int(Partes["cosyd_p_5_missing_doc_in_prisma"]) == 1 else ""
                Partes["cosyd"] += "6 Other ("+str(Partes["cosyd_p_6_other"])+")" if Partes["cosyd_p_6_other"] is not None else ""
                Partes["cosyd"] += " [DESTINO "+str(Partes["cosyd_p_destino"])+"]"
                Datos = {
                    "OS&D": Partes["cosyd"],
                    "Folio": str(Partes["cosyd_id"]).zfill(5),
                    "Paking Slip": str(Partes["cosyd_p_pakingslip"]),
                    "Caja":str(Partes["cosyd_caja"]),
                    "Parte": str(Partes["cosyd_parte"]),
                    "Cantidad ASN" : str(Partes["cosyd_cantidad_asn"]),
                    "Ruta / Fecha" : str(Partes["cosyd_ruta"]) + " / " + str(Partes["cosyd_ruta_fecha_hora"])
                }
            
                Formato(a,sheet_obj,Datos)
                if i%2:
                    a += 24
                else:
                    a += 20
                i += 1
        wb.save(str(current_app.root_path).replace("\\","/")+"/Files/Gen/"+"/OS&D.xlsx")
        Resultado["Archivo"] = "OS&D.xlsx"
        Resultado["Estado"] = 1
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Formato(renglon_inicia,Hoja,Datos):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    Hoja.cell(row = renglon_inicia, column = 1).value = "OS&D [ODC]"
    Hoja.merge_cells("A"+str(renglon_inicia)+":I"+str(renglon_inicia+2))
    Hoja.cell(row = renglon_inicia, column = 1).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
    Hoja.cell(row = renglon_inicia, column = 1).font = Font(size = "36",color="FFFFFF",bold=True)
    Hoja.cell(row = renglon_inicia, column = 1).fill = PatternFill(bgColor=colors.Color(indexed=32), fill_type = "solid")

    renglon_inicia+=3

    Hoja.cell(row = renglon_inicia, column = 1).value = str(Datos["OS&D"])
    Hoja.merge_cells("A"+str(renglon_inicia)+":I"+str(renglon_inicia+4))
    Hoja.cell(row = renglon_inicia, column = 1).font = Font(size = "16",bold=True)
    Hoja.cell(row = renglon_inicia, column = 1).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)

    for i in range(1,10):
        Hoja.cell(row = renglon_inicia, column = i).border = thin_border
    for i in range(1,10):
        Hoja.cell(row = renglon_inicia+1, column = i).border = thin_border
    for i in range(1,10):
        Hoja.cell(row = renglon_inicia+2, column = i).border = thin_border
    for i in range(1,10):
        Hoja.cell(row = renglon_inicia+3, column = i).border = thin_border
    for i in range(1,10):
        Hoja.cell(row = renglon_inicia+4, column = i).border = thin_border

    renglon_inicia+=5
    
    for K in Datos.keys():
        if "OS&D" != K:
            Hoja.cell(row = renglon_inicia, column = 1).value = str(K)
            Hoja.merge_cells("A"+str(renglon_inicia)+":C"+str(renglon_inicia+1))
            Hoja.cell(row = renglon_inicia, column = 1).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            Hoja.cell(row = renglon_inicia, column = 1).font = Font(size = "16",color="555555",bold=True)
            Hoja.cell(row = renglon_inicia, column = 4).value = str(Datos[K])
            Hoja.merge_cells("D"+str(renglon_inicia)+":I"+str(renglon_inicia+1))
            Hoja.cell(row = renglon_inicia, column = 4).alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
            Hoja.cell(row = renglon_inicia, column = 4).font = Font(size = "16",bold=True)
            for i in range(1,10):
                Hoja.cell(row = renglon_inicia, column = i).border = thin_border
            for i in range(1,10):
                Hoja.cell(row = renglon_inicia+1, column = i).border = thin_border

            renglon_inicia+=2


def Direccionar(Datos):
    try:
        if "K" in session.keys():
            fernet = Fernet(session["K"])
        else:
            fernet = Fernet(LibDM_2023.Compartido().Dame_K2())
        if Datos is None:
            return Inicio()
        else:
            if "Fun" in Datos:
                Funcion = fernet.decrypt(Datos["Fun"].encode()).decode()
                Par = {}
                for P in Datos.keys():
                    if str(P) != "Fun":
                        Par[str(P)] = Datos[str(P)]
                if "IDu" in session.keys():
                    Par["ID_User"] = session["IDu"]
                return globals()[Funcion](Par)
            else:
                return Inicio()
    except:
        return str(sys.exc_info())
