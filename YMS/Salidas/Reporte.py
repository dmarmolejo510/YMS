from flask import request,session,render_template,current_app
from cryptography.fernet import Fernet
from datetime import datetime,date,timedelta
import sys
import json
import os
import openpyxl
from Componentes import LibDM_2023
Url = ""
fernet = Fernet(LibDM_2023.Compartido().Dame_K2())
BD_Nombre = "public"
Bandera_Dock = "YMS"
def Inicio():
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    Cur = ""
    try:
        Activo = "YMS"
        Compartido = LibDM_2023.Compartido()
        Menu = LibDM_2023.Menu().Menu(Activo,request.url_root,session["IDu"])
        Titulo = LibDM_2023.Menu().Get_Titulo(Activo)
        Contenido = ""

        Contenido += "<div class='h2 fw-lighter mt-2 mb-2 pb-2 text-center'><i class='mdi mdi-progress-upload'></i> Reporte de Estado de Outbound</div>"
        #Rutas = DB.Get_Dato("SELECT * FROM odc_slp.crutas WHERE cr_tipo = 'Outbound'")
        Contenido += """
        <div class='container'>
        """
        Formulario = {"Col":"12", "Campos": [],"Clase": "Estado_Cajas" }
        Formulario["Campos"].append({"tipo":"fecha-rango","id":"fecha","campo":"Fecha","titulo":"Rango de Fecha","Requerido":1,"Col":12,"min":1,"max":10,"valor":""})
        Contenido += str(Compartido.Formulario(Formulario))

        Contenido += """
        <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Cargar_Fecha();'><i class='mdi mdi-cloud-download'></i> Descargar</button></div>
        <div id='Res'>
        </div>
        <script>
            function Cargar_Fecha(){
                var Info = Dame_Formulario(".Estado_Cajas",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Cargar_Fecha".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info)};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            $("#Res").html(Resultado["Contenido"]);
                            swal.close();
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
            }            
        </script>
        </div>
        """

        Cur += render_template("general.html",Contenido=Contenido,Componentes=Compartido.Complementos(None),Menu=Menu,Titulo=Titulo)
    except:
        Cur += str(sys.exc_info())
    return Cur
def Cargar_Fecha(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Info_Datos = json.loads(str(Datos["Info"]))
        Resultado["Contenido"] += "<div class='h2 fw-lighter mt-2 mb-2 pb-2 text-center'>"+str(Info_Datos["Fecha"][0])+" to "+str(Info_Datos["Fecha"][1])+"</div>"
        Historico = []
        Cajas_Aqui = []

        Fecha_I = datetime.strptime(str(Info_Datos["Fecha"][0]), "%Y-%m-%d")
        Fecha_F = datetime.strptime(str(Info_Datos["Fecha"][1]), "%Y-%m-%d")
        lista_fechas = [Fecha_I + timedelta(days=d) for d in range((Fecha_F - Fecha_I).days + 1)]

        Likes = []
        for F in lista_fechas:
            Likes.append(" cc_informacion_actual LIKE '%"+str(F.strftime("%Y-%m-%d"))+"%' ")

        Programacion_Abiertos = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE " + str("OR".join(Likes)) )

        for Programacion in Programacion_Abiertos:
            if Programacion["cc_id"] is not None:
                Cajas_Aqui.append(str(Programacion["cc_id"]))

        Movimientos = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas_moviemiento WHERE cch_master IN ("+str(",".join(Cajas_Aqui))+") ORDER BY cch_fecha_hora")

        for Programacion in Programacion_Abiertos:
            Info_Ahora = json.loads(str(Programacion["cc_informacion_actual"]))
            Info_T = {}
            Info_T["Fecha"] = str(Info_Ahora["Fecha_Salida"])
            Ruta = ""
            try:
                Tipo = Info_Ahora[Programacion["cc_tipo_actual"]]
                Ruta += str(Tipo)
                Tipo_1 = Info_Ahora[Tipo]
                Ruta += "/"+str(Tipo_1)
                Tipo_2 = Info_Ahora[Tipo_1]
                Ruta += "/"+str(Tipo_2)
            except:
                pass
            Info_T["Ruta"] = str(Ruta)

            Info_T["Caja"] = str(Programacion["cc_contenedor"])
            Info_T["Carrier"] = str(Info_Ahora["Carrier"])
            Info_T["Entrada"] = str(Programacion["cc_fecha_hora"])

            Movimientos_Aqui = []
            for M in Movimientos:
                if int(M["cch_master"]) == int(Programacion["cc_id"]):
                    Movimientos_Aqui.append(M)
            
            Definir_Ruta = ""
            for M in Movimientos_Aqui:
                if "Fecha_Salida" in str(M["cch_informacion_actual"]):
                    Definir_Ruta = M["cch_fecha_hora"].strftime("%m-%d %H:%M")

            Info_T["DEFINIR RUTA"] = str(Definir_Ruta)

            Libera_Operaciones = ""
            for M in Movimientos_Aqui:
                if M["cch_movimiento"] == "LIBERA OPERACIONES":
                    Libera_Operaciones = M["cch_fecha_hora"].strftime("%m-%d %H:%M")
            Info_T["OPERACIONES"] = str(Libera_Operaciones)

            Salida = ""
            for M in Movimientos_Aqui:
                if M["cch_movimiento"] == "SALIDA":
                    Salida = M["cch_fecha_hora"].strftime("%m-%d %H:%M")
                if M["cch_movimiento"] == "ELIMINADO POR YARD":
                    Salida = M["cch_fecha_hora"].strftime("%m-%d %H:%M")
            
            Info_T["SALIO"] = str(Salida)
            Historico.append(Info_T)

            
        Resultado["Contenido"] += """
        <button class='btn btn-success mb-1' id="download-xlsx-ruteo"><i class='mdi mdi-microsoft-excel'></i> Descargar Excel</button>
        <button class='btn btn-dark mb-1' onclick='Descargar_Reporte()'><i class='mdi mdi-microsoft-excel'></i> Descargar Reporte</button>
        <div id='Tabla_Historico' class='border border-dark bg-dark-subtle'></div>
        <script>
            delete Tabla_Historico;
            var Tabla_Historico = new Tabulator("#Tabla_Historico", {
                minHeight:500,
                layout:"fitColumns",
                data:"""+str(Historico)+""",
                columns:[
                    {field:"Fecha","title":"Fecha de Salida"},
                    {field:"Ruta","title":"Ruta"},
                    {field:"Caja","title":"Caja"},
                    {field:"Carrier","title":"Carrier"},
                    {field:"Entrada","title":"Entrada"},
                    {field:"DEFINIR RUTA","title":"Define Ruta"},
                    {field:"OPERACIONES","title":"Material Cargado"},
                    {field:"SALIO","title":"Salida"}
                ]
            });
            Tabla_Historico.on("tableBuilt", function(){ 
                setTimeout(() => {
                    Tabla_Historico.clearFilter();
                    Tabla_Historico.redraw();
                }, 100);
            });
            document.getElementById("download-xlsx-ruteo").addEventListener("click", function(){
                Tabla_Historico.download("xlsx", "Historia.xlsx", {sheetName:"My Data"});
            });
        </script>
        
        """
        Resultado["Contenido"] += """
        <script>
            function Descargar_Reporte(){
                var Info = Dame_Formulario(".Estado_Cajas",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Descargar_Reporte".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info)};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1){
                                if("Pase Salida" in Resultado)
                                {
                                    var win1 = window.open('"""+str(request.url_root)+"""/Portal_File/Gen/'+Resultado["Pase Salida"], '_blank');
                                    win1.focus();
                                }
                                if("Manifiesto" in Resultado)
                                {
                                    var win1 = window.open('"""+str(request.url_root)+"""/Portal_File/Gen/'+Resultado["Manifiesto"], '_blank');
                                    win1.focus();
                                }
                                Mensaje(2);
                            }else{
                                Mensaje(0,Resultado["Contenido"]);
                            }
                            
                        },
                        error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,textStatus);}
                    });
                }
            }

        </script>
        </div>
        """
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Descargar_Reporte(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Error = ""
        Nombre_Archivo = "Reporte_Out"
        wb_obj = openpyxl.load_workbook(str(current_app.root_path).replace("\\","/")+"/Formatos/Reporte_E_N.xlsx")
        sheet_obj = wb_obj.worksheets[0]
        Info_Datos = json.loads(str(Datos["Info"]))
        sheet_obj.cell(row = 2, column = 3).value = str(datetime.strptime(Info_Datos["Fecha"][0],"%Y-%m-%d").strftime("%b %d"))  + " - " + str(datetime.strptime(Info_Datos["Fecha"][1],"%Y-%m-%d").strftime("%b %d"))
        Info_Datos = json.loads(str(Datos["Info"]))
        Resultado["Contenido"] += "<div class='h2 fw-lighter mt-2 mb-2 pb-2 text-center'>"+str(Info_Datos["Fecha"][0])+" to "+str(Info_Datos["Fecha"][1])+"</div>"
        Historico = []
        Cajas_Aqui = []
        
        Fecha_I = datetime.strptime(str(Info_Datos["Fecha"][0]), "%Y-%m-%d")
        Fecha_F = datetime.strptime(str(Info_Datos["Fecha"][1]), "%Y-%m-%d")
        lista_fechas = [Fecha_I + timedelta(days=d) for d in range((Fecha_F - Fecha_I).days + 1)]

        Likes = []
        for F in lista_fechas:
            Likes.append(" cc_informacion_actual LIKE '%"+str(F.strftime("%Y-%m-%d"))+"%' ")

        Programacion_Abiertos = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE " + str("OR".join(Likes)) )

        for Programacion in Programacion_Abiertos:
            if Programacion["cc_id"] is not None:
                Cajas_Aqui.append(str(Programacion["cc_id"]))

        Movimientos = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas_moviemiento WHERE cch_master IN ("+str(",".join(Cajas_Aqui))+") ORDER BY cch_fecha_hora")

        for Programacion in Programacion_Abiertos:
            Info_Ahora = json.loads(str(Programacion["cc_informacion_actual"]))
            Info_T = {}
            Info_T["Fecha"] = str(Info_Ahora["Fecha_Salida"])
            Ruta = ""
            try:
                Tipo = Info_Ahora[Programacion["cc_tipo_actual"]]
                Ruta += str(Tipo)
                Tipo_1 = Info_Ahora[Tipo]
                Ruta += "/"+str(Tipo_1)
                Tipo_2 = Info_Ahora[Tipo_1]
                Ruta += "/"+str(Tipo_2)
            except:
                pass
            Info_T["Ruta"] = str(Ruta)

            Info_T["Caja"] = str(Programacion["cc_contenedor"])
            Info_T["Carrier"] = str(Info_Ahora["Carrier"])
            Info_T["Entrada"] = str(Programacion["cc_fecha_hora"].strftime("%m-%d %H:%M"))

            Movimientos_Aqui = []
            for M in Movimientos:
                if int(M["cch_master"]) == int(Programacion["cc_id"]):
                    Movimientos_Aqui.append(M)
            
            Definir_Ruta = ""
            for M in Movimientos_Aqui:
                if "Fecha_Salida" in str(M["cch_informacion_actual"]):
                    Definir_Ruta = M["cch_fecha_hora"].strftime("%m-%d %H:%M")

            Info_T["DEFINIR RUTA"] = str(Definir_Ruta)

            Libera_Operaciones = ""
            for M in Movimientos_Aqui:
                if M["cch_movimiento"] == "LIBERA OPERACIONES":
                    Libera_Operaciones = M["cch_fecha_hora"].strftime("%m-%d %H:%M")
            Info_T["OPERACIONES"] = str(Libera_Operaciones)

            Salida = ""
            for M in Movimientos_Aqui:
                if M["cch_movimiento"] == "SALIDA":
                    Salida = M["cch_fecha_hora"].strftime("%m-%d %H:%M")
                if M["cch_movimiento"] == "ELIMINADO POR YARD":
                    Salida = M["cch_fecha_hora"].strftime("%m-%d %H:%M")
            
            Info_T["SALIO"] = str(Salida)
            Historico.append(Info_T)
        Row = 5
        
        for H in Historico:
            try:
                sheet_obj.cell(row = Row, column = 1).value = H["Fecha"]
                sheet_obj.cell(row = Row, column = 2).value = H["Ruta"]
                sheet_obj.cell(row = Row, column = 3).value = H["Caja"]
                sheet_obj.cell(row = Row, column = 4).value = H["Carrier"]
                sheet_obj.cell(row = Row, column = 5).value = H["Entrada"]
                sheet_obj.cell(row = Row, column = 6).value = H["DEFINIR RUTA"]
                sheet_obj.cell(row = Row, column = 7).value = H["OPERACIONES"]
                sheet_obj.cell(row = Row, column = 8).value = H["SALIO"]
                Row += 1
            except:
                pass

        wb_obj.save(str(current_app.root_path).replace("\\","/")+"/Files/Gen/"+str(Nombre_Archivo)+".xlsx")
        Resultado["Pase Salida"] = str(Nombre_Archivo)+".xlsx"
        if Error == "":
            Resultado["Estado"] = 1
        else:
            Resultado["Contenido"] += Error
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur

def Direccionar(Datos):
    try:
        if "K" in session.keys():
            fernet = Fernet(session["K"])
        else:
            fernet = Fernet(LibDM_2023.Compartido().Dame_K2())
        if Datos is None:
            return Inicio()
        else:
            if "Fun" in Datos:
                Funcion = fernet.decrypt(Datos["Fun"].encode()).decode()
                Par = {}
                for P in Datos.keys():
                    if str(P) != "Fun":
                        Par[str(P)] = Datos[str(P)]
                if "IDu" in session.keys():
                    Par["ID_User"] = session["IDu"]
                return globals()[Funcion](Par)
            else:
                return Inicio()
    except:
        return str(sys.exc_info())
