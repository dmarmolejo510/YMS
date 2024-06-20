from flask import request,session,render_template,current_app
from cryptography.fernet import Fernet
from datetime import datetime,date,timedelta
import sys
import json
import os
import openpyxl
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from Componentes import LibDM_2023
BD_Nombre = LibDM_2023.Compartido().Dame_Base_Datos("YMS")
Bandera_Dock = "YMS"
fernet = Fernet(LibDM_2023.Compartido().Dame_K2())
def Inicio_OLD():
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    Cur = ""
    try:
        DB = LibDM_2023.DataBase()
        Activo = "YMS"
        Compartido = LibDM_2023.Compartido()
        Menu = LibDM_2023.Menu().Menu(Activo,request.url_root,session["IDu"])
        Titulo = LibDM_2023.Menu().Get_Titulo(Activo)
        Contenido = ""
        Rutas = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo != 'Empty'")
        Patio = []
        for Ruta in Rutas:
            Opciones = "<div class='text-center'><div class='btn-group' role='group' aria-label='Basic mixed styles example'>"
            Opciones += "</div></div>"
            Patio.append({"Tipo":str(Ruta["cr_tipo"]),"Opciones":Opciones})


        Contenido += "<div class='container'><div class='h2 fw-lighter mt-1 mb-1 text-center border-bottom'><i class='mdi mdi-card-bulleted'></i> Route Master</div>"
        for Ruta in Rutas:
            Contenido += """
            <div class='w-100 border border-dark bg-body mb-2 p-3'>
                <div class='fs-1 text-center'>"""+str(Ruta["cr_tipo"])+""" <a class='link-primary' onclick='Agregar_Nivel("N1",\""""+str(Ruta["cr_tipo"])+"""\")' style='cursor:pointer'><i class='mdi mdi-plus'></i></a></div>
            """
            if Ruta["cr_niveles"] is not None and str(Ruta["cr_niveles"]) != "":
                Contenido += "<div class='row'>"
                Niveles = json.loads(Ruta["cr_niveles"]) 
                for N1 in Niveles.keys():
                    Tiene = False
                    Contenido += "<div class='col p-1'><div class='border border-dark bg-body-tertiary'>"
                    Contenido += "<div class='text-center fs-2'>"+str(N1)+" <a class='link-warning' style='cursor:pointer' onclick='Agregar_Nivel(\"N2\",\""+str(Ruta["cr_tipo"])+"/"+str(N1)+"\")' ><i class='mdi mdi-pencil'></i></a> <a class='link-danger' style='cursor:pointer' onclick='Eliminar_Nivel(\"N2\",\""+str(Ruta["cr_tipo"])+"/"+str(N1)+"\")'><i class='mdi mdi-trash-can'></i></a></div>"
                    try:
                        Keys = Niveles[N1].keys()
                        Contenido += "<div class='row p-3'>"
                        for N2 in Keys:
                            Contenido += "<div class='col p-1'><div class='border border-dark bg-dark-subtle'>"
                            Contenido += "<div class='text-center fs-3'>"+str(N2)+"</div>"
                            try:
                                Docks = Niveles[N1][N2]["DOCKS"]
                                Contenido += "<div class='row p-3'>"
                                for N3 in Docks:
                                    Contenido += "<div class='col p-1'><div class='border border-dark bg-warning'>"
                                    Contenido += "<div class='text-center fs-4'>"+str(N3)+"</div>"
                                    Contenido += "</div></div>"
                                Contenido += "</div>"
                            except:
                                pass

                            Contenido += "</div></div>"
                        Contenido += "</div>"
                    except:
                        pass
                    Contenido += "</div></div>"
                Contenido += "</div>"
            Contenido += """
            </div>
            """
        Contenido += "</div>"
        Contenido += """
        <script>
            function Agregar_Nivel(Nivel,Direccion){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-plus'></i>" + Direccion );
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm');
                var parametros = {"Fun":'"""+str(fernet.encrypt("Agregar_Nivel".encode()).decode("utf-8"))+"""',"Nivel":Nivel,"Direccion":Direccion};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });
            }
            function Eliminar_Nivel(Nivel,Direccion){
                Swal.fire({
                title: 'Are you sure delete ['+Direccion+']?',
                buttonsStyling: false,showCancelButton: true,confirmButtonText: "<i class='mdi mdi-check'></i> Yes",cancelButtonText: "<i class='mdi mdi-close'></i> No",showLoaderOnConfirm: true,
                customClass: {confirmButton: 'btn btn-success ms-1 me-1',cancelButton: 'btn btn-danger ms-1 me-1'},
                preConfirm: () => {
                
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Eliminar_Nivel".encode()).decode("utf-8"))+"""',"Nivel":Nivel,"Direccion":Direccion};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                Mensaje(2);
                                Llamar_Funcion(\""""+str(request.url)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                    
                }
                })
            }
        </script>
        """
        Cur += render_template("general.html",Contenido=Contenido,Componentes=Compartido.Complementos(None),Menu=Menu,Titulo=Titulo)
    except:
        Cur += str(sys.exc_info())
    return Cur



def Inicio():
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Rutas = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo != 'Empty'")

        Concentrado = {"Nacional":{},"Internacional":{},"MilkRuns":{}}
        for Ruta in Rutas:
            Niveles = json.loads(Ruta["cr_niveles"])
            if "Nacional" in Niveles.keys():
                for R in Niveles["Nacional"].keys():
                    if R in Concentrado["Nacional"].keys():
                        Concentrado["Nacional"][R][Ruta["cr_tipo"]] = 1
                    else:
                        for k in Niveles["Nacional"][R].keys():
                            if Niveles["Nacional"][R][k] is None:
                                Niveles["Nacional"][R][k] = ""
                        Concentrado["Nacional"][R] = {}
                        Concentrado["Nacional"][R][Ruta["cr_tipo"]] = 1
                        Concentrado["Nacional"][R]["NOMBRE"] = Niveles["Nacional"][R]["NOMBRE"]
                        Concentrado["Nacional"][R]["ZONA"] = Niveles["Nacional"][R]["ZONA"]
                        Concentrado["Nacional"][R]["SCAC"] = Niveles["Nacional"][R]["SCAC"]
                        Concentrado["Nacional"][R]["DOCKS"] = Niveles["Nacional"][R]["DOCKS"]
                        Concentrado["Nacional"][R]["RUTAS"] = Niveles["Nacional"][R]["RUTAS"]
                        Concentrado["Nacional"][R]["DL"] =  Niveles["Nacional"][R]["DL"]
                        if "DESTINO" in Niveles["Nacional"][R].keys():
                            Concentrado["Nacional"][R]["DESTINO"] =  Niveles["Nacional"][R]["DESTINO"]
                        else:
                            Concentrado["Nacional"][R]["DESTINO"] = ""
                        if "CUTTIME_IN" in Niveles["Nacional"][R].keys():
                            Concentrado["Nacional"][R]["CUTTIME_IN"] =  Niveles["Nacional"][R]["CUTTIME_IN"]
                        else:
                            Concentrado["Nacional"][R]["CUTTIME_IN"] = ""
                        if "CUTTIME_OUT" in Niveles["Nacional"][R].keys():
                            Concentrado["Nacional"][R]["CUTTIME_OUT"] =  Niveles["Nacional"][R]["CUTTIME_OUT"]
                        else:
                            Concentrado["Nacional"][R]["CUTTIME_OUT"] = ""
            if "Importacion" in Niveles.keys():
                for R in Niveles["Importacion"].keys():
                    if R in Concentrado["Internacional"].keys():
                        Concentrado["Internacional"][R][Ruta["cr_tipo"]] = 1
                    else:
                        for k in Niveles["Importacion"][R].keys():
                            if Niveles["Importacion"][R][k] is None:
                                Niveles["Importacion"][R][k] = ""
                        Concentrado["Internacional"][R] = {}
                        Concentrado["Internacional"][R][Ruta["cr_tipo"]] = 1
                        Concentrado["Internacional"][R]["NOMBRE"] = Niveles["Importacion"][R]["NOMBRE"]
                        Concentrado["Internacional"][R]["ZONA"] = Niveles["Importacion"][R]["ZONA"]
                        Concentrado["Internacional"][R]["SCAC"] = Niveles["Importacion"][R]["SCAC"]
                        Concentrado["Internacional"][R]["DOCKS"] = Niveles["Importacion"][R]["DOCKS"]
                        Concentrado["Internacional"][R]["RUTAS"] = Niveles["Importacion"][R]["RUTAS"]
                        Concentrado["Internacional"][R]["DL"] =  Niveles["Importacion"][R]["DL"]
                        if "DESTINO" in Niveles["Importacion"][R].keys():
                            Concentrado["Internacional"][R]["DESTINO"] =  Niveles["Importacion"][R]["DESTINO"]
                        else:
                            Concentrado["Internacional"][R]["DESTINO"] = ""
                        if "CUTTIME_IN" in Niveles["Importacion"][R].keys():
                            Concentrado["Internacional"][R]["CUTTIME_IN"] =  Niveles["Importacion"][R]["CUTTIME_IN"]
                        else:
                            Concentrado["Internacional"][R]["CUTTIME_IN"] = ""
                        if "CUTTIME_OUT" in Niveles["Importacion"][R].keys():
                            Concentrado["Internacional"][R]["CUTTIME_OUT"] =  Niveles["Importacion"][R]["CUTTIME_OUT"]
                        else:
                            Concentrado["Internacional"][R]["CUTTIME_OUT"] = ""          
            if "Exportacion" in Niveles.keys():
                for R in Niveles["Exportacion"].keys():
                    if R in Concentrado["Internacional"].keys():
                        Concentrado["Internacional"][R][Ruta["cr_tipo"]] = 1
                    else:
                        for k in Niveles["Exportacion"][R].keys():
                            if Niveles["Exportacion"][R][k] is None:
                                Niveles["Exportacion"][R][k] = ""
                        Concentrado["Internacional"][R] = {}
                        Concentrado["Internacional"][R][Ruta["cr_tipo"]] = 1
                        Concentrado["Internacional"][R]["NOMBRE"] = Niveles["Exportacion"][R]["NOMBRE"]
                        Concentrado["Internacional"][R]["ZONA"] = Niveles["Exportacion"][R]["ZONA"]
                        Concentrado["Internacional"][R]["SCAC"] = Niveles["Exportacion"][R]["SCAC"]
                        Concentrado["Internacional"][R]["DOCKS"] = Niveles["Exportacion"][R]["DOCKS"]
                        Concentrado["Internacional"][R]["RUTAS"] = Niveles["Exportacion"][R]["RUTAS"]
                        Concentrado["Internacional"][R]["DL"] =  Niveles["Exportacion"][R]["DL"]
                        if "DESTINO" in Niveles["Exportacion"][R].keys():
                            Concentrado["Internacional"][R]["DESTINO"] =  Niveles["Exportacion"][R]["DESTINO"]
                        else:
                            Concentrado["Internacional"][R]["DESTINO"] = ""
                        if "CUTTIME_IN" in Niveles["Exportacion"][R].keys():
                            Concentrado["Internacional"][R]["CUTTIME_IN"] =  Niveles["Exportacion"][R]["CUTTIME_IN"]
                        else:
                            Concentrado["Internacional"][R]["CUTTIME_IN"] = ""
                        if "CUTTIME_OUT" in Niveles["Exportacion"][R].keys():
                            Concentrado["Internacional"][R]["CUTTIME_OUT"] =  Niveles["Exportacion"][R]["CUTTIME_OUT"]
                        else:
                            Concentrado["Internacional"][R]["CUTTIME_OUT"] = ""
            if "MilkRun" in Niveles.keys():
                for R in Niveles["MilkRun"].keys():
                    if Niveles["MilkRun"][R]["UNION"] in Concentrado["MilkRuns"].keys():
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]][Ruta["cr_tipo"]] = 1
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["DIAS"][R] = Niveles["MilkRun"][R]["RUTAS"]
                        pass
                    else:
                        for k in Niveles["MilkRun"][R].keys():
                            if Niveles["MilkRun"][R][k] is None:
                                Niveles["MilkRun"][R][k] = ""
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]] = {}
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]][Ruta["cr_tipo"]] = 1
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["NOMBRE"] = Niveles["MilkRun"][R]["NOMBRE"]
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["ZONA"] = Niveles["MilkRun"][R]["ZONA"]
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["SCAC"] = Niveles["MilkRun"][R]["SCAC"]
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["DOCKS"] = Niveles["MilkRun"][R]["DOCKS"]
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["DIAS"] = {}
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["DIAS"][R] = Niveles["MilkRun"][R]["RUTAS"]
                        Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["DL"] =  Niveles["MilkRun"][R]["DL"]
                        if "DESTINO" in Niveles["MilkRun"][R].keys():
                            Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["DESTINO"] =  Niveles["MilkRun"][R]["DESTINO"]
                        else:
                           Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["DESTINO"] = ""
                        if "CUTTIME_IN" in Niveles["MilkRun"][R].keys():
                            Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["CUTTIME_IN"] =  Niveles["MilkRun"][R]["CUTTIME_IN"]
                        else:
                            Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["CUTTIME_IN"] = ""
                        if "CUTTIME_OUT" in Niveles["MilkRun"][R].keys():
                            Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["CUTTIME_OUT"] =  Niveles["MilkRun"][R]["CUTTIME_OUT"]
                        else:
                            Concentrado["MilkRuns"][Niveles["MilkRun"][R]["UNION"]]["CUTTIME_OUT"] = ""

        Cur += "<div class='container'><div class='h2 fw-lighter mt-1 mb-1 text-center border-bottom'><i class='mdi mdi-card-bulleted'></i> Route Master</div>"
        Informacion = []
        for Tipo in Concentrado.keys():
            aux = {"Tipo":str(Tipo),"_children":[]}
            for R in Concentrado[Tipo].keys():
                idRuta = str(R).replace(" ","")
                Ruta = ""
                if "DIAS" in Concentrado[Tipo][R].keys():
                    Ruta = "<div>Días de la semana de recolección</div>"
                    Dias = {"M":"Luneas","T":"Martes","W":"Miercoles","R":"Jueves","F":"Viernes"}
                    Ruta += """
                    <table class='table table-sm table-bordered'>
                    <thead class='table-dark'>
                        <tr>
                    """
                    for D in Dias:
                        Ruta += "<th class='text-center'>"+str(R)+str(D)+"</th>"
                    Ruta += "</tr></thead><tbody><tr>"
                    index = 0
                    for D in Dias:
                        Ruta += "<td>"
                        Aqui = None
                        for k in Concentrado[Tipo][R]["DIAS"]:
                            if str(k) == str(R)+str(D):
                                Aqui = Concentrado[Tipo][R]["DIAS"][k]
                                break
                        if Aqui is not None:
                            if index == 0:
                                Ruta += """<div><a onclick='Agregar_item_DUNS(\"DUNS_"""+str(str(R)+str(D))+"""\")' class='link-success' style='cursor:pointer;'><i class='mdi mdi-plus'></i> Nuevo</a><a onclick='Duplicar_item_DUNS(\"DUNS_"""+str(str(R)+str(D)[:-1])+"""\")' class='link-primary' style='cursor:pointer;'><i class='mdi mdi-content-copy'></i> Duplicar ></a></div>"""
                            else:
                                Ruta += """<div><a onclick='Agregar_item_DUNS(\"DUNS_"""+str(str(R)+str(D))+"""\")' class='link-success' style='cursor:pointer;'><i class='mdi mdi-plus'></i> Nuevo</a></div>"""
                            Ruta += """
                            <ol class='list-group list-group-numbered Lista-DUNS' rutaduns='DUNS_"""+str(str(R)+str(D))+"""' ruta='"""+str(idRuta)+"""'>
                            """
                            for DUNS in Aqui:
                                Ruta += "<li class='list-group-item d-flex justify-content-between align-items-start p-0 m-0 ps-1 pe-1' valor='"+str(DUNS["DUNS"]).strip()+"'><div class='ms-2 me-auto'>"+str(DUNS["DUNS"]).strip()+"</div><span style='cursor:pointer' onclick='$(this).parent().remove(); Cambio_Algo(\""+str(idRuta)+"\");' class='badge bg-danger rounded-pill'><i class='mdi mdi-trash-can'></i></span></li>"
                            Ruta += "</ol>"
                        Ruta += "</td>"
                        index += 1
                    Ruta += "</tr></tbody></table>"
                    Ruta += "<hr>"
                Configuracion = ""
                Configuracion += """
                    <div class='p-3 bg-warning-subtle Contenedor_Configuracion'>
                        <div class='Estado fw-bold' ruta='"""+str(idRuta)+"""'>|</div>
                        <div class='fw-bold'>La ruta es:</div>
                        <div class="form-check form-check-inline">
                """
                if "Inbound" in Concentrado[Tipo][R].keys():
                    Configuracion += """<input checked class="form-check-input Fomulario_Ruta" campo='Inbound' type="checkbox" ruta='"""+str(idRuta)+"""' id='"""+str(idRuta)+"""_1' value="option1" onclick='Cambio_Algo(\""""+str(idRuta)+"""\")'>"""
                else:
                    Configuracion += """<input class="form-check-input Fomulario_Ruta" campo='Inbound' type="checkbox" ruta='"""+str(idRuta)+"""' id='"""+str(idRuta)+"""_1' value="option1" onclick='Cambio_Algo(\""""+str(idRuta)+"""\")'>"""
                Configuracion += """
                            <label class="form-check-label" for='"""+str(idRuta)+"""_1'>Inbound</label>
                        </div>
                        <div class="form-check form-check-inline">
                """
                if "Outbound" in Concentrado[Tipo][R].keys():
                    Configuracion += """<input checked class="form-check-input Fomulario_Ruta" campo='Outbound' ruta='"""+str(idRuta)+"""' type="checkbox" id='"""+str(idRuta)+"""_2' value="option2" onclick='Cambio_Algo(\""""+str(idRuta)+"""\")'>"""
                else:
                    Configuracion += """<input class="form-check-input Fomulario_Ruta" campo='Outbound' type="checkbox" ruta='"""+str(idRuta)+"""' id='"""+str(idRuta)+"""_2' value="option2" onclick='Cambio_Algo(\""""+str(idRuta)+"""\")'>"""
                Configuracion += """
                            <label class="form-check-label" for='"""+str(idRuta)+"""_2'>Outbound</label>
                        </div>
                        <hr>
                        <div class='fw-bold'>información general</div>
                        <div class='row'>
                                <div class='col-4'>
                                    <div class="mb-3">
                                        <label class="form-label">Nombre</label>
                                        <input type="text" class="form-control Fomulario_Ruta" campo='NOMBRE' ruta='"""+str(idRuta)+"""' value='"""+str(Concentrado[Tipo][R]["NOMBRE"])+"""' onchange='Cambio_Algo(\""""+str(idRuta)+"""\")'>
                                    </div>
                                </div>
                                <div class='col-4'>
                                    <div class="mb-3">
                                        <label class="form-label">Zona</label>
                                        <input type="text" class="form-control Fomulario_Ruta" campo='ZONA' ruta='"""+str(idRuta)+"""' value='"""+str(Concentrado[Tipo][R]["ZONA"])+"""' onchange='Cambio_Algo(\""""+str(idRuta)+"""\")'>
                                    </div>
                                </div>
                                <div class='col-4'>
                                    <div class="mb-3">
                                        <label class="form-label">SCAC</label>
                                        <input type="text" class="form-control Fomulario_Ruta" campo='SCAC' ruta='"""+str(idRuta)+"""' value='"""+str(Concentrado[Tipo][R]["SCAC"])+"""' onchange='Cambio_Algo(\""""+str(idRuta)+"""\")'>
                                    </div>
                                </div>
                                <div class='col-4'>
                                    <div class="mb-3">
                                        <label class="form-label">Destino</label>
                                        <select class="form-select Fomulario_Ruta" campo='DESTINO' ruta='"""+str(idRuta)+"""' onchange='Cambio_Algo(\""""+str(idRuta)+"""\")'>
                """
                Destino_Ahora = ""
                if "DESTINO" in Concentrado[Tipo][R].keys() and Concentrado[Tipo][R]["DESTINO"] != "":
                    Destino_Ahora = str(Concentrado[Tipo][R]["DESTINO"])
                if Destino_Ahora == "":
                    Configuracion += "<option selected value=''></option>"
                else:
                    Configuracion += "<option value=''></option>"
                for Destino in DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cdestinos WHERE cactivo = 1"):
                    if Destino_Ahora == str(Destino["cdestino"]):
                        Configuracion += "<option selected value='"+str(Destino["cdestino"])+"'>"+str(Destino["cnombre"])+" ["+str(Destino["ctipo"])+"]</option>"
                    else:
                        Configuracion += "<option value='"+str(Destino["cdestino"])+"'>"+str(Destino["cnombre"])+" ["+str(Destino["ctipo"])+"]</option>"
                Configuracion += """
                                        </select>
                                    </div>
                                </div>
                                <div class='col-4'>
                                    <div class="mb-3">
                                        <label class="form-label">CutTime Entrada</label>
                                        <input type='text' class='form-control Fomulario_Ruta hora' campo='CUTTIME_IN' ruta='"""+str(idRuta)+"""' value='"""+str(Concentrado[Tipo][R]["CUTTIME_IN"])+"""' placeholder='' onchange='Cambio_Algo(\""""+str(idRuta)+"""\")'>
                                    </div>
                                </div>
                                <div class='col-4'>
                                    <div class="mb-3">
                                        <label class="form-label">CutTime Salida</label>
                                        <input type='text' class='form-control Fomulario_Ruta hora' campo='CUTTIME_OUT' ruta='"""+str(idRuta)+"""' value='"""+str(Concentrado[Tipo][R]["CUTTIME_OUT"])+"""' placeholder='' onchange='Cambio_Algo(\""""+str(idRuta)+"""\")'>
                                    </div>
                                </div>

                            </div>
                    <hr>
                """+str(Ruta)+"""
                    <div><span class='fw-bold'>Lista de distribución (Caja Lista)</span> <a onclick='Agregar_item(\""""+str(idRuta)+"""\")' class='link-success' style='cursor:pointer;' ><i class='mdi mdi-plus'></i> Nuevo</a></div>
                    <ol class='list-group list-group-numbered Fomulario_Ruta' campo='DL' ruta='"""+str(idRuta)+"""'>
                """
                for C in str(Concentrado[Tipo][R]["DL"]).split(";"):
                    if str(C).strip() != "":
                        Configuracion += "<li class='list-group-item d-flex justify-content-between align-items-start p-0 m-0 ps-1 pe-1' valor='"+str(C).strip()+"'><div class='ms-2 me-auto'>"+str(C).strip()+"</div><span style='cursor:pointer' onclick='$(this).parent().remove(); Cambio_Algo(\""+str(idRuta)+"\");' class='badge bg-danger rounded-pill'><i class='mdi mdi-trash-can'></i></span></li>"
                Configuracion += """
                    </ol>
                    <hr>
                    <div class='w-100 text-center'><button onclick='Guardar_Cambios(\""""+str(idRuta)+"""\",\""""+str(R)+"""\",\""""+str(Tipo)+"""\")' class='btn btn-success w-75'><i class='mdi mdi-floppy'></i> Guardar Cambios</button></div>
                    
                    <div class='w-100 text-end mt-2'><button onclick='Eliminar_Ruta(\""""+str(idRuta)+"""\",\""""+str(R)+"""\",\""""+str(Tipo)+"""\")' class='btn btn-danger w-25'><i class='mdi mdi-floppy'></i> Eliminar Ruta</button></div>
                </div>
                """
                aux["_children"].append({"Ruta": str(R),"_children":[{"Configuracion":Configuracion}]})
            aux["_children"].append({"Ruta": "<a onclick='Agregar_Nuevo(\""+str(Tipo)+"\")' class='link-success' style='cursor:pointer;'><i class='mdi mdi-plus'></i> Nuevo</a>"})
            Informacion.append(aux)
        DUNS_DB = {} 
        for Prov in DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".cproveedores WHERE crilcpr_activo = '1' ORDER BY crilcpr_codigo"):
            DUNS_DB[str(Prov["crilcpr_codigo"])] = str(Prov["crilcpr_codigo"])+" - "+str(Prov["crilcpr_nombre"])
        Cur += """
        <div id='Tabla_Conf' class='border border-dark bg-dark-subtle'></div>
        <script>
            var table = new Tabulator("#Tabla_Conf", {
                minHeight:500,
                layout:"fitColumns",
                data:"""+str(Informacion)+""",
                dataTree:true,
                columns:[
                {title:"Tipo", field:"Tipo",widthGrow:1,formatter:"html"},
                {title:"Ruta", field:"Ruta",widthGrow:1,formatter:"html"},
                {title:"Configuracion", field:"Configuracion",formatter:"html",widthGrow:10}
                ],
            });
            table.on("renderComplete", function(){
                $('.hora').flatpickr({enableTime: true,noCalendar: true,dateFormat: 'H:i',time_24hr: true,defaultHour:0,defaultMinute:0});
            });
            function Agregar_item(Ruta)
            {
                $.fn.modal.Constructor.prototype._initializeFocusTrap = function () { return { activate: function () { }, deactivate: function () { } } };
                Swal.fire({
                title: 'Agregar Correo',
                layout:'fitData',
                input: 'email',
                showCloseButton: true,
                showCancelButton: true,
                focusConfirm: false,
                confirmButtonText:'<i class="mdi mdi-plus"></i> Agregar',
                cancelButtonText:'<i class="mdi mdi-close"></i> Cancelar'
                ,preConfirm: (item) => {
                    $(".list-group[ruta='"+Ruta+"']").children().each(function() {
                        if($(this).attr('valor').trim().toUpperCase() == item.trim().toUpperCase())
                        {
                            Swal.showValidationMessage('Este ítem ya existe');
                            return false;
                        }
                    });
                    return item.trim();
                },
                allowOutsideClick: () => !Swal.isLoading()
                }).then((result) => {
                    if (result.isConfirmed)
                    {
                        $(".list-group[ruta='"+Ruta+"']").append("<li class='list-group-item d-flex justify-content-between align-items-start p-0 m-0 ps-1 pe-1' valor='"+result.value+"'><div class='ms-2 me-auto'>"+result.value+"</div><span style='cursor:pointer' onclick='$(this).parent().remove();' class='badge bg-danger rounded-pill'><i class='mdi mdi-trash-can'></i></span></li>")
                        $(".list-group[ruta='"+Ruta+"']").parent().parent().css('height','auto');
                        Cambio_Algo(Ruta);
                    }
                })

            }
            function Agregar_item_DUNS(Ruta){
                $.fn.modal.Constructor.prototype._initializeFocusTrap = function () { return { activate: function () { }, deactivate: function () { } } };
                Swal.fire({
                title: 'Agregar DUNS',
                layout:'fitData',
                input: 'select',
                inputOptions:"""+str(DUNS_DB)+""",
                showCloseButton: true,
                showCancelButton: true,
                focusConfirm: false,
                confirmButtonText:'<i class="mdi mdi-plus"></i> Agregar',
                cancelButtonText:'<i class="mdi mdi-close"></i> Cancelar'
                ,preConfirm: (item) => {
                    $(".list-group[rutaduns='"+Ruta+"']").children().each(function() {
                        if($(this).attr('valor').trim().toUpperCase() == item.trim().toUpperCase())
                        {
                            Swal.showValidationMessage('Este ítem ya existe');
                            return false;
                        }
                    });
                    return item.trim();
                },
                allowOutsideClick: () => !Swal.isLoading()
                }).then((result) => {
                    if (result.isConfirmed)
                    {
                        $(".list-group[rutaduns='"+Ruta+"']").append("<li class='list-group-item d-flex justify-content-between align-items-start p-0 m-0 ps-1 pe-1' valor='"+result.value+"'><div class='ms-2 me-auto'>"+result.value+"</div><span style='cursor:pointer' onclick='$(this).parent().remove();' class='badge bg-danger rounded-pill'><i class='mdi mdi-trash-can'></i></span></li>")
                        $(".list-group[rutaduns='"+Ruta+"']").parent().parent().css('height','auto');
                        Cambio_Algo(Ruta);
                    }
                })
            }
            function Duplicar_item_DUNS(Ruta){
                var Dias = ["T","W","R","F"];
                Dias.forEach((D) => {
                    $(".list-group[rutaduns='"+Ruta+D+"']").html($(".list-group[rutaduns='"+Ruta+"M']").html());
                });
            }

            function Cambio_Algo(Ruta){
                $(".Estado[ruta='"+Ruta+"']").html("| Sin guardar cambios").removeClass('text-success').addClass('text-danger');
            }

            function Guardar_Cambios(Ruta,Ruta_Nombre,Tipo){
                var Info = {};
                $(".Fomulario_Ruta[ruta='"+Ruta+"']").each(function() {
                    var Control = $(this);
                    if(Control.attr('campo') == "Inbound" || Control.attr('campo') == "Outbound"){
                        if(Control.prop('checked'))
                            Info[Control.attr('campo')] = 1
                        else
                            Info[Control.attr('campo')] = 0
                    }
                    if(Control.attr('campo') == "NOMBRE" || Control.attr('campo') == "ZONA" || Control.attr('campo') == "SCAC" || Control.attr('campo') == "CUTTIME_IN" || Control.attr('campo') == "CUTTIME_OUT"){
                        Info[Control.attr('campo')] = Control.val();
                    }
                    if(Control.attr('campo') == "DESTINO"){
                        Info[Control.attr('campo')] = Control.find('option:selected').val();
                    }
                    if(Control.attr('campo') == "DL"){
                        Info["DL"] = [];
                        var DL = [];
                        Control.children().each(function() {
                            var item = $(this);
                            Info["DL"].push(item.attr('valor'));
                        });
                    }
                });

                Info["DUNS"] = {};
                $(".Lista-DUNS[ruta='"+Ruta+"']").each(function() {
                    var rutaduns = $(this);
                    Info["DUNS"][rutaduns.attr('rutaduns').split('_')[1]] = [];
                    rutaduns.children().each(function() {
                        var item = $(this);
                        Info["DUNS"][rutaduns.attr('rutaduns').split('_')[1]].push(item.attr('valor'));
                    });
                });
                
                Swal.fire({
                title: '¿Estas segur@ de guardar los cambios de la ruta ['+Tipo+'/'+Ruta_Nombre+']?',
                buttonsStyling: false,showCancelButton: true,confirmButtonText: "<i class='mdi mdi-check'></i> Yes",cancelButtonText: "<i class='mdi mdi-close'></i> No",showLoaderOnConfirm: true,
                customClass: {confirmButton: 'btn btn-success ms-1 me-1',cancelButton: 'btn btn-danger ms-1 me-1'},
                preConfirm: () => {
                
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Guardar_Cambios".encode()).decode("utf-8"))+"""',"Ruta_Nombre":Ruta_Nombre,"Tipo":Tipo,"Info":JSON.stringify(Info)};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                Mensaje(2);
                                $(".Estado[ruta='"+Ruta+"']").html("| ¡Guardado con éxito!").removeClass('text-danger').addClass('text-success');;
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                    
                }
                })
            }

            function Agregar_Nuevo(Tipo){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-plus'></i>" + Tipo );
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm');
                var parametros = {"Fun":'"""+str(fernet.encrypt("Agregar_Nuevo".encode()).decode("utf-8"))+"""',"Tipo":Tipo};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });
            }

            function Eliminar_Ruta(Ruta,Ruta_Nombre,Tipo){                
                Swal.fire({
                title: '¿Estas segur@ de Eliminar la ruta ['+Tipo+'/'+Ruta_Nombre+']?',
                buttonsStyling: false,showCancelButton: true,confirmButtonText: "<i class='mdi mdi-check'></i> Yes",cancelButtonText: "<i class='mdi mdi-close'></i> No",showLoaderOnConfirm: true,
                customClass: {confirmButton: 'btn btn-success ms-1 me-1',cancelButton: 'btn btn-danger ms-1 me-1'},
                preConfirm: () => {
                
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Eliminar_Ruta".encode()).decode("utf-8"))+"""',"Ruta_Nombre":Ruta_Nombre,"Tipo":Tipo};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                Mensaje(2);
                                Llamar_Funcion(\""""+str(request.url)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                    
                }
                })
            }
        </script>
        """
    except:
         Cur = str(sys.exc_info())
    #Cur += json.dumps(Resultado)
    return Cur
def Guardar_Cambios(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Info_Datos = json.loads(str(Datos["Info"]))
        Rutas = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo != 'Empty'")
        Inbound = json.loads(Rutas[0]["cr_niveles"])
        Outbound = json.loads(Rutas[1]["cr_niveles"])
        if int(Info_Datos["Inbound"]) == 0 and int(Info_Datos["Outbound"]) == 0:
            Resultado["Contenido"] += "Debe ser Inbound o Outbound o ambos"
        else:
            if Datos["Tipo"] == "Nacional":
                if int(Info_Datos["Inbound"]) == 1:
                    Inbound["Nacional"][str(Datos["Ruta_Nombre"])] = {
                        "NOMBRE":str(Info_Datos["NOMBRE"]),
                        "ZONA":str(Info_Datos["ZONA"]),
                        "SCAC":str(Info_Datos["SCAC"]),
                        "DESTINO":str(Info_Datos["DESTINO"]),
                        "CUTTIME_IN":str(Info_Datos["CUTTIME_IN"]),
                        "CUTTIME_OUT":str(Info_Datos["CUTTIME_OUT"]),
                        "DOCKS":[],
                        "RUTAS":[],
                        "DL":";".join(Info_Datos["DL"])
                    }
                else:
                    try:
                        del Inbound["Nacional"][str(Datos["Ruta_Nombre"])]
                    except:
                        pass
                if int(Info_Datos["Outbound"]) == 1:
                    Outbound["Nacional"][str(Datos["Ruta_Nombre"])] = {
                        "NOMBRE":str(Info_Datos["NOMBRE"]),
                        "ZONA":str(Info_Datos["ZONA"]),
                        "SCAC":str(Info_Datos["SCAC"]),
                        "DESTINO":str(Info_Datos["DESTINO"]),
                        "CUTTIME_IN":str(Info_Datos["CUTTIME_IN"]),
                        "CUTTIME_OUT":str(Info_Datos["CUTTIME_OUT"]),
                        "DOCKS":[],
                        "RUTAS":[],
                        "DL":";".join(Info_Datos["DL"])
                    }
                else:
                    try:
                        del Outbound["Nacional"][str(Datos["Ruta_Nombre"])]
                    except:
                        pass
            if Datos["Tipo"] == "Internacional":
                if int(Info_Datos["Inbound"]) == 1:
                    Inbound["Importacion"][str(Datos["Ruta_Nombre"])] = {
                        "NOMBRE":str(Info_Datos["NOMBRE"]),
                        "ZONA":str(Info_Datos["ZONA"]),
                        "SCAC":str(Info_Datos["SCAC"]),
                        "DESTINO":str(Info_Datos["DESTINO"]),
                        "CUTTIME_IN":str(Info_Datos["CUTTIME_IN"]),
                        "CUTTIME_OUT":str(Info_Datos["CUTTIME_OUT"]),
                        "DOCKS":[],
                        "RUTAS":[],
                        "DL":";".join(Info_Datos["DL"])
                    }
                else:
                    try:
                        del Inbound["Importacion"][str(Datos["Ruta_Nombre"])]
                    except:
                        pass
                if int(Info_Datos["Outbound"]) == 1:
                    Outbound["Exportacion"][str(Datos["Ruta_Nombre"])] = {
                        "NOMBRE":str(Info_Datos["NOMBRE"]),
                        "ZONA":str(Info_Datos["ZONA"]),
                        "SCAC":str(Info_Datos["SCAC"]),
                        "DESTINO":str(Info_Datos["DESTINO"]),
                        "CUTTIME_IN":str(Info_Datos["CUTTIME_IN"]),
                        "CUTTIME_OUT":str(Info_Datos["CUTTIME_OUT"]),
                        "DOCKS":[],
                        "RUTAS":[],
                        "DL":";".join(Info_Datos["DL"])
                    }
                else:
                    try:
                        del Outbound["Exportacion"][str(Datos["Ruta_Nombre"])]    
                    except:
                        pass 
            if Datos["Tipo"] == "MilkRuns":
                Dias = {"M":"Luneas","T":"Martes","W":"Miercoles","R":"Jueves","F":"Viernes"}
                if int(Info_Datos["Inbound"]) == 1:
                    for D in Dias.keys():
                        Ruta = []
                        if "DUNS" in Info_Datos.keys():
                            if str(Datos["Ruta_Nombre"])+str(D) in Info_Datos["DUNS"].keys():
                                for DUNS in Info_Datos["DUNS"][str(Datos["Ruta_Nombre"])+str(D)]:
                                    Ruta.append({"DUNS":int(DUNS)})
                        Inbound["MilkRun"][str(Datos["Ruta_Nombre"])+str(D)] = {
                            "NOMBRE":str(Info_Datos["NOMBRE"]),
                            "ZONA":str(Info_Datos["ZONA"]),
                            "SCAC":str(Info_Datos["SCAC"]),
                            "DESTINO":str(Info_Datos["DESTINO"]),
                            "CUTTIME_IN":str(Info_Datos["CUTTIME_IN"]),
                            "CUTTIME_OUT":str(Info_Datos["CUTTIME_OUT"]),
                            "DOCKS":[],
                            "RUTAS":Ruta,
                            "DL":";".join(Info_Datos["DL"]),
                            "UNION":str(Datos["Ruta_Nombre"])
                        }
                else:
                    for D in Dias.keys():
                        try:
                            del Inbound["MilkRun"][str(Datos["Ruta_Nombre"])+str(D)]
                        except:
                            pass
                if int(Info_Datos["Outbound"]) == 1:
                    for D in Dias.keys():
                        Ruta = []
                        if "DUNS" in Info_Datos.keys():
                            if str(Datos["Ruta_Nombre"])+str(D) in Info_Datos["DUNS"].keys():
                                for DUNS in Info_Datos["DUNS"][str(Datos["Ruta_Nombre"])+str(D)]:
                                    Ruta.append({"DUNS":int(DUNS)})
                        Outbound["MilkRun"][str(Datos["Ruta_Nombre"])+str(D)] = {
                            "NOMBRE":str(Info_Datos["NOMBRE"]),
                            "ZONA":str(Info_Datos["ZONA"]),
                            "SCAC":str(Info_Datos["SCAC"]),
                            "DESTINO":str(Info_Datos["DESTINO"]),
                            "CUTTIME_IN":str(Info_Datos["CUTTIME_IN"]),
                            "CUTTIME_OUT":str(Info_Datos["CUTTIME_OUT"]),
                            "DOCKS":[],
                            "RUTAS":Ruta,
                            "DL":";".join(Info_Datos["DL"]),
                            "UNION":str(Datos["Ruta_Nombre"])
                        }
                else:
                    for D in Dias.keys():
                        try:
                            del Outbound["MilkRun"][str(Datos["Ruta_Nombre"])+str(D)] 
                        except:
                            pass
            Error = ""
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".crutas SET cr_niveles = '"+str(json.dumps(Inbound))+"' WHERE cr_tipo = 'Inbound'")
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".crutas SET cr_niveles = '"+str(json.dumps(Outbound))+"' WHERE cr_tipo = 'Outbound'")
            if Error == "":
                Resultado["Estado"] = 1
            else:
                Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Agregar_Nuevo(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Formulario = {"Col":"12", "Campos": [],"Clase": "Agregar_Nuevo" }
        Formulario["Campos"].append({"tipo":"texto","campo":"Nombre","titulo":"Ruta (sin espacios)","Requerido":1,"min":1,"max":150,"valor":""})
        Formulario["Campos"].append({"tipo":"sep","campo":"","titulo":"<div>La ruta puede ser:</div>"})
        Formulario["Campos"].append({"tipo":"checkbox","campo":"Inbound","titulo":"Inbound","Requerido":1,"valor":False})
        Formulario["Campos"].append({"tipo":"checkbox","campo":"Outbound","titulo":"Outbound","Requerido":1,"valor":False})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += """
        <hr>
        <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Agregar_Nuevo_Guardar(\""""+str(Datos["Tipo"])+"""\")'><i class='mdi mdi-floppy'></i> Save</button></div>
        """
        Resultado["Contenido"] += """
        <script>
            function Agregar_Nuevo_Guardar(Tipo){
                var Info = Dame_Formulario(".Agregar_Nuevo",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Agregar_Nuevo_Guardar".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"Tipo":Tipo};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                $("#Vent_1").modal("hide");
                                Mensaje(2);
                                Llamar_Funcion(\""""+str(request.url)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
            }
        </script>
        """
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Agregar_Nuevo_Guardar(Datos):
    DB = LibDM_2023.DataBase()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Info_Datos = json.loads(str(Datos["Info"]))
        Rutas = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo != 'Empty'")
        Inbound = json.loads(Rutas[0]["cr_niveles"])
        Outbound = json.loads(Rutas[1]["cr_niveles"])
        Ya_Existe = False
        for Ruta in Rutas:
            Niveles = json.loads(Ruta["cr_niveles"])
            for Tipo in Niveles.keys():
                if str(Info_Datos["Nombre"]).upper() in Niveles[Tipo].keys():
                    Ya_Existe = True
        if Ya_Existe:
            Resultado["Contenido"] += "Este Ruta ya existe!"
        else:
            if int(Info_Datos["Inbound"]) == 0 and int(Info_Datos["Outbound"]) == 0:
                Resultado["Contenido"] += "Debe ser Inbound o Outbound o ambos"
            else:
                Info_Datos["Nombre"] = str(Info_Datos["Nombre"]).replace(" ","")
                if Datos["Tipo"] == "Nacional":
                    if int(Info_Datos["Inbound"]) == 1:
                        Inbound["Nacional"][str(Info_Datos["Nombre"]).upper()] = {
                            "NOMBRE":"",
                            "ZONA":"",
                            "SCAC":"",
                            "DESTINO":"",
                            "CUTTIME_IN":"",
                            "CUTTIME_OUT":"",
                            "DOCKS":[],
                            "RUTAS":[],
                            "DL":""
                        }
                    if int(Info_Datos["Outbound"]) == 1:
                        Outbound["Nacional"][str(Info_Datos["Nombre"]).upper()] = {
                            "NOMBRE":"",
                            "ZONA":"",
                            "SCAC":"",
                            "DESTINO":"",
                            "CUTTIME_IN":"",
                            "CUTTIME_OUT":"",
                            "DOCKS":[],
                            "RUTAS":[],
                            "DL":""
                        }
                if Datos["Tipo"] == "Internacional":
                    if int(Info_Datos["Inbound"]) == 1:
                        Inbound["Importacion"][str(Info_Datos["Nombre"]).upper()] = {
                            "NOMBRE":"",
                            "ZONA":"",
                            "SCAC":"",
                            "DESTINO":"",
                            "CUTTIME_IN":"",
                            "CUTTIME_OUT":"",
                            "DOCKS":[],
                            "RUTAS":[],
                            "DL":""
                        }
                    if int(Info_Datos["Outbound"]) == 1:
                        Outbound["Exportacion"][str(Info_Datos["Nombre"]).upper()] = {
                            "NOMBRE":"",
                            "ZONA":"",
                            "SCAC":"",
                            "DESTINO":"",
                            "CUTTIME_IN":"",
                            "CUTTIME_OUT":"",
                            "DOCKS":[],
                            "RUTAS":[],
                            "DL":""
                        }
                if Datos["Tipo"] == "MilkRuns":
                    Dias = {"M":"Luneas","T":"Martes","W":"Miercoles","R":"Jueves","F":"Viernes"}
                    if int(Info_Datos["Inbound"]) == 1:
                        for D in Dias.keys():
                            Inbound["MilkRun"][str(Info_Datos["Nombre"]).upper()+str(D)] = {
                                "NOMBRE":"",
                                "ZONA":"",
                                "SCAC":"",
                                "DESTINO":"",
                                "CUTTIME_IN":"",
                                "CUTTIME_OUT":"",
                                "DOCKS":[],
                                "RUTAS":[],
                                "DL":"",
                                "UNION":str(Info_Datos["Nombre"]).upper()
                            }
                    if int(Info_Datos["Outbound"]) == 1:
                        for D in Dias.keys():
                            Outbound["MilkRun"][str(Info_Datos["Nombre"]).upper()+str(D)] = {
                                "NOMBRE":"",
                                "ZONA":"",
                                "SCAC":"",
                                "DESTINO":"",
                                "CUTTIME_IN":"",
                                "CUTTIME_OUT":"",
                                "DOCKS":[],
                                "RUTAS":[],
                                "DL":"",
                                "UNION":str(Info_Datos["Nombre"]).upper()
                            }
                Error = ""
                Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".crutas SET cr_niveles = '"+str(json.dumps(Inbound))+"' WHERE cr_tipo = 'Inbound'")
                Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".crutas SET cr_niveles = '"+str(json.dumps(Outbound))+"' WHERE cr_tipo = 'Outbound'")
                if Error == "":
                    Resultado["Estado"] = 1
                else:
                    Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Eliminar_Ruta(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Resultado["Contenido"] += str(Datos)
        Rutas = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo != 'Empty'")
        Inbound = json.loads(Rutas[0]["cr_niveles"])
        Outbound = json.loads(Rutas[1]["cr_niveles"])
        if Datos["Tipo"] == "Nacional":
            try:
                del Inbound["Nacional"][str(Datos["Ruta_Nombre"])]
            except:
                pass
            try:
                del Outbound["Nacional"][str(Datos["Ruta_Nombre"])]
            except:
                pass
        if Datos["Tipo"] == "Internacional":
            try:
                del Inbound["Importacion"][str(Datos["Ruta_Nombre"])]
            except:
                pass
            try:
                del Outbound["Exportacion"][str(Datos["Ruta_Nombre"])]
            except:
                pass
        if Datos["Tipo"] == "MilkRuns":
            Dias = {"M":"Luneas","T":"Martes","W":"Miercoles","R":"Jueves","F":"Viernes"}
            for D in Dias.keys():
                try:
                    del Inbound["MilkRun"][str(Datos["Ruta_Nombre"])+str(D)]
                except:
                    pass
                try:
                    del Outbound["MilkRun"][str(Datos["Ruta_Nombre"])+str(D)]
                except:
                    pass           

        Error = ""
        Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".crutas SET cr_niveles = '"+str(json.dumps(Inbound))+"' WHERE cr_tipo = 'Inbound'")
        Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".crutas SET cr_niveles = '"+str(json.dumps(Outbound))+"' WHERE cr_tipo = 'Outbound'")
        if Error == "":
            Resultado["Estado"] = 1
        else:
            Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur



def Agregar_Nivel(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        if str(Datos["Nivel"]) == "N1":
            Formulario = {"Col":"12", "Campos": [],"Clase": "Alta_Ruta" }
            Formulario["Campos"].append({"tipo":"texto","campo":"Nombre","titulo":"Name","Requerido":1,"min":1,"max":30,"valor":""})
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        if str(Datos["Nivel"]) == "N2":
            Resultado["Contenido"] += "<div class='w-100 text-center'><button class='btn btn-primary' onclick='Descargar_Excel(\""+str(Datos["Direccion"])+"\")'><i class='mdi mdi-cloud-download'></i> Download configuration file (Excel)</button></div>"
            Formulario = {"Col":"12", "Campos": [],"Clase": "Alta_Ruta" }
            Formulario["Campos"].append({"tipo":"archivo","campo":"Nueva Rutas","titulo":"New route","Requerido":1,"Col":12,"min":1,"max":1,"tipo_archivo":["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],"valor":""})
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += """
        <hr>
        <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Guardar_Nivel(\""""+str(Datos["Nivel"])+"""\",\""""+str(Datos["Direccion"])+"""\")'><i class='mdi mdi-floppy'></i> Save</button></div>
        <script>
            function Guardar_Nivel(Nivel,Direccion){
                var Info = Dame_Formulario(".Alta_Ruta",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Guardar_Nivel".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"Nivel":Nivel,"Direccion":Direccion};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                $("#Vent_1").modal("hide");
                                Mensaje(2);
                                Llamar_Funcion(\""""+str(request.url)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
            }
            function Descargar_Excel(Direccion){
                Mostrar_Ventana_Cargando(false);
                var parametros = {"Fun":'"""+str(fernet.encrypt("Descargar_Excel".encode()).decode("utf-8"))+"""',"Direccion":Direccion};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        if(Resultado["Estado"] == 1)
                        {
                            if("Archivo" in Resultado)
                            {
                                var win1 = window.open('"""+str(request.url_root)+"""/Portal_File/Gen/'+Resultado["Archivo"], '_blank');
                                win1.focus();
                            }
                            Mensaje(2);
                        }
                        else
                            Mensaje(0,Resultado["Contenido"]);
                            
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        Mensaje(0,textStatus);
                    }
                });
            }
        </script>
        """
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Guardar_Nivel(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Error = ""
        Info_Datos = json.loads(str(Datos["Info"]))
        Tipo = str(Datos["Direccion"]).split('/')[0]
        Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo = '"+str(Tipo)+"'")[0]
        if Info_Ahora["cr_niveles"] is None or str(Info_Ahora["cr_niveles"]).strip() == "":
            Info_Ruta = {}
        else:
            Info_Ruta = json.loads(str(Info_Ahora["cr_niveles"]))
        if Datos["Nivel"] == "N1":
            Info_Ruta[str(Info_Datos["Nombre"])] = 0
        if Datos["Nivel"] == "N2":
            Tipo_2 = str(Datos["Direccion"]).split('/')[1]
            Info_Ruta[str(Tipo_2)] = {}
            wb_obj = openpyxl.load_workbook(str(current_app.root_path).replace("\\","/")+"/Files/"+str(Info_Datos["Nueva Rutas"][0]))
            
            Dia = ["M","T","W","R","F"]
            Runas_Letras = []
            for N1 in wb_obj.sheetnames:
                for D in Dia:
                    if str(D) in str(N1) and str(N1) not in Runas_Letras:
                        Runas_Letras.append(N1)

            for N1 in wb_obj.sheetnames:
                if str(Tipo_2) == "MilkRun":
                    if str(N1) in Runas_Letras:
                        N2 = {}
                        wb_obj.active = wb_obj[N1]
                        Shell = wb_obj.active
                        N2["NOMBRE"] = Shell["B1"].value
                        N2["ZONA"] = Shell["B2"].value
                        N2["SCAC"] = Shell["B3"].value
                        
                        index = 6
                        N2["DOCKS"] = []
                        A = "."
                        while A is not None and str(A).strip() != "":
                            A = Shell["A"+str(index)].value
                            if A != "" and A is not None:
                                N2["DOCKS"].append(A)
                                index += 1
                        
                        index = 7
                        N2["RUTAS"] = []
                        A = "."
                        while A is not None and str(A).strip() != "":
                            A = Shell["D"+str(index)].value
                            if A != "" and A is not None:
                                DUNS = Shell["D"+str(index)].value

                                INICIO = ""
                                try:
                                    INICIO = Shell["E"+str(index)].value.strftime("%Y-%m-%d %H:%M:%S")
                                except:
                                    INICIO = datetime.strptime(Shell["E"+str(index)].value,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")
                                
                                FIN = ""
                                try:
                                    FIN = Shell["F"+str(index)].value.strftime("%Y-%m-%d %H:%M:%S")
                                except:
                                    FIN = datetime.strptime(Shell["F"+str(index)].value,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")

                                R = {"DUNS":DUNS,"INICIA":INICIO,"FIN":FIN}
                                N2["RUTAS"].append(R)
                                index += 1

                        DL = []
                        index = 6
                        A = "."
                        while A is not None and str(A).strip() != "":
                            A = Shell["H"+str(index)].value
                            if A != "" and A is not None:
                                DL.append(A.replace(" ","").replace(";",""))
                                index += 1

                        N2["DL"] = ";".join(DL)


                        # if len(N2.keys()) == 0:
                        #     Info_Ruta[str(Tipo_2)][N1] = 0
                        # else:
                        Info_Ruta[str(Tipo_2)][N1] = N2
                    else:
                        wb_obj.active = wb_obj[N1]
                        Dia = ["M","T","W","R","F"]
                        for D in Dia:
                            if str(N1)+str(D) not in Runas_Letras:
                                N2 = {}
                                Shell = wb_obj.active
                                N2["NOMBRE"] = Shell["B1"].value
                                N2["ZONA"] = Shell["B2"].value
                                N2["SCAC"] = Shell["B3"].value
                                
                                index = 6
                                N2["DOCKS"] = []
                                A = "."
                                while A is not None and str(A).strip() != "":
                                    A = Shell["A"+str(index)].value
                                    if A != "" and A is not None:
                                        N2["DOCKS"].append(A)
                                        index += 1
                                
                                index = 7
                                N2["RUTAS"] = []
                                A = "."
                                while A is not None and str(A).strip() != "":
                                    A = Shell["D"+str(index)].value
                                    if A != "" and A is not None:
                                        DUNS = Shell["D"+str(index)].value

                                        INICIO = ""
                                        try:
                                            INICIO = Shell["E"+str(index)].value.strftime("%Y-%m-%d %H:%M:%S")
                                        except:
                                            INICIO = datetime.strptime(Shell["E"+str(index)].value,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")
                                        
                                        FIN = ""
                                        try:
                                            FIN = Shell["F"+str(index)].value.strftime("%Y-%m-%d %H:%M:%S")
                                        except:
                                            FIN = datetime.strptime(Shell["F"+str(index)].value,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")

                                        R = {"DUNS":DUNS,"INICIA":INICIO,"FIN":FIN}
                                        N2["RUTAS"].append(R)
                                        index += 1

                                DL = []
                                index = 6
                                A = "."
                                while A is not None and str(A).strip() != "":
                                    A = Shell["H"+str(index)].value
                                    if A != "" and A is not None:
                                        DL.append(A.replace(" ","").replace(";",""))
                                        index += 1

                                N2["DL"] = ";".join(DL)


                                # if len(N2.keys()) == 0:
                                #     Info_Ruta[str(Tipo_2)][N1] = 0
                                # else:
                                Info_Ruta[str(Tipo_2)][N1+str(D)] = N2
                else:
                    N2 = {}
                    wb_obj.active = wb_obj[N1]
                    Shell = wb_obj.active
                    N2["NOMBRE"] = Shell["B1"].value
                    N2["ZONA"] = Shell["B2"].value
                    N2["SCAC"] = Shell["B3"].value
                    
                    index = 6
                    N2["DOCKS"] = []
                    A = "."
                    while A is not None and str(A).strip() != "":
                        A = Shell["A"+str(index)].value
                        if A != "" and A is not None:
                            N2["DOCKS"].append(A)
                            index += 1
                    
                    index = 7
                    N2["RUTAS"] = []
                    A = "."
                    while A is not None and str(A).strip() != "":
                        A = Shell["D"+str(index)].value
                        if A != "" and A is not None:
                            DUNS = Shell["D"+str(index)].value

                            INICIO = ""
                            try:
                                INICIO = Shell["E"+str(index)].value.strftime("%Y-%m-%d %H:%M:%S")
                            except:
                                INICIO = datetime.strptime(Shell["E"+str(index)].value,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")
                            
                            FIN = ""
                            try:
                                FIN = Shell["F"+str(index)].value.strftime("%Y-%m-%d %H:%M:%S")
                            except:
                                FIN = datetime.strptime(Shell["F"+str(index)].value,"%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d %H:%M:%S")

                            R = {"DUNS":DUNS,"INICIA":INICIO,"FIN":FIN}
                            N2["RUTAS"].append(R)
                            index += 1

                    DL = []
                    index = 6
                    A = "."
                    while A is not None and str(A).strip() != "":
                        A = Shell["H"+str(index)].value
                        if A != "" and A is not None:
                            DL.append(A.replace(" ","").replace(";",""))
                            index += 1

                    N2["DL"] = ";".join(DL)
                    Info_Ruta[str(Tipo_2)][N1] = N2
                

        Error = DB.Instruccion("UPDATE "+str(BD_Nombre)+".crutas SET cr_niveles = '"+str(json.dumps(Info_Ruta))+"' WHERE cr_tipo = '"+str(Tipo)+"'")

        #Error = str(Runas_Letras)
 
        if Error == "":
            Resultado["Estado"] = 1
        else:
            Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Descargar_Excel(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Borrar_Primera = False
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NUEVA RUTA"
        Tipo = str(Datos["Direccion"]).split('/')[0]
        Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo = '"+str(Tipo)+"'")[0]
        Info_Ruta = json.loads(str(Info_Ahora["cr_niveles"]))
        if Info_Ruta[str(Datos["Direccion"]).split('/')[1]] == 0 or "NOMBRE" in Info_Ruta[str(Datos["Direccion"]).split('/')[1]].keys():

            ws['A1'] = "NOMBRE"
            ws['B1'].fill = PatternFill("solid", fgColor="fffd00")
            ws['B1'].border = Border(bottom=Side(border_style="thin", color="000000"))
            ws['B1'] = ""

            ws['A2'] = "ZONA"
            ws['B2'].fill = PatternFill("solid", fgColor="fffd00")
            ws['B2'].border = Border(bottom=Side(border_style="thin", color="000000"))
            ws['B2'] = ""

            ws['A3'] = "SCAC"
            ws['B3'].fill = PatternFill("solid", fgColor="fffd00")
            ws['B3'].border = Border(bottom=Side(border_style="thin", color="000000"))
            ws['B3'] = ""
            
            index = 6
            for r in range(0,10):
                ws['A'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                ws['A'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                ws['A'+str(index)] = ""
                index += 1
            
            ws['D5'] = "RUTA"
            ws['D6'] = "DUNS"
            ws['E6'] = "INICIA"
            ws['F6'] = "FIN"

            index = 7
            for r in range(0,10):
                ws['D'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                ws['D'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                ws['D'+str(index)] = ""
                ws['E'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                ws['E'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                ws['E'+str(index)] = ""
                ws['F'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                ws['F'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                ws['F'+str(index)] = ""
                index += 1
            
            ws['H5'] = "DL"

            index = 6
            for r in range(0,10):
                ws['H'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                ws['H'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                ws['H'+str(index)] = ""
                index += 1
        else:
            for N1 in Info_Ruta[str(Datos["Direccion"]).split('/')[1]].keys():
                
                Nombre = ""
                Zona = ""
                SCAC = ""
                Rutas = []
                Docks = []

                if Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1] == 0:
                    pass
                else:
                    Nombre = Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1]["NOMBRE"]
                    Zona = Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1]["ZONA"]
                    SCAC = Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1]["SCAC"]
                    Rutas = Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1]["RUTAS"]
                    Docks = Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1]["DOCKS"]
                    
                    if "DL" in Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1].keys():
                        DL = str(Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1]["DL"]).split(";")
                    else:
                        DL = []

               

                Borrar_Primera = True
                ws = wb.create_sheet(str(N1))
                ws['A1'] = "NOMBRE"
                ws['B1'].fill = PatternFill("solid", fgColor="fffd00")
                ws['B1'].border = Border(bottom=Side(border_style="thin", color="000000"))
                ws['B1'] = Nombre

                ws['A2'] = "ZONA"
                ws['B2'].fill = PatternFill("solid", fgColor="fffd00")
                ws['B2'].border = Border(bottom=Side(border_style="thin", color="000000"))
                ws['B2'] = Zona

                ws['A3'] = "SCAC"
                ws['B3'].fill = PatternFill("solid", fgColor="fffd00")
                ws['B3'].border = Border(bottom=Side(border_style="thin", color="000000"))
                ws['B3'] = SCAC



                
                ws['A5'] = "DOCKS"

                index = 6
                if len(Docks) == 0:
                    for r in range(0,10):
                        ws['A'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['A'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['A'+str(index)] = ""
                        index += 1
                else:
                    for D in Docks:
                        ws['A'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['A'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['A'+str(index)] = D
                        index += 1

                ws['D5'] = "RUTA"
                ws['D6'] = "DUNS"
                ws['E6'] = "INICIA"
                ws['F6'] = "FIN"

                index = 7
                if len(Rutas) == 0:
                    for r in range(0,10):
                        ws['D'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['D'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['D'+str(index)] = ""
                        ws['E'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['E'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['E'+str(index)] = ""
                        ws['F'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['F'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['F'+str(index)] = ""
                        index += 1
                else:
                    for R in Rutas:
                        ws['D'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['D'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['D'+str(index)] = R["DUNS"]
                        ws['E'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['E'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['E'+str(index)] = R["INICIA"]
                        ws['F'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['F'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['F'+str(index)] = R["FIN"]
                        index += 1

                ws['H5'] = "DL"

                index = 6
                if len(DL) == 0:
                    for r in range(0,10):
                        ws['H'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['H'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['H'+str(index)] = ""
                        index += 1
                else:
                    for D in DL:
                        ws['H'+str(index)].fill = PatternFill("solid", fgColor="fffd00")
                        ws['H'+str(index)].border = Border(bottom=Side(border_style="thin", color="000000"))
                        ws['H'+str(index)] = str(D).strip()
                        index += 1


                if Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1] == 0 or "NOMBRE" in Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1].keys():
                    pass
                else:
                    index = 1
                    for N2 in Info_Ruta[str(Datos["Direccion"]).split('/')[1]][N1].keys():
                        ws['A'+str(index)] = str(N2)
                        index += 1
        if Borrar_Primera:
            wb.remove_sheet(wb.get_sheet_by_name('NUEVA RUTA'))
        
        wb.save(str(current_app.root_path).replace("\\","/")+"/Files/Gen/"+str(Tipo)+'.xlsx')
        Resultado["Contenido"] += str(Info_Ruta)
        Resultado["Archivo"] = str(Tipo)+'.xlsx'
        Resultado["Estado"] = 1
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Eliminar_Nivel(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Tipo = str(Datos["Direccion"]).split('/')[0]
        Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo = '"+str(Tipo)+"'")[0]
        Info_Ruta = json.loads(str(Info_Ahora["cr_niveles"]))
        del Info_Ruta[str(Datos["Direccion"]).split('/')[1]]
        Error = DB.Instruccion("UPDATE "+str(BD_Nombre)+".crutas SET cr_niveles = '"+str(json.dumps(Info_Ruta))+"' WHERE cr_tipo = '"+str(Tipo)+"'")
        if Error == "":
            Resultado["Estado"] = 1
        else:
            Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur


def Direccionar(Datos):
    try:
        if "K" in session.keys():
            fernet = Fernet(session["K"])
        else:
            fernet = Fernet(LibDM_2023.Compartido().Dame_K2())
        if Datos is None:
            return Inicio()
        else:
            if "Fun" in Datos:
                Funcion = fernet.decrypt(Datos["Fun"].encode()).decode()
                Par = {}
                for P in Datos.keys():
                    if str(P) != "Fun":
                        Par[str(P)] = Datos[str(P)]
                if "IDu" in session.keys():
                    Par["ID_User"] = session["IDu"]
                return globals()[Funcion](Par)
            else:
                return Inicio()
    except:
        return str(sys.exc_info())
