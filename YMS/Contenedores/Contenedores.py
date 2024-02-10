from flask import request,session,render_template
from cryptography.fernet import Fernet
from datetime import datetime,date,timedelta
import sys
import json
import os
from Componentes import LibDM_2023
import hashlib
import random
import string
import qrcode
from io import BytesIO
import openpyxl
Url = ""
fernet = Fernet(LibDM_2023.Compartido().Dame_K2())
BD_Nombre = "public"
Bandera_Dock = "YMS"
def Inicio():
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    Cur = ""
    try:
        DB = LibDM_2023.DataBase()
        Activo = "YMS"
        Compartido = LibDM_2023.Compartido()
        Menu = LibDM_2023.Menu().Menu(Activo,request.url_root,session["IDu"])
        Titulo = LibDM_2023.Menu().Get_Titulo(Activo)
        Contenido = ""
        Contenedores_Patio = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas where cc_activo = 1 AND cc_caja_danada = 0 AND cc_bloqueo_rg = 0 AND cc_ubicacion = 'Patio'")
        Contenedores_Dock = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas where cc_activo = 1 AND cc_caja_danada = 0 AND cc_bloqueo_rg = 0 AND cc_ubicacion = 'Dock'")
        Patio = []
        for Contenedor in Contenedores_Patio:
            Opciones = "<div class='btn-group' role='group' aria-label='Basic mixed styles example'>"
            Info_Actual = json.loads(str(Contenedor["cc_informacion_actual"]))

            Opciones += "<button class='btn btn-sm btn-primary p-0 ps-1 pe-1' onclick='Ver_Historico("+str(Contenedor["cc_id"])+",\""+str(Contenedor["cc_contenedor"])+"\")'><i class='mdi mdi-history'></i></button>"
            if int(Contenedor["cc_bloquear"]) == 0:
                Opciones += "<button class='btn btn-sm btn-warning p-0 ps-1 pe-1' onclick='Modificar("+str(Contenedor["cc_id"])+",\""+str(Contenedor["cc_contenedor"])+"\")'><i class='mdi mdi-pencil'></i></button>"
            else:
                Opciones += "<span class='ms-1 me-1'><i class='mdi mdi-lock'></i></span>"
                

            Sello = ""
            if "Sello Proveedor" in Info_Actual.keys():
                Sello = "<span style='color:blue'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Proveedor"])
            if "Sello Temporal" in Info_Actual.keys():
                Sello = "<span style='color:#7a7a7a'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Temporal"])
            if "Sello Blanco" in Info_Actual.keys():
                Sello = "<span><i class='mdi mdi-label-outline'></i></span> "+str(Info_Actual["Sello Temporal"])
            if "Sello Rojo" in Info_Actual.keys():
                Sello = "<span style='color:#ff0000'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Temporal"])

            Ruta = ""
            Ruta += "<span class='h-100 border' style='width:20px; background:var(--Color_"+str(Contenedor["cc_tipo_actual"])+");Color:var(--Color_"+str(Contenedor["cc_tipo_actual"])+");'>__</span>"
            try:
                Tipo = Info_Actual[Contenedor["cc_tipo_actual"]]
                Ruta += str(Tipo)
                Tipo_1 = Info_Actual[Tipo]
                Ruta += "/"+str(Tipo_1)
                Tipo_2 = Info_Actual[Tipo_1]
                Ruta += "/"+str(Tipo_2)
            except:
                pass
            
            if "Fecha_Salida" in Info_Actual.keys():
                Ruta += " <"+str(Info_Actual["Fecha_Salida"])+">"


            Estado = ""
            if "Etapa" not in Info_Actual.keys():
                if Contenedor["cc_tipo_actual"] is None:
                    Estado = "<span class='fw-bold text-danger'>Not assigned</span>"
                else:
                    Estado = "Waiting"
            else:
                Estado = Info_Actual["Etapa"]
                if Info_Actual["Etapa"] == "Pending Exit":
                    #Opciones += "<button class='btn btn-sm btn-dark p-0 ps-1 pe-1' onclick='Imprimir_Pase("+str(Contenedor["cc_id"])+",\""+str(Contenedor["cc_contenedor"])+"\")'><i class='mdi mdi-printer'></i></button>"
                    Opciones += "<button class='btn btn-sm btn-danger p-0 ps-1 pe-1' onclick='Salida("+str(Contenedor["cc_id"])+",\""+str(Contenedor["cc_contenedor"])+"\")'><i class='mdi mdi-exit-to-app'></i></button>"

            Opciones += "</div>"

            for K in Contenedor.keys():
                if Contenedor[K] is None:
                    Contenedor[K] = ""
            Patio.append({"Tiempo":"<span class='tiempo' tiempo='"+str(Contenedor["cc_ultimo_mov"])+"'></span>","Caja":str(Contenedor["cc_contenedor"]),"Carrier":Info_Actual["Carrier"],"Tipo":Contenedor["cc_tipo_actual"],"Estado":str(Estado),"Sello":Sello,"Ruta":Ruta,"Opciones":"<div class='text-center'>"+str(Opciones)+"</div>"})

        Dock = []
        for Contenedor in Contenedores_Dock:
            Opciones = "<div class='btn-group' role='group' aria-label='Basic mixed styles example'>"
            Info_Actual = json.loads(str(Contenedor["cc_informacion_actual"]))
            Opciones += "<button class='btn btn-sm btn-primary p-0 ps-1 pe-1' onclick='Ver_Historico("+str(Contenedor["cc_id"])+",\""+str(Contenedor["cc_contenedor"])+"\")'><i class='mdi mdi-history'></i></button>"
            if Contenedor["cc_tipo_actual"] == "Empty" and "Empty" in Info_Actual.keys() and Info_Actual["Empty"] == "Ready to Load":
                Opciones += "<button class='btn btn-sm btn-warning p-0 ps-1 pe-1' onclick='Opciones("+str(Contenedor["cc_id"])+",\""+str(Contenedor["cc_contenedor"])+"\")'><i class='mdi mdi-pencil'></i></button>"
            else:
                Opciones += "<button class='btn btn-sm btn-warning p-0 ps-1 pe-1' onclick='Opciones("+str(Contenedor["cc_id"])+",\""+str(Contenedor["cc_contenedor"])+"\")'><i class='mdi mdi-pencil'></i></button>"
            
            Ruta = ""
            Ruta += "<span class='h-100 border' style='width:20px; background:var(--Color_"+str(Contenedor["cc_tipo_actual"])+");Color:var(--Color_"+str(Contenedor["cc_tipo_actual"])+");'>__</span>"
            try:
                Tipo = Info_Actual[Contenedor["cc_tipo_actual"]]
                Ruta += str(Tipo)
                Tipo_1 = Info_Actual[Tipo]
                Ruta += "/"+str(Tipo_1)
                Tipo_2 = Info_Actual[Tipo_1]
                Ruta += "/"+str(Tipo_2)
            except:
                pass
                
            if "Fecha_Salida" in Info_Actual.keys():
                    Ruta += " <"+str(Info_Actual["Fecha_Salida"])+">"

            Sello = ""
            if "Sello Proveedor" in Info_Actual.keys():
                Sello = "<span style='color:blue'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Proveedor"])
            if "Sello Temporal" in Info_Actual.keys():
                Sello = "<span style='color:#7a7a7a'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Temporal"])
            if "Sello Blanco" in Info_Actual.keys():
                Sello = "<span><i class='mdi mdi-label-outline'></i></span> "+str(Info_Actual["Sello Temporal"])
            if "Sello Rojo" in Info_Actual.keys():
                Sello = "<span style='color:#ff0000'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Temporal"])
            
            Estado = ""
            if "Etapa" not in Info_Actual.keys():
                Estado = "Waiting"
            else:
                Estado = Info_Actual["Etapa"]
            
            for K in Contenedor.keys():
                if Contenedor[K] is None:
                    Contenedor[K] = ""

            Dock.append({"Dock":str(Contenedor["cc_dock"]),"Tiempo":"<span class='tiempo' tiempo='"+str(Contenedor["cc_ultimo_mov"])+"'></span>","Caja":str(Contenedor["cc_contenedor"]),"Carrier":Info_Actual["Carrier"],"Tipo":Contenedor["cc_tipo_actual"],"Estado":Estado,"Ruta":Ruta,"Opciones":"<div class='text-center'>"+str(Opciones)+"</div>"})
        Contenido += """<div class='text-end pe-1 pt-1'><small class='link-primary' style='cursor:pointer' onclick='Llamar_Funcion(\""""+str(request.url)+"""\");'>Actualizar <i class='mdi mdi-refresh'></i></small></div>"""
        Contenido += "<div class='h2 fw-lighter mt-1 mb-1 text-center border-bottom'><i class='mdi mdi-truck-trailer'></i> Contenedores</div>"
        Contenido += """
        <div class='text-end mb-1'> <button class='btn btn-success' onclick='Nueva_Caja()'><i class='mdi mdi-plus'></i> Nuevo Contenedor</button> </div>
        <div class='row'>
            <div class='col-6'>
                <div class='h3 text-center'><i class='mdi mdi-car-brake-parking'></i> Patio</div>
                <div id='Tabla_Patio' class='border border-dark bg-dark-subtle'></div>
                <script>
                    delete Tabla_Patio;
                    var Tabla_Patio = new Tabulator("#Tabla_Patio", {
                        minHeight:800,
                        layout:"fitColumns",
                        data:"""+str(Patio)+""",
                        columns:[
                            {field:"Tiempo","title":"Ultimo Movement",formatter:"html"},
                            {field:"Caja","title":"Contenedor/Caja",headerFilter:"input"},
                            {field:"Carrier","title":"Carrier"},
                            {field:"Tipo","title":"Tipo"},
                            {field:"Ruta","title":"Ruta",formatter:"html"},
                            {field:"Sello","title":"Sello",formatter:"html"},
                            {field:"Estado","title":"Estado",formatter:"html"},
                            {field:"Opciones","title":"Op",formatter:"html"}
                        ]
                    });
                    Tabla_Patio.on("tableBuilt", function(){ 
                        setTimeout(() => {
                            Tabla_Patio.clearFilter();
                            Tabla_Patio.redraw();
                        }, 100);
                    });
                </script>
            </div>
            <div class='col-6'>
                <div class='h3 text-center'><i class='mdi mdi-sign-caution'></i> Docks</div>
                <div id='Tabla_Docks' class='border border-dark bg-dark-subtle'></div>
                <script>
                    delete Tabla_Docks;
                    var Tabla_Docks = new Tabulator("#Tabla_Docks", {
                        minHeight:800,
                        layout:"fitColumns",
                        data:"""+str(Dock)+""",
                        columns:[
                            {field:"Dock","title":"Dock"},
                            {field:"Tiempo","title":"Ultimo Movement",formatter:"html"},
                            {field:"Caja","title":"Contenedor/Caja",headerFilter:"input"},
                            {field:"Carrier","title":"Carrier"},
                            {field:"Tipo","title":"Tipo"},
                            {field:"Ruta","title":"Ruta",formatter:"html"},
                            {field:"Opciones","title":"Op",formatter:"html"}
                        ]
                    });
                    Tabla_Docks.on("tableBuilt", function(){ 
                        setTimeout(() => {
                            Tabla_Docks.clearFilter();
                            Tabla_Docks.redraw();
                        }, 100);
                    });
                </script>
            </div>
        </div>
        <script>
            function Ver_Historico(ID,Contenedor){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-history'></i> History ["+Contenedor+"]");
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm').addClass('modal-xl');
                var parametros = {"Fun":'"""+str(fernet.encrypt("Ver_Historico".encode()).decode("utf-8"))+"""',"ID":ID};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });
            }
            function Modificar(ID,Contenedor){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-pencil'></i> Modify ["+Contenedor+"]");
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm')
                var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar".encode()).decode("utf-8"))+"""',"ID":ID};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });
            }
            function Opciones(ID,Contenedor){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-menu'></i> Options ["+Contenedor+"]");
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm')
                var parametros = {"Fun":'"""+str(fernet.encrypt("Opciones".encode()).decode("utf-8"))+"""',"ID":ID};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });
            }
            function Regresar(ID,Contenedor){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-history'></i> Return previous ["+Contenedor+"]");
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm').addClass('modal-xl');
                var parametros = {"Fun":'"""+str(fernet.encrypt("Regresar".encode()).decode("utf-8"))+"""',"ID":ID};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });
            }
            function Imprimir_Pase(ID,Contenedor){
                Mostrar_Ventana_Cargando(false);
                var parametros = {"Fun":'"""+str(fernet.encrypt("Imprimir_Pase".encode()).decode("utf-8"))+"""',"ID":ID};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        if(Resultado["Estado"] == 1){
                            if("Pase Salida" in Resultado)
                            {
                                var win1 = window.open(Resultado["Pase Salida"], '_blank');
                                win1.focus();
                            }
                            if("Manifiesto" in Resultado)
                            {
                                var win1 = window.open(Resultado["Manifiesto"], '_blank');
                                win1.focus();
                            }
                            Mensaje(2);
                        }else{
                            Mensaje(0,Resultado["Contenido"]);
                        }
                        
                    },
                    error: function (jqXHR, textStatus, errorThrown ){Mensaje(0,textStatus);}
                });
            }
            function Nueva_Caja(){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-plus'></i> Nuevo Contenedor");
                $("#Vent_1").removeClass('modal-xl modal-lg modal-sm');
                var parametros = {"Fun":'"""+str(fernet.encrypt("Nueva_Caja".encode()).decode("utf-8"))+"""'};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });
            }
            function Salida(ID,Contenedor){
                Swal.fire({
                title: '¿Estás seguro de darle salida al contenedor ['+Contenedor+']?',
                buttonsStyling: false,showCancelButton: true,confirmButtonText: "<i class='mdi mdi-check'></i> Si",cancelButtonText: "<i class='mdi mdi-close'></i> No",showLoaderOnConfirm: true,
                customClass: {confirmButton: 'btn btn-success ms-1 me-1',cancelButton: 'btn btn-danger ms-1 me-1'},
                preConfirm: () => {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Salida".encode()).decode("utf-8"))+"""',"ID":ID};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                $("#Vent_1").modal("hide");
                                Mensaje(2);
                                Llamar_Funcion(\""""+str(request.url)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
                })
            }
        </script>
        """
        
        Cur += render_template("general.html",Contenido=Contenido,Componentes=Compartido.Complementos(None),Menu=Menu,Titulo=Titulo)
    except:
        Cur += str(sys.exc_info())
    return Cur

def Tipo_Nuevo(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        if "Sin_Guardar" not in Datos.keys():
            Datos["Sin_Guardar"] = 1
        Info_Actual = json.loads(str(Datos["Info_Actual"]))
        De_Donde = str(Datos["De_Donde"])

        Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
        if str(Datos["Donde"]) == "Dock":
            Dock = []
            Dock_Ya = ""
            DOCK_DISPONIBLES = DB.Get_Dato("""
            SELECT * FROM """+str(BD_Nombre)+""".cdock_asignacion DOCK 
            left join """+str(BD_Nombre)+""".ccajas CAJAS on CAJAS.cc_activo = 1 and CAJAS.cc_dock = DOCK.cd_dock and CAJAS.cc_id != '"""+str(Datos["ID"])+"""'
            where DOCK.cd_ub = '"""+str(Bandera_Dock)+"""' and cd_activo = 1 and CAJAS.cc_id is null
            """)
            for D in DOCK_DISPONIBLES:
                Dock.append(str(D["cd_dock"]))
            if "Dock" in Info_Actual.keys():
                Dock_Ya = Info_Actual["Dock"]
            Formulario["Campos"].append({"tipo":"seleccion","campo":"Dock","titulo":"Dock","Requerido":1,"Tipo_Opciones":"Opciones","Opciones":Dock,"valor":Dock_Ya,"Col":12})

        Info = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo = '"+str(Datos["Tipo"])+"'")[0]["cr_niveles"]
        if Info is not None:
            Info = json.loads(Info)
            Opciones = []
            for K in Info.keys():
                Opciones.append(K)
            
            Formulario["Campos"].append({"tipo":"seleccion","id":"Tipo_Nuevo_1","campo":str(Datos["Tipo"]),"titulo":str(Datos["Tipo"]),"Requerido":1,"Tipo_Opciones":"Opciones","Opciones":Opciones,"valor":""})
            
            if str(Datos["Tipo"]) == "Inbound":
                if De_Donde == "Patio" and str(Datos["Donde"]) == "Patio":
                    if "Sello Proveedor" in Info_Actual.keys():
                        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Proveedor","titulo":"Supplier Seal","Requerido":1,"min":1,"max":150,"valor":str(Info_Actual["Sello Proveedor"])})
                    elif "Sello Temporal" in Info_Actual.keys():
                        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Proveedor","titulo":"Sello Temporal","Requerido":1,"min":1,"max":150,"valor":str(Info_Actual["Sello Temporal"])})
                    else:
                        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Proveedor","titulo":"Supplier Seal","Requerido":1,"min":1,"max":150,"valor":""})
                
                if De_Donde == "Patio" and str(Datos["Donde"]) == "Dock":
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                if De_Donde == "Dock" and str(Datos["Donde"]) == "Dock":
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
            
            if str(Datos["Tipo"]) == "Outbound":
                if De_Donde == "Patio" and str(Datos["Donde"]) == "Dock":
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                if De_Donde == "Dock" and str(Datos["Donde"]) == "Dock":
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
            
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))

            
            Resultado["Contenido"] += "<div id='Opciones_1'></div>"
            Resultado["Contenido"] += """
            <script>
                Actualizar_Cambia_Texto();
                $("#Tipo_Nuevo_1").on( "change", function() {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Tipo_Nuevo_1".encode()).decode("utf-8"))+"""',"Tipo":$(this).find('option:selected').val(),"Nivel_1":'"""+str(Datos["Tipo"])+"""',"Donde":'"""+str(Datos["Donde"])+"""',"ID":'"""+str(Datos["ID"])+"""',"Info_Actual":JSON.stringify("""+str(Info_Actual)+"""),"De_Donde":'"""+str(De_Donde)+"""',"Sin_Guardar":'"""+str(Datos["Sin_Guardar"])+"""'};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            $("#Opciones_1").html(Resultado["Contenido"]);
                            swal.close();
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            $("#Opciones_1").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                            swal.close();
                        }
                    });

                } );
            """
            if str(Datos["Tipo"]) in Info_Actual.keys():
                Resultado["Contenido"] += """
                $( document ).ready(function() {
                    $("#Tipo_Nuevo_1").val(\""""+str(Info_Actual[str(Datos["Tipo"])])+"""\").trigger("change");
                });
                """
            Resultado["Contenido"] += """
            </script>
            """
        else:
            if str(Datos["Tipo"]) == "Inbound":
                if De_Donde == "Patio" and str(Datos["Donde"]) == "Patio":
                    if "Sello Proveedor" in Info_Actual.keys():
                        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Proveedor","titulo":"Supplier Seal","Requerido":1,"min":1,"max":150,"valor":str(Info_Actual["Sello Proveedor"])})
                    else:
                        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Proveedor","titulo":"Supplier Seal","Requerido":1,"min":1,"max":150,"valor":""})
                if De_Donde == "Patio" and str(Datos["Donde"]) == "Dock":
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                if De_Donde == "Dock" and str(Datos["Donde"]) == "Dock":
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})

            if str(Datos["Tipo"]) == "Outbound":
                if De_Donde == "Patio" and str(Datos["Donde"]) == "Patio":
                    if "Sello Temporal" in Info_Actual.keys():
                        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Temporal","titulo":"Temporary Seal","Requerido":1,"min":1,"max":150,"valor":str(Info_Actual["Sello Temporal"])})
                    else:
                        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Temporal","titulo":"Temporary Seal","Requerido":1,"min":1,"max":150,"valor":""})
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                if De_Donde == "Patio" and str(Datos["Donde"]) == "Dock":
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                if De_Donde == "Dock" and str(Datos["Donde"]) == "Dock":
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
            
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            
            
            Resultado["Contenido"] += """
            <hr>
            <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_Guardar("""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
            """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Tipo_Nuevo_1(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        if "Sin_Guardar" not in Datos.keys():
            Datos["Sin_Guardar"] = 1
        Info_Actual = json.loads(str(Datos["Info_Actual"]))
        De_Donde = str(Datos["De_Donde"])
        
        Info = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo = '"+str(Datos["Nivel_1"])+"'")[0]["cr_niveles"]
        Info = json.loads(Info)
        Opciones = []
        if Info[Datos["Tipo"]] == 0 or "NOMBRE" in Info[Datos["Tipo"]].keys():
            if De_Donde == "Patio" and str(Datos["Donde"]) == "Patio":
                if str(Datos["Nivel_1"]) == "Empty" and str(Datos["Tipo"]) == "Return Empty":
                    Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                    Formulario["Campos"].append({"tipo":"checkbox","campo":"Salida","titulo":"¿Lista para salir? (GENERA PASA DE SALIDA)","Requerido":1,"valor":True,"editable":False})
                    Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            if De_Donde == "Patio" and str(Datos["Donde"]) == "Dock":
                if str(Datos["Nivel_1"]) == "Empty":
                    Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                    Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            if De_Donde == "Dock" and str(Datos["Donde"]) == "Dock":
                if str(Datos["Nivel_1"]) == "Empty":
                    Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
                    Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                    Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            if int(Datos["Sin_Guardar"]) == 1:
                Resultado["Contenido"] += """
                <hr>
                <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_Guardar("""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
                """
        else:
            for K in Info[Datos["Tipo"]].keys():
                Opciones.append(K)
            Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
            Formulario["Campos"].append({"tipo":"seleccion","id":"Tipo_Nuevo_2","campo":str(Datos["Tipo"]),"titulo":str(Datos["Tipo"]),"Requerido":1,"Tipo_Opciones":"Opciones","Opciones":Opciones,"valor":""})
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            Resultado["Contenido"] += "<div id='Opciones_2'></div>"
            Resultado["Contenido"] += """
            <script>
                Actualizar_Cambia_Texto();
                $("#Tipo_Nuevo_2").on( "change", function() {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Tipo_Nuevo_2".encode()).decode("utf-8"))+"""',"Tipo":$(this).find('option:selected').val(),"Nivel_1":'"""+str(Datos["Nivel_1"])+"""',"Nivel_2":'"""+str(Datos["Tipo"])+"""',"Donde":'"""+str(Datos["Donde"])+"""',"ID":'"""+str(Datos["ID"])+"""',"Info_Actual":JSON.stringify("""+str(Info_Actual)+"""),"De_Donde":'"""+str(De_Donde)+"""',"Sin_Guardar":'"""+str(Datos["Sin_Guardar"])+"""'};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            $("#Opciones_2").html(Resultado["Contenido"]);
                            swal.close();
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            $("#Opciones_2").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                            swal.close();
                        }
                    });

                } );
            """
            if str(Datos["Tipo"]) in Info_Actual.keys():
                Resultado["Contenido"] += """
                $( document ).ready(function() {
                    $("#Tipo_Nuevo_2").val(\""""+str(Info_Actual[str(Datos["Tipo"])])+"""\").trigger("change");
                });
                """
            Resultado["Contenido"] += """
            </script>
            """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Tipo_Nuevo_2(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Info_Actual = json.loads(str(Datos["Info_Actual"]))
        De_Donde = str(Datos["De_Donde"])
        if Datos["Nivel_1"] == "Outbound":
            Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
            if "Fecha_Salida" in Info_Actual.keys():
                Formulario["Campos"].append({"tipo":"fecha","campo":"Fecha_Salida","titulo":"Fecha de Salida","Requerido":1,"valor":str(Info_Actual["Fecha_Salida"]),"editable":False})
            else:
                Formulario["Campos"].append({"tipo":"fecha","campo":"Fecha_Salida","titulo":"Fecha de Salida","Requerido":1,"valor":""})
            if De_Donde == "Patio" and str(Datos["Donde"]) == "Patio":
                Formulario["Campos"].append({"tipo":"checkbox","campo":"Salida","titulo":"¿Lista para salir? (GENERA PASA DE SALIDA)","Requerido":1,"valor":False})
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            if De_Donde == "Patio" and str(Datos["Donde"]) == "Patio":
                Resultado["Contenido"] += """
                <div id='Foto_Lista_Salir'></div>
                <script>
                    $(document).ready(function() {
                        $('#checkbox_1').change(function() {
                            if(this.checked) {
                                Mostrar_Ventana_Cargando(false);
                                var parametros = {"Fun":'"""+str(fernet.encrypt("Foto_Lista_Salir".encode()).decode("utf-8"))+"""'};
                                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                                    success:  function (response)
                                    {
                                        var Resultado = JSON.parse(response);
                                        $("#Foto_Lista_Salir").html(Resultado["Contenido"]);
                                        swal.close();
                                    },
                                    error: function (jqXHR, textStatus, errorThrown )
                                    {
                                        $("#Foto_Lista_Salir").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                                        swal.close();
                                    }
                                });
                            }else{
                                $("#Foto_Lista_Salir").html("");
                            }      
                        });
                    });
                </script>
                """

        if "Sin_Guardar" not in Datos.keys():
            Datos["Sin_Guardar"] = 1
       

        Info = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo = '"+str(Datos["Nivel_1"])+"'")[0]["cr_niveles"]
        Info = json.loads(Info)
        Opciones = []
        if Info[Datos["Nivel_2"]][Datos["Tipo"]] == 0:
            if int(Datos["Sin_Guardar"]) == 1:
                Resultado["Contenido"] += """
                <hr>
                <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_Guardar("""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
                """
        else:
            if "DOCKS" in Info[Datos["Nivel_2"]][Datos["Tipo"]].keys() and len(Info[Datos["Nivel_2"]][Datos["Tipo"]]["DOCKS"]) > 0:
                for K in Info[Datos["Nivel_2"]][Datos["Tipo"]]["DOCKS"]:
                    Opciones.append(K)
                Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
                if str(Datos["Tipo"]) in Info_Actual.keys():
                    Formulario["Campos"].append({"tipo":"seleccion","id":"Tipo_Nuevo_3","campo":str(Datos["Tipo"]),"titulo":str(Datos["Tipo"]),"Requerido":1,"Tipo_Opciones":"Opciones","Opciones":Opciones,"valor":Info_Actual[str(Datos["Tipo"])]})
                else:
                    Formulario["Campos"].append({"tipo":"seleccion","id":"Tipo_Nuevo_3","campo":str(Datos["Tipo"]),"titulo":str(Datos["Tipo"]),"Requerido":1,"Tipo_Opciones":"Opciones","Opciones":Opciones,"valor":""})
                Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))

            if int(Datos["Sin_Guardar"]) == 1:
                Resultado["Contenido"] += """
                <hr>
                <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_Guardar("""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
                """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    print(Cur)
def Modificar(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Caja = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
        Info_Actual = json.loads(str(Caja["cc_informacion_actual"]))
        De_Donde = Caja["cc_ubicacion"]
        Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
        Formulario["Campos"].append({"tipo":"texto","campo":"Container","titulo":"Contenedor","Requerido":1,"min":1,"max":150,"valor":str(Caja["cc_contenedor"])})
        Formulario["Campos"].append({"tipo":"seleccion","campo":"Carrier","titulo":"Carrier","Requerido":1,"Tipo_Opciones":"Query","Opciones":"SELECT cca_nombre as Valor, cca_nombre as Texto FROM "+str(BD_Nombre)+".ccarrier WHERE cca_activo = '1'","valor":str(Info_Actual["Carrier"]),"Col":12})
        Formulario["Campos"].append({"tipo":"seleccion","id":"Donde_Nuevo","campo":"Donde","titulo":"Ubicacion","Requerido":1,"Tipo_Opciones":"Opciones","Opciones":["Patio","Dock"],"valor":str(De_Donde),"Col":12})
        Formulario["Campos"].append({"tipo":"seleccion","id":"Tipo_Nuevo","campo":"Tipo","titulo":"Tipo","Requerido":1,"Tipo_Opciones":"Query","Opciones":"SELECT cr_tipo as Valor,cr_tipo as Texto FROM "+str(BD_Nombre)+".crutas","valor":"","Col":12})
        
        if Caja["cc_tipo_actual"] is None:
            Resultado["Contenido"] += """
            <div class='text-center w-100'><button class='btn btn-warning' onclick='Error_Ingreso("""+str(Datos["ID"])+""",\""""+str(Caja["cc_contenedor"])+"""\")'><i class='mdi mdi-alert'></i> Error de ingreso (<span class='text-danger fw-bold'>GENERAR PASE DE SALIDA</span>)</button></div>
            <hr>
            """
        
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += "<div id='Opciones'></div>"
        Resultado["Contenido"] += """
        <script>
            Actualizar_Cambia_Texto();
            $("#Donde_Nuevo").on( "change", function() {
                $("#Opciones").html("");
                $("#Tipo_Nuevo").val("");
                Actualizar_Cambia_Texto();
        """
        if "Tipo" in Info_Actual.keys():
            Resultado["Contenido"] += """
                $("#Tipo_Nuevo").val(\""""+str(Info_Actual["Tipo"])+"""\").trigger("change");
            """
        Resultado["Contenido"] += """
            } );
            $("#Tipo_Nuevo").on( "change", function() {
                Mostrar_Ventana_Cargando(false);
                var parametros = {"Fun":'"""+str(fernet.encrypt("Tipo_Nuevo".encode()).decode("utf-8"))+"""',"Tipo":$(this).find('option:selected').val(),"ID":'"""+str(Datos["ID"])+"""',"Donde":$("#Donde_Nuevo").val(),"Info_Actual":JSON.stringify("""+str(Info_Actual)+"""),"De_Donde":'"""+str(De_Donde)+"""'};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Opciones").html(Resultado["Contenido"]);
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Opciones").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });

            } );
            function Modificar_Guardar(ID){
                var Info = Dame_Formulario(".Asignar",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_Guardar".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID":ID};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                $("#Vent_1").modal("hide");
                                Mensaje(2);
                                Llamar_Funcion(\""""+str(request.url)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
            }
            function Error_Ingreso(ID,Contenedor){
                Swal.fire({
                    title: '¿Estás seguro de generar el pase de salida del contenedor ['+Contenedor+']?',
                    buttonsStyling: false,showCancelButton: true,confirmButtonText: "<i class='mdi mdi-check'></i> Yes",cancelButtonText: "<i class='mdi mdi-close'></i> No",showLoaderOnConfirm: true,
                    customClass: {confirmButton: 'btn btn-success ms-1 me-1',cancelButton: 'btn btn-danger ms-1 me-1'},
                    preConfirm: () => {
                        
                        Mostrar_Ventana_Cargando(false);
                            var parametros = {"Fun":'"""+str(fernet.encrypt("Error_Ingreso".encode()).decode("utf-8"))+"""',"ID":ID};
                            $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                                success:  function (response)
                                {
                                    var Resultado = JSON.parse(response);
                                    if(Resultado["Estado"] == 1)
                                    {
                                        $("#Vent_1").modal("hide");
                                        Mensaje(2);
                                        Llamar_Funcion(\""""+str(request.url)+"""\");
                                    }
                                    else
                                        Mensaje(0,Resultado["Contenido"]);
                                        
                                },
                                error: function (jqXHR, textStatus, errorThrown )
                                {
                                    Mensaje(0,textStatus);
                                }
                            });
                        
                    }
                })
            }
        """
        if "Tipo" in Info_Actual.keys():
            Resultado["Contenido"] += """
            $( document ).ready(function() {
                $("#Tipo_Nuevo").val(\""""+str(Info_Actual["Tipo"])+"""\").trigger("change");
            });
            """
        Resultado["Contenido"] += """
        </script>
        """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Modificar_Guardar(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Salida = 0
        Datos_Info = json.loads(str(Datos["Info"]))
        Contenedos = Datos_Info["Container"]
        del Datos_Info["Container"]
        if "Salida" in Datos_Info.keys():
            Salida = Datos_Info["Salida"]
            del Datos_Info["Salida"]
        
        Ubicacion = "Patio"
        Dock = "null"
        if "Dock" in Datos_Info.keys():
            Ubicacion = "Dock"
            Dock = "'"+str(Datos_Info["Dock"])+"'"
        Antes_De = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]

        Error = ""
        if Salida == 0:
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_contenedor = '"+str(Contenedos)+"', cc_informacion_actual = '"+str(json.dumps(Datos_Info))+"',cc_tipo_actual = '"+str(Datos_Info["Tipo"])+"',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"' WHERE cc_id = '"+str(Datos["ID"])+"' ")
        else:
            Codigo = str(hashlib.md5(str("LOGISTICS INSIGHT CORPORATION - PASE SALIDA - "+str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))+" - "+str(Datos["ID"])).encode()).hexdigest())
            Folio = ''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(8))
            Folio += str(Datos["ID"])
            Datos_Info["Etapa"] = "Pending Exit"
            if Codigo is None:
                Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_contenedor = '"+str(Contenedos)+"', cc_informacion_actual = '"+str(json.dumps(Datos_Info))+"',cc_tipo_actual = '"+str(Datos_Info["Tipo"])+"',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"',cc_bloquear = 1 WHERE cc_id = '"+str(Datos["ID"])+"' ")
            else:
                Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_contenedor = '"+str(Contenedos)+"', cc_informacion_actual = '"+str(json.dumps(Datos_Info))+"',cc_tipo_actual = '"+str(Datos_Info["Tipo"])+"',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"',cc_qr_salida = '"+str(Codigo)+"', cc_folio_salida = '"+str(Folio)+"',cc_bloquear = 1 WHERE cc_id = '"+str(Datos["ID"])+"' ")

        if Error == "":
            Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
            Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
            Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
            Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
            Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"

            if Salida == 1:
                TIPO_MOV = "LIBERA OPERACIONES"
            else:
                if Antes_De["cc_tipo_actual"] is None:
                    TIPO_MOV = "ASIGNAR"
                else:
                    TIPO_MOV = "MODIFICAR"

            Error += DB.Instruccion(""" 
            INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
            (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
            VALUES
            ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','"""+str(TIPO_MOV)+"""')
            """)
            if Error == "":
                Resultado["Estado"] = 1
            else:
                Resultado["Contenido"] = Error
        else:
            Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Opciones(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Caja = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
        Info_Actual = json.loads(str(Caja["cc_informacion_actual"]))
        De_Donde = Caja["cc_ubicacion"]
        if Caja["cc_tipo_actual"] == "Inbound":
            Resultado["Contenido"] = """
            <div>Opciones de Contenedor/Caja</div>
            <div class='row'>
                <div class='col-12 p-1'>
                    <button class='btn btn-dark w-100 btn-lg' onclick='Modificar_A("Ready to Load","","""+str(Datos["ID"])+""");'><div>Empty [Ready to Load]</div><div><small>(Dejar caja en Dock)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-dark w-100 btn-lg' onclick='Modificar_A("Ready to Load","Patio","""+str(Datos["ID"])+""");'><div>Empty [Ready to Load]</div><div><small>(Mover caja a patio)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-dark w-100 btn-lg' onclick='Modificar_A("Return Empty","Patio","""+str(Datos["ID"])+""");'><div>Empty [Return empty]</div><div><small>(Mover caja a patio <span class='text-danger fw-bold'>LISTA PARA SALIR</span>)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-primary w-100 btn-lg' onclick='Modificar_A("Inbound","Patio","""+str(Datos["ID"])+""");'><div>Inbound</div><div><small>(Mover caja a patio <span class='text-warning fw-bold'>CAJA AUN CON MATERIAL DE PROVEEDOR</span>)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-info w-100 btn-lg' onclick='Evidencia("""+str(Datos["ID"])+""");'><div> Add evidence </div><div><small></small></div></button>
                </div>
            </div>
            <div>Opciones de Material</div>
            <div class='row'>
                <div class='col-6 p-1'>
                    <button class='btn btn-warning w-100 btn-lg' onclick='Generar_OSyD("Exceso de Material","""+str(Datos["ID"])+""")'>Exceso de Material</button>
                </div>
                <div class='col-6 p-1'>
                    <button class='btn btn-warning w-100 btn-lg' onclick='Generar_OSyD("Faltante de Material","""+str(Datos["ID"])+""")'>Faltante de Material</button>
                </div>
                <div class='col-6 p-1'>
                    <button class='btn btn-warning w-100 btn-lg' onclick='Generar_OSyD("Material Dañado","""+str(Datos["ID"])+""")'>Material Dañado</button>
                </div>
                <div class='col-6 p-1'>
                    <button class='btn btn-warning w-100 btn-lg' onclick='Generar_OSyD("Otra Discrepancia","""+str(Datos["ID"])+""")'>Otra Discrepancia</button>
                </div>
            </div>
            <script>
                function Modificar_A(Poner_En,A_Donde,ID){
                    Mostrar_Ventana_Cargando(false);
                    $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-pencil'></i> "+ Poner_En);
                    $("#Vent_1").removeClass('modal-xl modal-lg modal-sm')
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_A".encode()).decode("utf-8"))+"""',"ID":ID,"Poner_En":Poner_En,"A_Donde":A_Donde};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                            $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                            swal.close();
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                            swal.close();
                        }
                    });
                }
                function Generar_OSyD(Tipo,ID){
                    Mostrar_Ventana_Cargando(false);
                    $("#Vent_2").find(".modal-title").html("<i class='mdi mdi-plus'></i> "+ Tipo);
                    $("#Vent_2").removeClass('modal-xl modal-lg modal-sm')
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Generar_OSyD".encode()).decode("utf-8"))+"""',"ID":ID,"Tipo":Tipo};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            $("#Vent_2").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                            $("#Vent_2").find(".modal-footer").find("button").attr('onclick',"$('#Vent_2').modal('hide'); delete table; ")
                            swal.close();
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            $("#Vent_2").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                            swal.close();
                        }
                    });
                }
            </script>
            """
        if Caja["cc_tipo_actual"] == "Outbound":
            Resultado["Contenido"] = """
            <div>Opciones de Contenedor</div>
            <div class='row'>
                <div class='col-12 p-1'>
                    <button class='btn btn-dark w-100 btn-lg' onclick='Modificar_A("Ready to Load","","""+str(Datos["ID"])+""");'><div>Empty [Ready to Load]</div><div><small>(Asignación incorrecta)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-warning w-100 btn-lg' onclick='Modificar_A("Outbound","Patio","""+str(Datos["ID"])+""");'><div>Outbound</div><div><small>(Mover Caja a Patio <span class='fw-bold text-danger'>AÚN FALTA MATERIAL POR CARGAR</span>)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-warning w-100 btn-lg' onclick='Modificar_A("Outbound","Patio","""+str(Datos["ID"])+""",1);'><div>Outbound</div><div><small>(Mover Caja a Patio <span class='fw-bold text-danger'>LISTA PARA SALIR</span>)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-info w-100 btn-lg' onclick='Evidencia("""+str(Datos["ID"])+""",\""""+str(Caja["cc_contenedor"])+"""\");'><div> Add evidence </div><div><small></small></div></button>
                </div>
            </div>
            <script>
                function Modificar_A(Poner_En,A_Donde,ID,Salida=0){
                    Mostrar_Ventana_Cargando(false);
                    $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-pencil'></i> "+ Poner_En);
                    $("#Vent_1").removeClass('modal-xl modal-lg modal-sm')
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_A".encode()).decode("utf-8"))+"""',"ID":ID,"Poner_En":Poner_En,"A_Donde":A_Donde,"Salida":Salida};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                            $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                            swal.close();
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                            swal.close();
                        }
                    });
                }
            </script>
            """
        if Caja["cc_tipo_actual"] == "Empty":
            Resultado["Contenido"] = """
            <div>Opciones de Containedor</div>
            <div class='row'>
                <div class='col-12 p-1'>
                    <button class='btn btn-dark w-100 btn-lg' onclick='Modificar_A("Ready to Load","Patio","""+str(Datos["ID"])+""");'><div>Empty [Ready to Load]</div><div><small>(Mover caja a patio)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-dark w-100 btn-lg' onclick='Modificar_A("Return Empty","Patio","""+str(Datos["ID"])+""");'><div>Empty [Return empty]</div><div><small>(Mover caja a patio <span class='text-danger fw-bold'>LISTA PARA SALIR</span>)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-warning w-100 btn-lg' onclick='Modificar_A("Outbound","Dock","""+str(Datos["ID"])+""");'><div>Outbound</div><div><small>(Dejar en dock <span class='text-danger fw-bold'>DEFINIR RUTA OUTBOUND</span>)</small></div></button>
                </div>
                <div class='col-12 p-1'>
                    <button class='btn btn-info w-100 btn-lg' onclick='Evidencia("""+str(Datos["ID"])+""",\""""+str(Caja["cc_contenedor"])+"""\");'><div> Add evidence </div><div><small></small></div></button>
                </div>
            </div>
            <script>
                function Modificar_A(Poner_En,A_Donde,ID,Salida=0){
                    Mostrar_Ventana_Cargando(false);
                    $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-pencil'></i> "+ Poner_En);
                    $("#Vent_1").removeClass('modal-xl modal-lg modal-sm')
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_A".encode()).decode("utf-8"))+"""',"ID":ID,"Poner_En":Poner_En,"A_Donde":A_Donde,"Salida":Salida};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                            $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                            swal.close();
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                            swal.close();
                        }
                    });
                }
                function Caja_Danada(ID,Contenedor){
                     Mostrar_Ventana_Cargando(false);
                    $("#Vent_1").find(".modal-title").html("<i class='mdi mdi-pipe-leak'></i> "+ Contenedor);
                    $("#Vent_1").removeClass('modal-xl modal-lg modal-sm')
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Caja_Danada".encode()).decode("utf-8"))+"""',"ID":ID};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            $("#Vent_1").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                            $("#Vent_1").find(".modal-footer").find("button").attr('onclick',"$('#Vent_1').modal('hide'); delete table; ")
                            swal.close();
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            $("#Vent_1").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                            swal.close();
                        }
                    });
                }
            </script>
            """
        Resultado["Contenido"] += """
        <script>
            function Evidencia(ID,Contenedor){
                Mostrar_Ventana_Cargando(false);
                $("#Vent_2").find(".modal-title").html("<i class='mdi mdi-camera-burst'></i> "+ Contenedor);
                $("#Vent_2").removeClass('modal-xl modal-lg modal-sm')
                var parametros = {"Fun":'"""+str(fernet.encrypt("Evidencia".encode()).decode("utf-8"))+"""',"ID":ID};
                $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                    success:  function (response)
                    {
                        var Resultado = JSON.parse(response);
                        $("#Vent_2").modal("show").find(".modal-body").html(Resultado["Contenido"]);
                        $("#Vent_2").find(".modal-footer").find("button").attr('onclick',"$('#Vent_2').modal('hide'); delete table; ")
                        swal.close();
                    },
                    error: function (jqXHR, textStatus, errorThrown )
                    {
                        $("#Vent_2").modal("show").find(".modal-body").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                        swal.close();
                    }
                });
            }
        </script>
        """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Modificar_A(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        if "Salida" not in Datos.keys():
            Datos["Salida"] = 0
        if "A_Donde" not in Datos.keys():
            Datos["A_Donde"] = ""
            
        Formulario = {"Col":"12", "Campos": [],"Clase": "Modificar_A" }
        Caja = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
        Info_Actual = json.loads(str(Caja["cc_informacion_actual"]))
        if Datos["Poner_En"] == "Ready to Load" and Datos["A_Donde"] == "":
            Dock = []
            Dock_Ya = ""
            DOCK_DISPONIBLES = DB.Get_Dato("""
            SELECT * FROM """+str(BD_Nombre)+""".cdock_asignacion DOCK 
            left join """+str(BD_Nombre)+""".ccajas CAJAS on CAJAS.cc_activo = 1 and CAJAS.cc_dock = DOCK.cd_dock and CAJAS.cc_id != '"""+str(Datos["ID"])+"""'
            where DOCK.cd_ub = '"""+str(Bandera_Dock)+"""' and cd_activo = 1 and CAJAS.cc_id is null
            """)
            for D in DOCK_DISPONIBLES:
                Dock.append(str(D["cd_dock"]))
            
            if "Dock" in Info_Actual.keys():
                Dock_Ya = Info_Actual["Dock"]
            Formulario["Campos"].append({"tipo":"seleccion","campo":"Dock","titulo":"Dock","Requerido":1,"Tipo_Opciones":"Opciones","Opciones":Dock,"valor":Dock_Ya,"Col":12})

            if Caja["cc_tipo_actual"] == "Empty" and "Fotos" in Info_Actual.keys():
                pass
            else:
                Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Fotos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})

            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            Resultado["Contenido"] += """
            <hr>
            <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_A_Guardar(\""""+str(Datos["Poner_En"])+"""\",\""""+str(Datos["A_Donde"])+"""\","""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
            """
        if Datos["Poner_En"] == "Ready to Load" and Datos["A_Donde"] != "":
            if Caja["cc_tipo_actual"] == "Empty" and "Fotos" in Info_Actual.keys():
                Resultado["Contenido"] += """
                <script>
                    $( document ).ready(function() {
                        Modificar_A_Guardar(\""""+str(Datos["Poner_En"])+"""\",\""""+str(Datos["A_Donde"])+"""\","""+str(Datos["ID"])+""");
                    });
                </script>
                """
            else:
                Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
                Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
                Resultado["Contenido"] += """
                <hr>
                <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_A_Guardar(\""""+str(Datos["Poner_En"])+"""\",\""""+str(Datos["A_Donde"])+"""\","""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
                """
        if Datos["Poner_En"] == "Return Empty":
            Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            Resultado["Contenido"] += """
            <hr>
            <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_A_Guardar(\""""+str(Datos["Poner_En"])+"""\",\""""+str(Datos["A_Donde"])+"""\","""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
            """
        if Datos["Poner_En"] == "Inbound":
            Formulario["Campos"].append({"tipo":"texto","campo":"Sello Temporal","titulo":"Temporary Seal","Requerido":1,"min":1,"max":150,"valor":""})
            Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Photos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            Resultado["Contenido"] += """
            <hr>
            <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_A_Guardar(\""""+str(Datos["Poner_En"])+"""\",\""""+str(Datos["A_Donde"])+"""\","""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
            """
        if Datos["Poner_En"] == "Outbound" and Datos["A_Donde"] == "Patio":
            Formulario["Campos"].append({"tipo":"texto","campo":"Sello Temporal","titulo":"Sello Termporal","Requerido":1,"min":1,"max":150,"valor":""})
            Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Fotos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
            if "Fecha_Salida" in Info_Actual.keys():
                Formulario["Campos"].append({"tipo":"fecha","campo":"Fecha_Salida","titulo":"Fecha de Salida","Requerido":1,"min":1,"max":30,"valor":Info_Actual["Fecha_Salida"],"editable":False})
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            Resultado["Contenido"] += """
            <hr>
            <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_A_Guardar(\""""+str(Datos["Poner_En"])+"""\",\""""+str(Datos["A_Donde"])+"""\","""+str(Datos["ID"])+""","""+str(Datos["Salida"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
            """
        if Datos["Poner_En"] == "Outbound" and Datos["A_Donde"] == "Dock":
            Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
            Dock = []
            Dock_Ya = ""
            DOCK_DISPONIBLES = DB.Get_Dato("""
            SELECT * FROM """+str(BD_Nombre)+""".cdock_asignacion DOCK 
            left join """+str(BD_Nombre)+""".ccajas CAJAS on CAJAS.cc_activo = 1 and CAJAS.cc_dock = DOCK.cd_dock and CAJAS.cc_id != '"""+str(Datos["ID"])+"""'
            where DOCK.cd_ub = '"""+str(Bandera_Dock)+"""' and cd_activo = 1 and CAJAS.cc_id is null
            """)
            for D in DOCK_DISPONIBLES:
                Dock.append(str(D["cd_dock"]))
            if "Dock" in Info_Actual.keys():
                Dock_Ya = Info_Actual["Dock"]
            Formulario["Campos"].append({"tipo":"seleccion","campo":"Dock","titulo":"Dock","Requerido":1,"Tipo_Opciones":"Opciones","Opciones":Dock,"valor":Dock_Ya,"Col":12})
            Info = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".crutas WHERE cr_tipo = 'Outbound'")[0]["cr_niveles"]
            if Info is not None:
                Info = json.loads(Info)
                Opciones = []
                for K in Info.keys():
                    Opciones.append(K)
                
                Formulario["Campos"].append({"tipo":"seleccion","id":"Tipo_Nuevo_1","campo":"Outbound","titulo":"Outbound","Requerido":1,"Tipo_Opciones":"Opciones","Opciones":Opciones,"valor":""})
                Resultado["Contenido"] += """
                <script>
                    Actualizar_Cambia_Texto();
                    $("#Tipo_Nuevo_1").on( "change", function() {
                        Mostrar_Ventana_Cargando(false);
                        var parametros = {"Fun":'"""+str(fernet.encrypt("Tipo_Nuevo_1".encode()).decode("utf-8"))+"""',"Tipo":$(this).find('option:selected').val(),"Nivel_1":'Outbound',"Donde":'Dock',"ID":'"""+str(Datos["ID"])+"""',"Info_Actual":JSON.stringify("""+str(Info_Actual)+"""),"De_Donde":'Dock',"Sin_Guardar":0};
                        $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                            success:  function (response)
                            {
                                var Resultado = JSON.parse(response);
                                $("#Opciones_1").html(Resultado["Contenido"]);
                                swal.close();
                            },
                            error: function (jqXHR, textStatus, errorThrown )
                            {
                                $("#Opciones_1").html("<i class='mdi mdi-alert'></i> "+ textStatus);
                                swal.close();
                            }
                        });

                    } );
                </script>
                """
            Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Fotos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
            if "Fecha_Salida" in Info_Actual.keys():
                Formulario["Campos"].append({"tipo":"fecha","campo":"Fecha_Salida","titulo":"Fecha de Salida","Requerido":1,"min":1,"max":30,"valor":Info_Actual["Fecha_Salida"],"editable":False})
            Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
            Resultado["Contenido"] += "<div id='Opciones_1'></div>"
            Resultado["Contenido"] += """
            <hr>
            <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Modificar_A_Guardar(\""""+str(Datos["Poner_En"])+"""\",\"Dock\","""+str(Datos["ID"])+""","""+str(Datos["Salida"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
            """
        if Formulario["Clase"] == "Asignar":
            Resultado["Contenido"] += """
            <script>
                function Modificar_A_Guardar(Poner_En,A_Donde,ID,Salida){
                    var Info = Dame_Formulario(".Asignar",true);
                    if(Info != null)
                    {
                        Mostrar_Ventana_Cargando(false);
                        var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_A_Guardar".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID":ID,"Poner_En":Poner_En,"A_Donde":A_Donde,"Info_Actual":JSON.stringify("""+str(Info_Actual)+"""),"Salida":Salida};
                        $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                            success:  function (response)
                            {
                                var Resultado = JSON.parse(response);
                                if(Resultado["Estado"] == 1)
                                {
                                    if("Pase Salida" in Resultado)
                                    {
                                        var win1 = window.open(Resultado["Pase Salida"], '_blank');
                                        win1.focus();
                                    }
                                    $("#Vent_1").modal("hide");
                                    Mensaje(2);
                                    Llamar_Funcion(\""""+str(request.url)+"""\");
                                }
                                else
                                    Mensaje(0,Resultado["Contenido"]);
                                    
                            },
                            error: function (jqXHR, textStatus, errorThrown )
                            {
                                Mensaje(0,textStatus);
                            }
                        });
                    }
                }
            </script>
            """
        else:
            Resultado["Contenido"] += """
            <script>
                function Modificar_A_Guardar(Poner_En,A_Donde,ID,Salida){
                    var Info = Dame_Formulario(".Modificar_A",true);
                    if(Info != null)
                    {
                        Mostrar_Ventana_Cargando(false);
                        var parametros = {"Fun":'"""+str(fernet.encrypt("Modificar_A_Guardar".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID":ID,"Poner_En":Poner_En,"A_Donde":A_Donde,"Info_Actual":JSON.stringify("""+str(Info_Actual)+"""),"Salida":Salida};
                        $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                            success:  function (response)
                            {
                                var Resultado = JSON.parse(response);
                                if(Resultado["Estado"] == 1)
                                {
                                    if("Pase Salida" in Resultado)
                                    {
                                        var win1 = window.open(Resultado["Pase Salida"], '_blank');
                                        win1.focus();
                                    }
                                    $("#Vent_1").modal("hide");
                                    Mensaje(2);
                                    Llamar_Funcion(\""""+str(request.url)+"""\");
                                }
                                else
                                    Mensaje(0,Resultado["Contenido"]);
                                    
                            },
                            error: function (jqXHR, textStatus, errorThrown )
                            {
                                Mensaje(0,textStatus);
                            }
                        });
                    }
                }
            </script>
            """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Modificar_A_Guardar(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    Error = ""
    try:
        Info_Datos = json.loads(str(Datos["Info"]))
        Info_Actual = json.loads(str(Datos["Info_Actual"]))
        if "A_Donde" not in Datos.keys():
            Datos["A_Donde"] = ""
        if Datos["Poner_En"] == "Ready to Load" and Datos["A_Donde"] == "":
            Info_Nueva = {}
            Info_Nueva["Carrier"] = Info_Actual["Carrier"]
            Info_Nueva["Donde"] = "Dock"
            Info_Nueva["Dock"] = Info_Datos["Dock"]
            Info_Nueva["Tipo"] = "Empty"
            Info_Nueva["Empty"] = "Ready to Load"
            Info_Nueva["Fotos"] = Info_Datos["Fotos"]
            Ubicacion = "Patio"
            Dock = "null"
            if "Dock" in Info_Nueva.keys():
                Ubicacion = "Dock"
                Dock = "'"+str(Info_Nueva["Dock"])+"'"
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_tipo_actual = 'Empty',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"' WHERE cc_id = '"+str(Datos["ID"])+"' ")
            if Error == "":
                Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
                Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
                Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
                Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
                Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
                Error = DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
                (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
                VALUES
                ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','MODIFICAR')
                """)
        if Datos["Poner_En"] == "Ready to Load" and Datos["A_Donde"] != "":
            Info_Nueva = {}
            Info_Nueva["Carrier"] = Info_Actual["Carrier"]
            Info_Nueva["Donde"] = "Patio"
            Info_Nueva["Tipo"] = "Empty"
            Info_Nueva["Empty"] = "Ready to Load"
            if "Fotos" in Info_Datos.keys():
                Info_Nueva["Fotos"] = Info_Datos["Fotos"]
            Ubicacion = "Patio"
            Dock = "null"
            if "Dock" in Info_Nueva.keys():
                Ubicacion = "Dock"
                Dock = "'"+str(Info_Nueva["Dock"])+"'"
            
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_tipo_actual = 'Empty',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"' WHERE cc_id = '"+str(Datos["ID"])+"' ")

            if Error == "":
                Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
                Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
                Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
                Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
                Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
                Error = DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
                (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
                VALUES
                ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','MODIFICAR')
                """)
        if Datos["Poner_En"] == "Return Empty":
            Info_Nueva = {}
            Info_Nueva["Carrier"] = Info_Actual["Carrier"]
            Info_Nueva["Donde"] = "Patio"
            Info_Nueva["Tipo"] = "Empty"
            Info_Nueva["Empty"] = "Return Empty"
            Info_Nueva["Fotos"] = Info_Datos["Fotos"]
            Info_Nueva["Etapa"] = "Pending Exit"
            Ubicacion = "Patio"
            Dock = "null"
            if "Dock" in Info_Nueva.keys():
                Ubicacion = "Dock"
                Dock = "'"+str(Info_Nueva["Dock"])+"'"
            
            Codigo = str(hashlib.md5(str("LOGISTICS INSIGHT CORPORATION - PASE SALIDA - "+str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))+" - "+str(Datos["ID"])).encode()).hexdigest())
            Folio = ''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(8))
            Folio += str(Datos["ID"])
            
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_tipo_actual = 'Empty',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"',cc_qr_salida = '"+str(Codigo)+"',cc_folio_salida = '"+str(Folio)+"',cc_bloquear = 1 WHERE cc_id = '"+str(Datos["ID"])+"' ")
            if Error == "":
                Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
                Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
                Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
                Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
                Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
                Error += DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
                (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
                VALUES
                ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','MODIFICAR')
                """)
                #Resultado["Pase Salida"] = Generar_Pase_Salida(str(Datos["ID"]),"Pase_Salida")
        if Datos["Poner_En"] == "Inbound":
            Info_Nueva = Info_Actual
            Info_Nueva["Donde"] = "Patio"
            Info_Nueva["Fotos"] = Info_Datos["Fotos"]
            Info_Nueva["Sello Temporal"] = Info_Datos["Sello Temporal"]
            Ubicacion = "Patio"
            Dock = "null"
            
            Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"' WHERE cc_id = '"+str(Datos["ID"])+"' ")
            if Error == "":
                Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
                Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
                Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
                Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
                Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
                Error += DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
                (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
                VALUES
                ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','MODIFICAR')
                """)
        if Datos["Poner_En"] == "Outbound" and Datos["A_Donde"] == "Patio":
            Tipo_Modificacion = "MODIFICAR"
            Info_Nueva = Info_Actual
            Info_Nueva["Donde"] = "Patio"
            Info_Nueva["Fotos"] = Info_Datos["Fotos"]
            Info_Nueva["Sello Temporal"] = Info_Datos["Sello Temporal"]
            Ubicacion = "Patio"
            Dock = "null"
            Codigo = None
            Folio = None
            if int(Datos["Salida"]) == 1:
                Tipo_Modificacion = "LIBERA OPERACIONES"
                Codigo = str(hashlib.md5(str("LOGISTICS INSIGHT CORPORATION - PASE SALIDA - "+str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))+" - "+str(Datos["ID"])).encode()).hexdigest())
                Folio = ''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(8))
                Folio += str(Datos["ID"])
                Info_Nueva["Etapa"] = "Pending Exit"
                if Codigo is None:
                    Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"',cc_bloquear = 1 WHERE cc_id = '"+str(Datos["ID"])+"' ")
                else:
                    Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"',cc_qr_salida = '"+str(Codigo)+"', cc_folio_salida = '"+str(Folio)+"',cc_bloquear = 1 WHERE cc_id = '"+str(Datos["ID"])+"' ")
            else:
                Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_ultimo_mov = NOW(),cc_dock= "+str(Dock)+",cc_ubicacion = '"+str(Ubicacion)+"' WHERE cc_id = '"+str(Datos["ID"])+"' ")
            if Error == "":
                Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
                Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
                Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
                Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
                Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
                Error += DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
                (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
                VALUES
                ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','"""+str(Tipo_Modificacion)+"""')
                """)
        if Datos["Poner_En"] == "Outbound" and Datos["A_Donde"] == "Dock":
            Info_Nueva = Info_Actual
            Datos_Info = json.loads(str(Datos["Info"]))
            del Info_Nueva["Empty"]
            Info_Nueva["Tipo"] = "Outbound"
            for k in Datos_Info.keys():
                Info_Nueva[k] = Datos_Info[k]
            Error += str(Datos_Info)
            Error = DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_tipo_actual = 'Outbound',cc_ultimo_mov = NOW(),cc_dock= "+str(Datos_Info["Dock"])+",cc_ubicacion = 'Dock' WHERE cc_id = '"+str(Datos["ID"])+"' ")
            if Error == "":
                Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
                Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
                Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
                Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
                Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
                Error = DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
                (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
                VALUES
                ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','MODIFICAR')
                """)
                if Error == "":
                    Resultado["Estado"] = 1
                else:
                    Resultado["Contenido"] = Error
            else:
                Resultado["Contenido"] += str(Error)

        if Error == "":
            Resultado["Estado"] = 1
        else:
            Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Evidencia(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Formulario = {"Col":"12", "Campos": [],"Clase": "Evidencia_Guardar" }
        Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Fotos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += """
        <hr>
        <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Evidencia_Guardar("""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
        <script>
            function Evidencia_Guardar(ID){
                var Info = Dame_Formulario(".Evidencia_Guardar",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Evidencia_Guardar".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID":ID};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                if("Pase Salida" in Resultado)
                                {
                                    var win1 = window.open(Resultado["Pase Salida"], '_blank');
                                    win1.focus();
                                }
                                $("#Vent_2").modal("hide");
                                Mensaje(2);
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
            }
        </script>
        """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    print(Cur)
def Evidencia_Guardar(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    Error = ""
    try:
        Caja = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
        Info_Actual = json.loads(str(Caja["cc_informacion_actual"]))
        Datos_Info = json.loads(str(Datos["Info"]))
        #del Info_Actual["Dock"]
        for k in Datos_Info.keys():
            Info_Actual[k] = Datos_Info[k]
        Error = DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Actual))+"',cc_ultimo_mov = NOW() WHERE cc_id = '"+str(Datos["ID"])+"' ")
        if Error == "":
            Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
            Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
            Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
            Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
            Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
            Error = DB.Instruccion(""" 
            INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
            (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
            VALUES
            ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','EVIDENCIA')
            """)
            if Error == "":
                Resultado["Estado"] = 1
            else:
                Resultado["Contenido"] = Error
        else:
            Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Ver_Historico(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Historico = []
        Usuarios = DB.Get_Dato("SELECT cusrid,\"Nombre\" FROM public.cuser")
        for H in DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas_moviemiento WHERE cch_master = '"+str(Datos["ID"])+"' ORDER BY cch_fecha_hora"):
            Info_Actual = json.loads(str(H["cch_informacion_actual"]))
            Usr_Aqui = None
            for U in Usuarios:
                if str(U["cusrid"]) == str(H["cch_usuario"]):
                    Usr_Aqui = U
                    break
            Tipo_Gen = "Not assigned" if H["cch_tipo_actual"] is None else H["cch_tipo_actual"]
            Usuario = str(H["cch_usuario"]) if Usr_Aqui is None else str(Usr_Aqui["Nombre"])
            Ubicacion = str(H["cch_ubicacion"]) if H["cch_dock"] is None else str(H["cch_ubicacion"]) + "["+str(H["cch_dock"])+"]"
            Archivos = ""
            Estado = ""
            Tipo = ""
            try:
                Tipo = Info_Actual[H["cch_tipo_actual"]]
                Estado += str(Tipo)
                Tipo_1 = Info_Actual[Tipo]
                Estado += "/"+str(Tipo_1)
                Tipo_2 = Info_Actual[Tipo_1]
                Estado += "/"+str(Tipo_2)
            except:
                pass

            if "Fotos" in Info_Actual.keys():
                for F in Info_Actual["Fotos"]:
                    Archivos += "<a href='http://10.4.7.219:8080/Portal_File/"+str(F)+"' target='_blank'><i class='mdi mdi-file-image'></i></a>"
            if "Archivos" in Info_Actual.keys():
                for F in Info_Actual["Archivos"]:
                    Archivos += "<a href='http://10.4.7.219:8080/Portal_File/"+str(F)+"' target='_blank'><i class='mdi mdi-file'></i></a>"
            
            Sello = ""
            if "Sello Proveedor" in Info_Actual.keys():
                Sello = "<span style='color:blue'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Proveedor"])
            if "Sello Temporal" in Info_Actual.keys():
                Sello = "<span style='color:#7a7a7a'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Temporal"])
            if "Sello Blanco" in Info_Actual.keys():
                Sello = "<span><i class='mdi mdi-label-outline'></i></span> "+str(Info_Actual["Sello Temporal"])
            if "Sello Rojo" in Info_Actual.keys():
                Sello = "<span style='color:#ff0000'><i class='mdi mdi-label'></i></span> "+str(Info_Actual["Sello Temporal"])

            Historico.append({"Fecha":H["cch_fecha_hora"].strftime("%Y-%m-%d %H:%M:%S"),"Usuario":str(Usuario),"Ubicacion":str(Ubicacion),"Carrier":Info_Actual["Carrier"],"Tipo":Tipo_Gen,"Estado":Estado,"Movimiento":str(H["cch_movimiento"]),"Archivos":Archivos,"Sello":Sello})
        Resultado["Contenido"] += """
        <div id='Tabla_Historico' class='border border-dark bg-dark-subtle'></div>
        <script>
            delete Tabla_Historico;
            var Tabla_Historico = new Tabulator("#Tabla_Historico", {
                minHeight:500,
                layout:"fitColumns",
                data:"""+str(Historico)+""",
                columns:[
                    {field:"Fecha","title":"Date"},
                    {field:"Usuario","title":"User"},
                    {field:"Ubicacion","title":"Location"},
                    {field:"Carrier","title":"Carrier"},
                    {field:"Tipo","title":"Type"},
                    {field:"Estado","title":"Status"},
                    {field:"Sello","title":"Seal",formatter:"html"},
                    {field:"Movimiento","title":"Step"},
                    {field:"Archivos","title":"Attached",formatter:"html"}
                ]
            });
            Tabla_Historico.on("tableBuilt", function(){ 
                setTimeout(() => {
                    Tabla_Historico.clearFilter();
                    Tabla_Historico.redraw();
                }, 100);
            });
        </script>
        """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Error_Ingreso(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Info_Nueva = json.loads(DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]["cc_informacion_actual"])
        Codigo = str(hashlib.md5(str("LOGISTICS INSIGHT CORPORATION - PASE SALIDA - "+str(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))+" - "+str(Datos["ID"])).encode()).hexdigest())
        Folio = ''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(8))
        Folio += str(Datos["ID"])
        Info_Nueva["Etapa"] = "Pending Exit"
        Error = DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Nueva))+"',cc_tipo_actual = 'Outbound', cc_qr_salida = '"+str(Codigo)+"' ,cc_folio_salida = '"+str(Folio)+"' ,cc_ultimo_mov = NOW(),cc_bloquear = 1 WHERE cc_id = '"+str(Datos["ID"])+"' ")
        if Error == "":
            Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
            Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
            Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
            Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
            Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
            Error = DB.Instruccion(""" 
            INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
            (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
            VALUES
            ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','ERROR DE INGRESO')
            """)
            if Error == "":
                Resultado["Estado"] = 1
            else:
                Resultado["Contenido"] = Error
        else:
            Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Imprimir_Pase(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = "content-type: text/html;charset=ISO-8859-1\n\n "
    Resultado = {"Contenido":"","Estado":0}
    try:
        Resultado["Pase Salida"] = Generar_Pase_Salida(Datos["ID"],"Pase_Salida_RILC_Toluca")
        Resultado["Estado"] = 1
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    print(Cur)
def Generar_Pase_Salida(ID_Caja,Nombre_Archivo):
    DB = LibDM_2023.DataBase()
    try:
        Error = ""
        CAJA = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(ID_Caja)+"'")[0]
        if CAJA["cc_qr_salida"] is not None:
            Codigo = str(CAJA["cc_qr_salida"])
            Folio = str(CAJA["cc_folio_salida"])
            Info_Cajas_Dtalle = json.loads(str(CAJA["cc_informacion_actual"]))
            buffered = BytesIO()
            img = qrcode.make(Codigo)
            f = open(str(os.path.dirname(os.path.realpath(__file__)))+"/"+str(Nombre_Archivo)+".png", "wb")
            img.save(f)
            f.close()
            wb_obj = openpyxl.load_workbook("C:/inetpub/wwwroot/Portal2/WTC/Plantilla_Pase_Salida.xlsx")
            sheet_obj = wb_obj.worksheets[0]

            Sello = ""
            try:
                K = None
                for Key in Info_Cajas_Dtalle.keys():
                    if "Sello" in str(Key):
                        K = str(Key)
                if K is not None:
                    Sello = str(Info_Cajas_Dtalle[K])
            except:
                pass

            sheet_obj.cell(row = 7, column = 3).value = str(CAJA["cc_contenedor"])[-4:]
            sheet_obj.cell(row = 11, column = 3).value = str(Sello)
            sheet_obj.cell(row = 15, column = 3).value = str(CAJA["cc_contenedor"])
            sheet_obj.cell(row = 15, column = 8).value = str(Info_Cajas_Dtalle["Carrier"])
            sheet_obj.cell(row = 16, column = 3).value = str(CAJA["cc_negocio"])
            sheet_obj.cell(row = 17, column = 3).value = str(CAJA["cc_ultimo_mov"])

            try:
                Ruta = ""
                Ruta = str(CAJA["cc_tipo_actual"])
                if CAJA["cc_tipo_actual"] in Info_Cajas_Dtalle.keys():
                    Ruta = Info_Cajas_Dtalle[CAJA["cc_tipo_actual"]]
                    if Info_Cajas_Dtalle[CAJA["cc_tipo_actual"]] in Info_Cajas_Dtalle.keys():
                        Ruta = Info_Cajas_Dtalle[Info_Cajas_Dtalle[CAJA["cc_tipo_actual"]]]
                Tipo = Tipo[""]
            except:
                pass
            sheet_obj.cell(row = 16, column = 8).value = str(Ruta)
            sheet_obj.cell(row = 36, column = 3).value = str(Folio)
            img = openpyxl.drawing.image.Image(str(os.path.dirname(os.path.realpath(__file__)))+'/'+str(Nombre_Archivo)+'.png')
            img.height = 310
            img.width = 310
            img.anchor = 'C20'
            sheet_obj.add_image(img)
            # sheet_obj.protection.sheet = True
            # sheet_obj.protection.password = 'LC_2022'
            # sheet_obj.protection.formatColumns = False
            # wb_obj.security  = WorkbookProtection(workbookPassword = 'LC_2022', lockStructure = True)
            wb_obj.save("C:/inetpub/wwwroot/Portal2/rilc_toluca/"+str(Nombre_Archivo)+".xlsx")
            return str(Nombre_Archivo)+".xlsx"
    except:
        return "ERROR: " + str(sys.exc_info())
def Generar_OSyD(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Caja = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
        Info_Actual = json.loads(str(Caja["cc_informacion_actual"]))
        Formulario = {"Col":"12", "Campos": [],"Clase": "OSyD" }
        if Datos["Tipo"] == "Otra Discrepancia":
            Datos["Tipo"] = ""
        Formulario["Campos"].append({"tipo":"texto","campo":"Comentario","titulo":"Comentarios","Requerido":1,"min":1,"max":150,"valor":Datos["Tipo"]})
        Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Fotos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += """
        <hr>
        <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Guardar_OSyD(\""""+str(Datos["Tipo"])+"""\","""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
        <script>
            function Guardar_OSyD(Tipo,ID){
                var Info = Dame_Formulario(".OSyD",true);
                if(Info != null)
                {
                    Info["Caja"] = '"""+str(Caja["cc_contenedor"])+"""';
                    Info["Carrier"] = '"""+str(Info_Actual["Carrier"])+"""';
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Guardar_OSyD".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID":ID,"Tipo":Tipo};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                $("#Vent_2").modal("hide");
                                Mensaje(2);
                                //Llamar_Funcion(\""""+str(request.url)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
            }
        </script>
        """
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Guardar_OSyD(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    Error = ""
    try:
        Info_Datos = json.loads(str(Datos["Info"]))
        Error = ""
        Error += DB.Instruccion("""
        INSERT INTO """+str(BD_Nombre)+""".cosyd
        (cosyd_tipo,cosyd_usuario,cosyd_alta,cosyd_caja,cosyd_comentario,cosyd_archivos,cosyd_scac)
        VALUES
        ('PRODUCCION','"""+str(Datos["ID_User"])+"""',NOW(),'"""+str(Info_Datos["Caja"])+"""','"""+str(Info_Datos["Comentario"])+"""','"""+str(','.join(Info_Datos["Fotos"]))+"""','"""+str(Info_Datos["Carrier"])+"""')
        """)
        if Error == "":
            ID = DB.Get_Dato("SELECT MAX(cosyd_id) as id FROM "+str(BD_Nombre)+".cosyd WHERE cosyd_tipo = 'PRODUCCION' AND cosyd_usuario = '"+str(Datos["ID_User"])+"' ")[0]["id"]
            Error += DB.Instruccion("""
            INSERT INTO """+str(BD_Nombre)+""".cosyd_historico
            (cosyd_master,cosyd_usuario,cosyd_comentario,cosyd_evidencia,cosyd_movimiento,cosyd_fecha)
            VALUES
            ('"""+str(ID)+"""','"""+str(Datos["ID_User"])+"""','"""+str(Info_Datos["Comentario"])+"""','"""+str(','.join(Info_Datos["Fotos"]))+"""','ALTA',NOW())
            """)
        if Error == "":
            Resultado["Estado"] = 1
        else:
            Resultado["Contenido"] = str(Error)
        Resultado["Contenido"] = str(Datos)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Foto_Lista_Salir(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    Error = ""
    try:
        Formulario = {"Col":"12", "Campos": [],"Clase": "Asignar" }
        Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Fotos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Temporal","titulo":"Sello Termporal","Requerido":1,"min":1,"max":150,"valor":""})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Salida(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    Error = ""
    try:
        Error += DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_activo = 0,cc_fecha_hora_salida = NOW(), cc_ultimo_mov = NOW() WHERE cc_id = '"+str(Datos["ID"])+"' ")
        if Error == "":
            Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
            Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
            Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
            Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
            Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
            Error += DB.Instruccion(""" 
            INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
            (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
            VALUES
            ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'CASETA','SALIDA')
            """)
        if Error == "":
            Resultado["Estado"] = 1

    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Nueva_Caja(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Formulario = {"Col":"12", "Campos": [],"Clase": "Alta_Caja" }
        Formulario["Campos"].append({"tipo":"texto","campo":"Container","titulo":"Contenedor","Requerido":1,"min":1,"max":150,"valor":""})
        Formulario["Campos"].append({"tipo":"seleccion","campo":"Carrier","titulo":"Carrier","Requerido":1,"Tipo_Opciones":"Query","Opciones":"SELECT cca_nombre as valor, cca_nombre as texto FROM "+str(BD_Nombre)+".ccarrier WHERE cca_activo = '1'","valor":"","Col":12})
        Formulario["Campos"].append({"tipo":"texto","campo":"Sello Proveedor","titulo":"Sello Proveedor","Requerido":1,"min":0,"max":20,"valor":""})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += "<div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Guardar_Caja()'><i class='mdi mdi-floppy'></i> Guardar</button></div>"
        Resultado["Contenido"] += """
        <script>
            Actualizar_Cambia_Texto();
            function Guardar_Caja(){
                var Info = Dame_Formulario(".Alta_Caja",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Guardar_Caja".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info)};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                $("#Vent_1").modal("hide");
                                Mensaje(2);
                                Llamar_Funcion(\""""+str(request.url_root)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
            }
        </script>
        """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Guardar_Caja(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Datos_Info = json.loads(str(Datos["Info"]))
        Existe = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_activo = 1 AND cc_contenedor = '"+str(Datos_Info["Container"])+"'")
        if len(Existe) == 0:
            Info = {}
            Info["Carrier"] = str(Datos_Info["Carrier"])
            Codigo = ''.join(random.SystemRandom().choice(string.ascii_uppercase + string.digits) for _ in range(8))
            if "Sello Proveedor" in Datos_Info.keys() and str(Datos_Info["Sello Proveedor"]).strip() != "":
                Info["Sello Proveedor"] = Datos_Info["Sello Proveedor"]
                Error = DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas
                (cc_contenedor,cc_ubicacion,cc_fecha_hora,cc_qr_entarda,cc_informacion_actual,cc_ultimo_mov)
                VALUES
                ('"""+str(Datos_Info["Container"])+"""','Patio',NOW(),'"""+str(Codigo)+"""','"""+str(json.dumps(Info))+"""',NOW())
                """)
            else:
                Info["Tipo"] = "Empty"
                Info["Empty"] = "Ready to Load"
                Error = DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas
                (cc_contenedor,cc_ubicacion,cc_fecha_hora,cc_qr_entarda,cc_informacion_actual,cc_ultimo_mov,cc_tipo_actual)
                VALUES
                ('"""+str(Datos_Info["Container"])+"""','Patio',NOW(),'"""+str(Codigo)+"""','"""+str(json.dumps(Info))+"""',NOW(),'Empty')
                """)
            if Error == "":
                IDMaster = DB.Get_Dato("SELECT MAX(cc_id) AS id FROM "+str(BD_Nombre)+".ccajas WHERE cc_contenedor = '"+str(Datos_Info["Container"])+"' AND cc_activo = 1 AND cc_ubicacion = 'Patio'")[0]["id"]
                Error = DB.Instruccion(""" 
                INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
                (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_usuario,cch_movimiento)
                VALUES
                ('"""+str(IDMaster)+"""',NOW(),'"""+str(Datos_Info["Container"])+"""','Patio','"""+str(json.dumps(Info))+"""','"""+str(Datos["ID_User"])+"""','ALTA')
                """)
                if Error == "":
                    Resultado["Estado"] = 1
                else:
                    Resultado["Contenido"] = Error
            else:
                Resultado["Contenido"] = Error
        else:
            Resultado["Contenido"] = "El Contenedor/Caja ["+str(Datos_Info["Container"])+"] ya existe!"
            
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Caja_Danada(Datos):
    if "K" in session.keys():
        fernet = Fernet(session["K"])
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    try:
        Formulario = {"Col":"12", "Campos": [],"Clase": "Caja_Danada" }
        Formulario["Campos"].append({"tipo":"archivo","campo":"Fotos","titulo":"Fotos","Requerido":1,"Col":12,"min":1,"max":5,"tipo_archivo":["image/*"],"valor":""})
        Resultado["Contenido"] += str(Compartido_2023.Formulario(Formulario))
        Resultado["Contenido"] += """
        <hr>
        <div class='w-100 text-center'><button class='btn btn-success w-75' onclick='Caja_Danada_Guardar("""+str(Datos["ID"])+""")'><i class='mdi mdi-floppy'></i> Save</button></div>
        <script>
            function Caja_Danada_Guardar(ID){
                var Info = Dame_Formulario(".Caja_Danada",true);
                if(Info != null)
                {
                    Mostrar_Ventana_Cargando(false);
                    var parametros = {"Fun":'"""+str(fernet.encrypt("Caja_Danada_Guardar".encode()).decode("utf-8"))+"""',"Info":JSON.stringify(Info),"ID":ID};
                    $.ajax({data:  parametros,url:\""""+str(request.url)+"""\",type:  "post",
                        success:  function (response)
                        {
                            var Resultado = JSON.parse(response);
                            if(Resultado["Estado"] == 1)
                            {
                                if("Pase Salida" in Resultado)
                                {
                                    var win1 = window.open(Resultado["Pase Salida"], '_blank');
                                    win1.focus();
                                }
                                $("#Vent_1").modal("hide");
                                Mensaje(2);
                                Llamar_Funcion(\""""+str(request.url)+"""\");
                            }
                            else
                                Mensaje(0,Resultado["Contenido"]);
                                
                        },
                        error: function (jqXHR, textStatus, errorThrown )
                        {
                            Mensaje(0,textStatus);
                        }
                    });
                }
            }
        </script>
        """
    except:
         Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur
def Caja_Danada_Guardar(Datos):
    DB = LibDM_2023.DataBase()
    Compartido_2023 = LibDM_2023.Compartido()
    Cur = ""
    Resultado = {"Contenido":"","Estado":0}
    Error = ""
    try:
        Caja = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
        Info_Actual = json.loads(str(Caja["cc_informacion_actual"]))
        Datos_Info = json.loads(str(Datos["Info"]))
        del Info_Actual["Dock"]
        for k in Datos_Info.keys():
            Info_Actual[k] = Datos_Info[k]
        Error = DB.Instruccion("UPDATE "+str(BD_Nombre)+".ccajas SET cc_informacion_actual = '"+str(json.dumps(Info_Actual))+"',cc_ultimo_mov = NOW(),cc_dock= null,cc_ubicacion = 'Patio',cc_caja_danada = 1 WHERE cc_id = '"+str(Datos["ID"])+"' ")
        if Error == "":
            Info_Ahora = DB.Get_Dato("SELECT * FROM "+str(BD_Nombre)+".ccajas WHERE cc_id = '"+str(Datos["ID"])+"'")[0]
            Info_Ahora["cc_dock"] = 'null' if Info_Ahora["cc_dock"] is None else "'"+str(Info_Ahora["cc_dock"])+"'"
            Info_Ahora["cc_tipo_actual"] = 'null' if Info_Ahora["cc_tipo_actual"] is None else "'"+str(Info_Ahora["cc_tipo_actual"])+"'"
            Info_Ahora["cc_zona"] = 'null' if Info_Ahora["cc_zona"] is None else "'"+str(Info_Ahora["cc_zona"])+"'"
            Info_Ahora["cc_negocio"] = 'null' if Info_Ahora["cc_negocio"] is None else "'"+str(Info_Ahora["cc_negocio"])+"'"
            Error = DB.Instruccion(""" 
            INSERT INTO """+str(BD_Nombre)+""".ccajas_moviemiento
            (cch_master,cch_fecha_hora,cch_contenedor,cch_ubicacion,cch_informacion_actual,cch_dock,cch_tipo_actual,cch_zona,cch_negocio,cch_usuario,cch_movimiento)
            VALUES
            ('"""+str(Info_Ahora["cc_id"])+"""',NOW(),'"""+str(Info_Ahora["cc_contenedor"])+"""','"""+str(Info_Ahora["cc_ubicacion"])+"""','"""+str(Info_Ahora["cc_informacion_actual"])+"""',"""+str(Info_Ahora["cc_dock"])+""","""+str(Info_Ahora["cc_tipo_actual"])+""","""+str(Info_Ahora["cc_zona"])+""","""+str(Info_Ahora["cc_negocio"])+""",'"""+str(Datos["ID_User"])+"""','CAJA DAÑADA')
            """)
            if Error == "":
                Resultado["Estado"] = 1
            else:
                Resultado["Contenido"] = Error
        else:
            Resultado["Contenido"] += str(Error)
    except:
        Resultado["Contenido"] = str(sys.exc_info())
    Cur += json.dumps(Resultado)
    return Cur

def Direccionar(Datos):
    try:
        if "K" in session.keys():
            fernet = Fernet(session["K"])
        else:
            fernet = Fernet(LibDM_2023.Compartido().Dame_K2())
        if Datos is None:
            return Inicio()
        else:
            if "Fun" in Datos:
                Funcion = fernet.decrypt(Datos["Fun"].encode()).decode()
                Par = {}
                for P in Datos.keys():
                    if str(P) != "Fun":
                        Par[str(P)] = Datos[str(P)]
                if "IDu" in session.keys():
                    Par["ID_User"] = session["IDu"]
                return globals()[Funcion](Par)
            else:
                return Inicio()
    except:
        return str(sys.exc_info())
